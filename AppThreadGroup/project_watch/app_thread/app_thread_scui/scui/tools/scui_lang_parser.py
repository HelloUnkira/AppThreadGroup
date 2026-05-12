import os
import sys
import json
import openpyxl
# -- coding: utf-8 --**


# for row in xlsx_sheet.rows: # 获取每一行的数据
#     for data in row:        # 获取每一行中单元格的数据
#         print(data.value)  # 打印单元格的值
# print(xlsx_sheet.cell(1, 1).value)


# 统计文件
def encode_scui_lang_parser_txt(file, xlsx_sheet, sheet_row, sheet_col):
    char_set = set()
    for item in xlsx_sheet.rows:
        for data in item:
            char_set = char_set | set(str(data.value))
            # print(set(str(data.value)))
    char_set = sorted(list(char_set))
    print("char_set:\n", char_set)
    for item in char_set:
        file.write(item)


# 编写集成化源文件
def encode_scui_lang_parser_bin(file, xlsx_sheet, sheet_row, sheet_col, args_list):
    for item in xlsx_sheet.iter_cols():
        for data in item:
            c_bytes = str(data.value).encode('utf-8')
            file.write(c_bytes)


# 编写集成化源文件
def encode_scui_lang_parser_c(file, xlsx_sheet, sheet_row, sheet_col, args_list):
    # 写点简要的说明
    file.write('/*本地静态的字符串表\n')
    file.write(' *通过scui_lang_parser.py生成\n */\n\n')
    file.write('#include "scui.h"\n\n')
    # 如果使用外源载入, 需要对应宏控制
    file.write('#if SCUI_LANG_PARSER_BIN_USE\n')
    # 提取所有外源依赖
    c_bytes_offset = 0
    c_bytes_length_max = 0
    for idx_col, item in enumerate(xlsx_sheet.iter_cols()):
        for idx_row, data in enumerate(item):
            c_bytes = str(data.value).encode('utf-8')
            c_bytes_len = len(c_bytes)
            c_bytes_prefix = 'static const scui_lang_item_t scui_lang_parser_item_'
            c_bytes_format = '{:03d}_{:03d} = {{.offset = 0x{:08X}, .length = 0x{:08X},}};\n'
            file.write(c_bytes_prefix + c_bytes_format.format(idx_col, idx_row, c_bytes_offset, c_bytes_len))
            c_bytes_offset += c_bytes_len
            if c_bytes_length_max <= c_bytes_len:
                c_bytes_length_max = c_bytes_len
    file.write('\n')
    file.write('const void * const scui_lang_parser_table[%d * %d] = {\n' % (sheet_row, sheet_col))
    for idx_col, item in enumerate(xlsx_sheet.iter_cols()):
        for idx_row, data in enumerate(item):
            file.write('\t(void *)&scui_lang_parser_item_%03d_%03d,\n' % (idx_col, idx_row))
    file.write('};\n')
    file.write('#else\n')
    # 提取所有外源依赖
    # 字符串太特殊了, 直接用原生表形式处理
    file.write('const void * const scui_lang_parser_table[%d * %d] = {\n' % (sheet_row, sheet_col))
    for idx_col, item in enumerate(xlsx_sheet.iter_cols()):
        for idx_row, data in enumerate(item):
            # 多国语这里不能很好的对齐,因为不同语言的空格宽度不一致
            c_array = str([hex(c) for c in str(data.value).encode('utf-8')])
            c_array = c_array.replace('\'', '')
            c_array = c_array.replace('[', '{')
            c_array = c_array.replace(']', ', 0}')
            c_array = c_array.replace('{', '(const char []){')
            file.write('\t//--->>>%s\n\t%s,\n' % (str(data.value).replace('\n', '\\n'), c_array))
    file.write('};\n')
    file.write('#endif\n')
    # 填充函数定义或者声明
    file.write('\nstatic scui_lang_type_t scui_lang_type = 0;\n\n')
    file.write('/*@brief 获取多国语语言类型\n')
    file.write(' *@param type 语言类型编号\n */\n\n')
    file.write('void scui_lang_get(scui_lang_type_t *type)\n{\n')
    file.write('\tSCUI_ASSERT(type != NULL);\n\t*type = scui_lang_type;\n}\n\n')
    # 填充函数定义或者声明
    file.write('/*@brief 设置多国语语言类型\n')
    file.write(' *@param type 语言类型编号\n */\n\n')
    file.write('void scui_lang_set(scui_lang_type_t *type)\n{\n')
    file.write('\tSCUI_ASSERT(type != NULL);\n\tscui_lang_type = *type;\n\t\n')
    file.write('\tscui_event_define(event, SCUI_HANDLE_SYSTEM, false, ')
    file.write('scui_event_lang_change, scui_event_absorb_none);\n')
    file.write('\tscui_event_notify(&event);\n}\n\n')
    # 填充函数定义或者声明
    file.write('/*@brief 多国语字符串转换\n')
    file.write(' *       需要同步拷贝使用\n')
    file.write(' *@param handle 字符串句柄\n')
    file.write(' *@param type   语言类型编号\n')
    file.write(' *@retval 字符串\n */\n')
    file.write('const char * scui_lang_str(scui_handle_t handle, scui_lang_type_t type)')
    file.write('\n{\n\tscui_handle_t string = handle;\n')
    file.write('\tswitch (type) {\n')
    file.write('\tdefault: string += type; break;\n')
    for idx, item in enumerate(args_list[1]):
        file.write('\tcase scui_lang_type_%s: string += scui_lang_type; break;\n' % item)
    file.write('\t}\n\tstring -= scui_lang_ofs_num;\n\t\n')
    file.write('\t#if SCUI_LANG_PARSER_BIN_USE\n')
    file.write('\tstatic uint8_t scui_lang_str_buffer[0x%X + 1] = {0};\n' % c_bytes_length_max)
    file.write('\tconst scui_lang_item_t *item_utf8 = scui_handle_source(string);\n')
    file.write('\tSCUI_ASSERT(item_utf8->length < 0x%X + 1);\n' % c_bytes_length_max);
    file.write('\tSCUI_ASSERT(item_utf8->offset + item_utf8->length < 0x%X);\n' % c_bytes_offset);
    file.write('\tscui_lang_src_read(scui_lang_str_buffer, item_utf8->offset, item_utf8->length);\n')
    file.write('\tscui_lang_str_buffer[item_utf8->length] = \'\\0\';\n')
    file.write('\tconst char *str_utf8 = scui_lang_str_buffer;\n')
    file.write('\t#else\n')
    file.write('\tconst char *str_utf8 = scui_handle_source(string);\n')
    file.write('\t#endif\n')
    file.write('\treturn str_utf8;\n')
    file.write('}\n\n\n')


# 编写集成化头文件
def encode_scui_lang_parser_h(file, xlsx_sheet, sheet_row, sheet_col, args_list):
    file.write('#ifndef SCUI_LANG_PARSER_H\n')
    file.write('#define SCUI_LANG_PARSER_H\n\n')
    # 写点简要的说明
    file.write('/*本地静态的字符串表\n')
    file.write(' *通过scui_lang_parser.py生成\n */\n\n')
    # 编写头部索引
    file.write('typedef enum {\n')
    file.write('\t%s = %s,\n' % (args_list[2], args_list[3]))
    for i, item in enumerate(xlsx_sheet.rows):
        c_str = str(item[0].value).replace('\n', '\\n')
        file.write('\tSCUI_LANG_0X%04x,\t/* %s */\n' % (i, c_str))
    file.write('} scui_lang_parser_str_t;\n\n')
    # 如果使用外源载入, 需要对应宏控制
    file.write('#if SCUI_LANG_PARSER_BIN_USE\n')
    file.write('typedef struct {\n')
    file.write('\tuintptr_t offset;\n')
    file.write('\tuintptr_t length;\n')
    file.write('} scui_lang_item_t;\n')
    file.write('#endif\n\n')
    file.write('extern const void * const scui_lang_parser_table[%d * %d];\n\n' % (sheet_row, sheet_col))
    # 编写头部索引
    file.write('typedef enum {\n')
    file.write('\tscui_lang_str_num = %s,\n' % str(sheet_row))
    file.write('\tscui_lang_ofs_num = 1,\n\t\n')
    file.write('\tscui_lang_type_num\t\t= %s,\n' % str(sheet_col))
    for idx, item in enumerate(args_list[1]):
        file.write('\tscui_lang_type_%s\t= SCUI_HANDLE_SYSTEM - %d,\n' % (item, idx + 1))
    for idx, item in enumerate(args_list[0]):
        file.write('\tscui_lang_type_%s\t\t= %s * %s + scui_lang_ofs_num,\n' % (item, 'scui_lang_str_num', str(idx)))
    file.write('} scui_lang_type_t;\n\n')
    # 填充函数定义或者声明
    file.write('/*@brief 获取多国语语言类型\n')
    file.write(' *@param type 语言类型编号\n */\n')
    file.write('void scui_lang_get(scui_lang_type_t *type);\n\n')
    # 填充函数定义或者声明
    file.write('/*@brief 设置多国语语言类型\n')
    file.write(' *@param type 语言类型编号\n */\n')
    file.write('void scui_lang_set(scui_lang_type_t *type);\n\n')
    # 填充函数定义或者声明
    file.write('/*@brief 多国语字符串转换\n')
    file.write(' *       需要同步拷贝使用\n')
    file.write(' *@param handle 字符串句柄\n')
    file.write(' *@param type   语言类型编号\n')
    file.write(' *@retval 字符串\n */\n')
    file.write('const char * scui_lang_str(scui_handle_t handle, scui_lang_type_t type);\n\n')
    file.write('#endif\n')


# 打印重定向到文件
class ScuiRedirectPrint(object):
    def __init__(self, stream=sys.stdout, path='.', file='log.txt'):
        if not os.path.exists(path):
            os.makedirs(path)
        self.log = open(os.path.join(path, file), mode='w+', encoding='utf-8')
        self.terminal = stream

    def write(self, message):
        self.log.write(message)
        self.terminal.write(message)

    def flush(self):
        pass


# 启用集成化资源总表生成
def encode_scui_lang_parser():
    # 参数列表:文件根目录,
    if len(sys.argv) != 3:
        print('argv list not match')
        return
    src_path = sys.argv[1]
    dst_path = sys.argv[2]
    # 获得文件处理根路径
    if not os.path.exists(src_path):
        print('src path is not exist')
        return
    if not os.path.exists(dst_path):
        print('dst path is not exist')
        return
    # print重定向
    sys.stdout = ScuiRedirectPrint(sys.stdout, file=os.path.join(dst_path, 'scui_lang_parser.out'))  # redirect print
    sys.stderr = ScuiRedirectPrint(sys.stderr, file=os.path.join(dst_path, 'scui_lang_parser.err'))  # redirect print
    print('src path:', src_path)
    print('dst path:', dst_path)
    # json转Python字符串并转标准字典
    json_file = open(os.path.join(src_path, 'scui_lang_parser.json'), 'r', encoding='utf-8')
    json_dict = json.loads(json_file.read())
    json_file.close()
    # 仅支持批量化字符串表管理
    if json_dict['type'] != 'scui lang parser':
        return
    # 从标准字典获取工作簿名字和单元
    xlsx_file = openpyxl.load_workbook(os.path.join(src_path, json_dict['xlsx']), read_only=False)
    xlsx_sheet = xlsx_file[json_dict['sheet']]
    sheet_row = xlsx_sheet.max_row
    sheet_col = xlsx_sheet.max_column
    # 获得多语言表,和句柄偏移
    args_list = [json_dict['language'], json_dict['custom'], json_dict['offset_name'], json_dict['offset_value']]
    if sheet_col != len(args_list[0]):
        print('args list not match')
        return
    # 开启四个文件
    scui_lang_parser_h = open(os.path.join(dst_path, 'scui_lang_parser.h'), mode='w', encoding='utf-8')
    scui_lang_parser_c = open(os.path.join(dst_path, 'scui_lang_parser.c'), mode='w', encoding='utf-8')
    scui_lang_parser_bin = open(os.path.join(dst_path, 'scui_lang_parser.bin'), mode='wb')
    scui_lang_parser_txt = open(os.path.join(dst_path, 'scui_lang_parser.txt'), mode='w', encoding='utf-8')
    # 解析
    encode_scui_lang_parser_h(scui_lang_parser_h, xlsx_sheet, sheet_row, sheet_col, args_list)
    encode_scui_lang_parser_c(scui_lang_parser_c, xlsx_sheet, sheet_row, sheet_col, args_list)
    encode_scui_lang_parser_bin(scui_lang_parser_bin, xlsx_sheet, sheet_row, sheet_col, args_list)
    encode_scui_lang_parser_txt(scui_lang_parser_txt, xlsx_sheet, sheet_row, sheet_col)
    # 关闭四个文件
    scui_lang_parser_h.close()
    scui_lang_parser_c.close()
    scui_lang_parser_bin.close()
    scui_lang_parser_txt.close()
    # 关闭工作簿
    xlsx_file.close()


if __name__ == '__main__':
    try:
        encode_scui_lang_parser()
    except Exception as e:
        print(e)
    print('scui lang parser finish')
    input('请按任意键退出...')
