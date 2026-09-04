# 资源打包统一前端(唯一入口, 仅 exe 启动)
# 位于 scui/tools/; 引用同目录下的 scui_pack_* 打包脚本(模块)
# 支持: widget / image / font / lang / cwf 五类
# 无参       -> 启动图形界面(五个子界面)
# 带参型     -> 命令行模式: python scui_pack_tools.py <type> [--src P] [--dst P] [--proj N]
# 说明: GUI 运行时把控制台 stdout/stderr 打印重定向输出到全局共享日志窗口;
#       运行产生的临时/缓存统一落在 __pack_tmp__(当前相对路径), 已被 .gitignore 忽略
import os
import re
import json
import sys
import importlib
import shutil
import queue
import threading

#============================================================
# 路径定位(基于 ui 路径推导, 兼容 dev 与 PyInstaller 单文件 exe)
#   ui 目录(scui_ui_res)的父级为 app_thread_scui
#   -> scui/tools(打包脚本)  scui/plugs(cwf 脚本)
#   注意: 不能用 __file__ 推导(exe 运行时指向临时解包目录)
#============================================================
def _dirs(ui):
    ui   = os.path.normpath(ui)
    app  = os.path.dirname(ui)                      # app_thread_scui
    scui = os.path.join(app, 'scui')
    tools= os.path.join(scui, 'tools')
    plugs= os.path.join(scui, 'plugs')
    return ui, app, scui, tools, plugs

def _ui_root(ui):
    return os.path.normpath(ui or os.getcwd())

# 日志路径显示: 相对 ui(scui_ui_res) 打印, 便于辨识(src/dst 均以其为基准)
def _rel_ui(p, ui):
    try:
        r = os.path.relpath(p, ui).replace('\\', '/')
        if r.startswith('..'):
            return p
        return r
    except Exception:
        return p

#============================================================
# 句柄偏移默认值(与各打包脚本头部常量一致; 弹窗修改后写入 scui_pack_handle.json)
#============================================================
_HANDLE_DEFAULT = {
    'widget': ('SCUI_HANDLE_OFFSET_WIDGET', '0x1000 - 1'),
    'image' : ('SCUI_HANDLE_OFFSET_IMAGE',  '0x2000 - 1'),
    'font'  : ('SCUI_HANDLE_OFFSET_FONT',   '0x4000 - 1'),
    'lang'  : ('SCUI_HANDLE_OFFSET_LANG',   '0x5000 - 1'),
}

# 图形配置默认(image 图形配置面板; 与 scui_pack_image.py 头部常量一致)
_IMAGE_CFG_DEFAULT = {
    'use_lz4':    True,
    'use_jpg':    True,
    'use_png':    True,
    'workers':    6,
    'alpha_bits': 4,
    'index_bits': 8,
    'endian':     False,
}
_IMG_CFG_DEFS = [      # (键, bool/int, 标签, 选项)
    ('use_lz4',    'bool', '启用 LZ4 压缩',   None),
    ('use_jpg',    'bool', '启用 JPG 打包',   None),
    ('use_png',    'bool', '启用 PNG 打包',   None),
    ('workers',    'combo', '并行线程数(1-6)', ('1', '2', '3', '4', '5', '6')),
    ('alpha_bits', 'combo', 'alpha 位宽(1/2/4/8)', ('1', '2', '4', '8')),
    ('index_bits', 'combo', 'index 位宽(1/2/4/8)', ('1', '2', '4', '8')),
    ('endian',     'bool', '大端字节序(整体)', None),
]

# 外源 lv_font_conv 命令模板(前置: cmd 中 lv_font_conv 可用)
_FONT_SYMBOL_RANGE = '61441,61448,61451,61452,61452,61453,61457,61459,61461,61465,61468,61473,61478,61479,61480,61502,61507,61512,61515,61516,61517,61521,61522,61523,61524,61543,61544,61550,61552,61553,61556,61559,61560,61561,61563,61587,61589,61636,61637,61639,61641,61664,61671,61674,61683,61724,61732,61787,61931,62016,62017,62018,62019,62020,62087,62099,62212,62189,62810,63426,63650'
_FONT_EU_RANGE  = '-r 0x00-0x7F -r 0x80-0xFF -r 0x100-0x17F -r 0x180-0x24F -r 0x2B0-0x2FF'
_FONT_CJK_RANGE = '-r 0x00-0x7F -r 0x3000-0x303F -r 0x3040-0x309F -r 0x30A0-0x30FF -r 0x4E00-0x9FFF'
_FONT_CMD_ORD   = 'lv_font_conv --font "{}" {} --size {} --format bin --bpp {} -o "{}" --force-fast-kern-format'
_FONT_CMD_DIY   = 'lv_font_conv --font "{}" {} --size {} --format bin --bpp {} -o "{}" --no-kerning'

#============================================================
# 任务定义
#============================================================
_TASK = {
    'widget': {'module': 'scui_pack_widget', 'entry': 'scui_widget_parser',
               'as_main': ['scui_pack_widget.py'], 'dir': 'tools'},
    'image' : {'module': 'scui_pack_image', 'entry': 'scui_image_parser',
               'as_main': ['scui_pack_image.py'], 'dir': 'tools'},
    'font'  : {'module': 'scui_pack_font', 'entry': 'scui_font_package',
               'as_main': ['scui_pack_font.py'], 'dir': 'tools'},
    'lang'  : {'module': 'scui_pack_lang', 'entry': 'encode_scui_lang_parser',
               'as_main': ['scui_pack_lang.py'], 'dir': 'tools'},
    'cwf'   : {'module': 'scui_pack_cwf', 'entry': 'scui_cwf_json_parser',
               'as_main': ['scui_pack_cwf.py'], 'dir': 'tools'},
}

# ui 默认子路径(src, dst)
_DEFAULT = {
    'widget': ('scene_src', 'scene_out'),
    'image' : ('image_src', 'image_out'),
    'font'  : ('font_bin',  'font_out'),
    'lang'  : ('lang_src',  'lang_out'),
    'cwf'   : ('cwf_json',  'cwf_json/bin'),
}

_IMG_EXT = ('.bmp', '.jpg', '.jpeg', '.png', '.gif', '.lottie.json', '.mp4')

#============================================================
# 脚本模块调用(注入 argv, 直接 import 复用主函数)
#============================================================
def _run_module(task, argv, ui):
    ui, app, scui, tools, plugs = _dirs(ui)
    if tools not in sys.path:
        sys.path.insert(0, tools)
    argv_old = sys.argv
    sys.argv = [task['as_main'][0]] + argv
    try:
        mod   = importlib.import_module(task['module'])
        mod.SCUI_UI_ROOT = ui     # 注入 ui(scui_ui_res) 基准, 后端日志路径相对化
        mod.SCUI_TOOLS  = tools   # 注入真实 tools 目录(读句柄偏移配置)
        if task['module'] == 'scui_pack_widget':
            mod.SCUI_WIDGET_TOOLS = tools     # 注入真实 tools 目录(打包环境下 __file__ 指向临时解包目录)
        entry = getattr(mod, task['entry'])
        entry()
        return 0
    except ImportError as e:
        print('[pack] 缺模块: %s' % e)
        print('[pack] 请确保依赖已安装(image需 pillow/lz4; image.7z解压需 py7zr)')
        return 1
    except Exception as e:
        print('[pack] 执行异常: %r' % e)
        return 1
    finally:
        sys.argv = argv_old
        shutil.rmtree(os.path.join(tools, '__pycache__'), ignore_errors=True)
        shutil.rmtree(os.path.join(plugs, '__pycache__'), ignore_errors=True)

def _image_src_unpack(src, tag):
    if os.path.isdir(src):
        return src, None
    if src.lower().endswith('.7z'):
        tmp = os.path.join(os.path.dirname(src), tag + '.src_tmp')
        try:
            import py7zr
            shutil.rmtree(tmp, ignore_errors=True)
            with py7zr.SevenZipFile(src, 'r') as z:
                z.extractall(tmp)
            return tmp, tmp
        except ImportError:
            print('[pack] 解压 image.7z 需要 py7zr: pip3 install py7zr')
            return None, None
    return None, None

def _do_task(name, ui, src, dst, proj):
    task   = _TASK[name]
    src    = os.path.normpath(src)
    dst    = os.path.normpath(dst)

    print()
    print('[pack] type : %s' % name)
    print('[pack] src  : scui_ui_res => %s' % _rel_ui(src, ui))
    print('[pack] dst  : scui_ui_res => %s' % _rel_ui(dst, ui))

    tmp = None
    if name == 'image':
        src, tmp = _image_src_unpack(src, 'image')
        if src is None:
            return 1
        argv = [src, dst, proj or 'prj']
    elif name == 'widget':
        argv = [src, dst, os.path.join(dst, 'scui_ui_maker.json')]
    elif name == 'cwf':
        return _do_cwf_task(ui, src, dst)
    else:
        argv = [src, dst]

    if not os.path.exists(src):
        print('[pack] src not exist: %s' % _rel_ui(src, ui))
        return 1
    if not os.path.isdir(dst):
        os.makedirs(dst, exist_ok=True)

    print('[pack] argv : %s' % [_rel_ui(a, ui) if os.path.isabs(a) else a for a in argv])
    ret = _run_module(task, argv, ui)
    if tmp:
        shutil.rmtree(tmp, ignore_errors=True)
    print()
    print('[pack] %s 执行完成: %s' % (name, 'OK' if ret == 0 else 'FAIL'))
    return ret

#============================================================
# cwf 多步打包(解压7z -> image parser -> cwf parser -> 清理)
#============================================================
def _do_cwf_task(ui, src, dst):
    ui_dir, app, scui, tools, plugs = _dirs(ui)
    img_task = _TASK['image']
    cwf_task = _TASK['cwf']
    if not os.path.isdir(src):
        print('[pack] cwf src not exist: %s' % _rel_ui(src, ui))
        return 1
    if not os.path.isdir(dst):
        os.makedirs(dst, exist_ok=True)

    # 收集所有 cwf 目录(有 image.7z + 同名 .json)
    cwf_list = []
    for name in sorted(os.listdir(src)):
        d = os.path.join(src, name)
        if not os.path.isdir(d):
            continue
        if os.path.exists(os.path.join(d, 'image.7z')) and \
           os.path.exists(os.path.join(d, name + '.json')):
            cwf_list.append(name)

    if not cwf_list:
        print('[pack] cwf: 未找到有效的 cwf 目录(需含 image.7z + 同名 .json)')
        return 1

    print('[pack] cwf list: %s' % ', '.join(cwf_list))
    print()

    # 确保工具路径在 sys.path
    if tools not in sys.path:
        sys.path.insert(0, tools)

    failed = []
    for wf in cwf_list:
        wf_dir = os.path.join(src, wf)
        print('=========== [%s] start ============' % wf)
        # step1: 解压 image.7z -> cwf 目录(7z 内含 image/ 子目录, 不能再嵌套)
        img_dir = os.path.join(wf_dir, 'image')
        shutil.rmtree(img_dir, ignore_errors=True)
        try:
            import py7zr
            with py7zr.SevenZipFile(os.path.join(wf_dir, 'image.7z'), 'r') as z:
                z.extractall(wf_dir)                    # 解压到 cwf 目录, 7z 自带 image/
            print('[%s] unzip image.7z -> OK' % wf)
        except Exception as e:
            print('[%s] unzip fail: %r' % (wf, e))
            failed.append(wf)
            continue

        # step2: image parser (image/ -> .)
        print('[%s] step1 image parser...' % wf)
        argv_old = sys.argv
        sys.argv = [img_task['as_main'][0], img_dir, wf_dir, 'cwf']
        try:
            mod = importlib.import_module(img_task['module'])
            mod.SCUI_UI_ROOT = ui     # 注入 ui(scui_ui_res) 基准, 后端日志路径相对化
            mod.SCUI_TOOLS  = tools   # 注入真实 tools 目录(读句柄偏移配置)
            mod.scui_image_parser()
        except Exception as e:
            print('[%s] image parser fail: %r' % (wf, e))
            failed.append(wf)
        finally:
            sys.argv = argv_old

        # step3: cwf json parser (. -> bin/)
        print('[%s] step2 cwf json parser...' % wf)
        sys.argv = [cwf_task['as_main'][0], wf_dir, dst]
        try:
            mod = importlib.import_module(cwf_task['module'])
            mod.SCUI_UI_ROOT = ui     # 注入 ui(scui_ui_res) 基准, 后端日志路径相对化
            mod.SCUI_TOOLS  = tools   # 注入真实 tools 目录(读句柄偏移配置)
            mod.SCUI_CWF_TOOLS = tools    # 注入真实 tools 目录
            mod.SCUI_CWF_PLUGS = plugs    # 注入真实 plugs 目录(写 scui_cwf_json_proto.h)
            mod.scui_cwf_json_parser()
        except Exception as e:
            print('[%s] cwf parser fail: %r' % (wf, e))
            failed.append(wf)
        finally:
            sys.argv = argv_old

        # step4: 清理临时文件
        shutil.rmtree(img_dir, ignore_errors=True)
        shutil.rmtree(os.path.join(wf_dir, 'image_array'), ignore_errors=True)
        for f in ('scui_image_parser.h', 'scui_image_parser.c',
                  'scui_image_parser.bin'):
            p = os.path.join(wf_dir, f)
            if os.path.exists(p):
                os.remove(p)
        # cwf parser 临时目录(统一 __pack_tmp__/, 相对 CWD, 见 .gitignore)
        shutil.rmtree('__pack_tmp__', ignore_errors=True)
        prog = os.path.join(dst, wf + '_json.prog')
        if os.path.exists(prog):
            os.remove(prog)
        # 清理 __pycache__
        shutil.rmtree(os.path.join(tools, '__pycache__'), ignore_errors=True)
        shutil.rmtree(os.path.join(plugs, '__pycache__'), ignore_errors=True)
        print('=========== [%s] finish ============' % wf)
        print()

    if failed:
        print('[pack] cwf FAILED: %s' % ', '.join(failed))
        return 1
    print('[pack] cwf 执行完成: OK')
    return 0

#============================================================
# 图像信息探测(对齐 scui_image_t 结构体字段)
#============================================================
# 与后端一致的无尾缀句柄: 去扩展名/去.idx/.dit标记/去点, 保留路径层级
def _img_handle_tag(file, src):
    rel = os.path.relpath(file, src)
    d, base = os.path.split(rel)
    stem, _ext = os.path.splitext(base)
    low = stem.lower()
    if low.endswith('.dit'):
        stem = stem[:-4]
    elif low.endswith('.idx'):
        stem = stem[:-4]
    clean = (d + '/' + stem) if d else stem
    return ('prj_' + clean).replace('.', '').replace('\\', '_').replace('/', '_').replace(' ', '_')


def _img_type_of(path):
    lo = path.lower()
    if lo.endswith('.gif'):           return 'scui_image_type_gif'
    if lo.endswith('.lottie.json'):   return 'scui_image_type_lottie'
    if lo.endswith('.mp4'):           return 'scui_image_type_mp4'
    if lo.endswith(('.jpg', '.jpeg')):return 'scui_image_type_jpg'
    if lo.endswith('.png'):           return 'scui_image_type_png'
    return 'scui_image_type_bmp'

# 图片定制标记解析/组装(# 前缀排除 + .dit/.idx 尾缀, 缀于扩展名前)
def _img_mkbar_split(name):
    n0, ext = os.path.splitext(name)
    exc  = n0.startswith('#')
    stem = n0[1:] if exc else n0
    low  = stem.lower()
    idx  = low.endswith('.idx')
    dit  = low.endswith('.dit')
    if idx: stem = stem[:-4]
    if dit: stem = stem[:-4]
    return stem, ext, exc, idx, dit

def _img_mkbar_join(stem, ext, exc, idx, dit):
    out = stem
    if idx: out += '.idx'
    elif dit: out += '.dit'
    if exc: out = '#' + out
    return out + ext


def _image_probe(path):
    info = {'path': path, 'name': os.path.basename(path)}
    info['type'] = _img_type_of(path)
    # size
    w = h = 0
    try:
        import PIL.Image as _img
        im = _img.open(path); w, h = im.size; im.close()
    except Exception:
        pass
    info['pixel.width']  = w
    info['pixel.height'] = h
    info['pixel.size_bin'] = os.path.getsize(path)
    # format & 每像素内存字节
    fmt = 'scui_pixel_cf_none'; pb = 0
    if info['type'] == 'scui_image_type_bmp': fmt, pb = 'scui_pixel_cf_bmp565', 2
    elif info['type'] == 'scui_image_type_jpg': fmt, pb = 'scui_pixel_cf_bmp565', 2
    elif info['type'] == 'scui_image_type_png': fmt, pb = 'scui_pixel_cf_bmp8565', 3
    info['pixel.format'] = fmt
    info['pixel.size_mem'] = w * h * pb
    info['com_pct'] = (float(info['pixel.size_bin']) / info['pixel.size_mem']) if info['pixel.size_mem'] else 1.0
    return info

#============================================================
# GUI
#============================================================
def _launch_gui(ui):
    try:
        import tkinter as tk
    except Exception as e:
        print('[pack] 无图形环境, 请使用命令行模式: %s' % e)
        return 1
    app = PackApp(ui)
    app.root.mainloop()
    return 0

class PackApp(object):
    def __init__(self, ui):
        import tkinter as tk
        from tkinter import ttk
        self.ui  = ui
        self.ui, self.app, self.scui, self.tools, self.plugs = _dirs(ui)
        self._imgs = []          # 保持 PhotoImage 引用
        self._pv  = {}           # (name,side) -> StringVar
        self._current_name = 'image'
        self._logq = None
        self._log_win = None    # 全局共享日志窗口(Toplevel)
        self._log_text = ''     # 日志内存缓冲(所有子界面共用)

        self.in_abs = {}; self.out_abs = {}
        for k, (si, so) in _DEFAULT.items():
            self.in_abs[k]  = os.path.join(ui, si)
            self.out_abs[k] = os.path.join(ui, so)

        self.root = tk.Tk()
        self.root.title('scui资源工具')
        self.root.geometry('980x720')
        self.root.minsize(860, 620)

        menubar = tk.Menu(self.root)
        m_set = tk.Menu(menubar, tearoff=0)
        m_set.add_command(label='路径…', command=self._open_paths)
        m_set.add_command(label='句柄偏移…', command=self._open_handles)
        menubar.add_cascade(label='设置', menu=m_set)
        m_log = tk.Menu(menubar, tearoff=0)
        m_log.add_command(label='打开 out/err 浏览…', command=self._open_logs)
        menubar.add_cascade(label='日志', menu=m_log)
        m_help = tk.Menu(menubar, tearoff=0)
        m_help.add_command(label='关于', command=self._about)
        menubar.add_cascade(label='帮助', menu=m_help)
        self.root.config(menu=menubar)

        self._build_header()
        self._build_notebook()

        self.root.protocol('WM_DELETE_WINDOW', self.root.destroy)
        self.root.after(120, self._poll)

    #--------------- 顶部信息条(只读, 无按钮) ---------------
    def _build_header(self):
        from tkinter import ttk
        bar = ttk.Frame(self.root, padding=(12, 6))
        bar.pack(fill='x')
        ttk.Label(bar, text='scui_ui: %s' % self.ui, foreground='#777').pack(side='left')
        self.path_status = ttk.Label(bar, text='', foreground='#888')
        self.path_status.pack(side='left', padx=(14, 0))

    #--------------- 路径设置子弹窗(显示所有路径, 前两项只读, 其余可改) ---------------
    def _open_paths(self):
        import tkinter as tk
        from tkinter import ttk, messagebox
        try:
            self._open_paths_dialog()
        except Exception as e:
            messagebox.showwarning('路径设置', '打开路径设置失败:\n%r' % e, parent=self.root)

    def _open_paths_dialog(self):
        import tkinter as tk
        from tkinter import ttk, filedialog
        top = tk.Toplevel(self.root)
        top.title('路径设置')
        top.geometry('620x600')
        top.transient(self.root)
        f = ttk.Frame(top, padding=(12, 10)); f.pack(fill='both', expand=True)
        ttk.Label(f, text='scui 路径: %s' % self.scui, font=('Consolas', 9)).pack(anchor='w')
        ttk.Label(f, text='scui_ui 路径: %s' % self.ui, font=('Consolas', 9)).pack(anchor='w', pady=(2, 8))

        grid = ttk.Frame(f); grid.pack(fill='x')
        varman = {}
        def _browse(entryvar):
            abs_p = filedialog.askdirectory(title='选择路径', initialdir=entryvar.get() or self.ui)
            if abs_p:
                entryvar.set(self._rel(abs_p))     # 显示为相对
        r = 0
        for name in ('widget', 'image', 'font', 'lang', 'cwf'):
            if name == 'font':                          # font 输入拆: font 源输入(font_src)
                fv = tk.StringVar(value=self._rel(self.font_src_dir))
                ttk.Label(grid, text='font 源输入:', width=14, anchor='e')\
                    .grid(row=r, column=0, padx=(0, 8), pady=3)
                ttk.Entry(grid, textvariable=fv, font=('Consolas', 9))\
                    .grid(row=r, column=1, sticky='ew', pady=3)
                ttk.Button(grid, text='浏览…', width=7, command=lambda v=fv: _browse(v))\
                    .grid(row=r, column=2, padx=(8, 0))
                varman[('in', 'font_src')] = fv
                r += 1
            for side, lab in (('in', '输入'), ('out', '输出')):
                cur = self.in_abs[name] if side == 'in' else self.out_abs[name]
                labt = 'font bin 输入:' if (name == 'font' and side == 'in') else '%s %s:' % (name, lab)
                ttk.Label(grid, text=labt, width=14, anchor='e')\
                    .grid(row=r, column=0, padx=(0, 8), pady=3)
                var = tk.StringVar(value=self._rel(cur))   # 子类型显示相对路径
                e = ttk.Entry(grid, textvariable=var, font=('Consolas', 9))
                e.grid(row=r, column=1, sticky='ew', pady=3)
                ttk.Button(grid, text='浏览…', width=7, command=lambda v=var: _browse(v))\
                    .grid(row=r, column=2, padx=(8, 0))
                varman[(side, name)] = var
                r += 1
        grid.columnconfigure(1, weight=1)

        note = ttk.Label(f, text='以下为相对 scui_ui 路径, 可修改；scui / scui_ui 为绝对路径且固定。修改即时生效。',
                         foreground='#888').pack(anchor='w', pady=(8, 4))
        ttk.Button(f, text='关闭', command=top.destroy).pack(side='right')
        for (side, name), var in varman.items():
            var.trace_add('write', lambda *a, s=side, n=name, v=var: self._apply_path(s, n, v))
        top.protocol('WM_DELETE_WINDOW', top.destroy)

    #--------------- 句柄偏移设置子弹窗(widget/image/font/lang 各一组: 偏移名/偏移值) ---------------
    def _open_handles(self):
        import tkinter as tk
        from tkinter import ttk, messagebox
        try:
            self._open_handles_dialog()
        except Exception as e:
            messagebox.showwarning('句柄偏移设置', '打开句柄偏移设置失败:\n%r' % e, parent=self.root)

    def _handle_json(self):
        return os.path.join(self.tools, 'scui_pack_handle.json')

    def _open_handles_dialog(self):
        import tkinter as tk
        from tkinter import ttk, messagebox
        top = tk.Toplevel(self.root)
        top.title('句柄偏移设置')
        top.geometry('580x340')
        top.transient(self.root)
        f = ttk.Frame(top, padding=(12, 10)); f.pack(fill='both', expand=True)
        # 读取已保存配置(无则用默认值)
        cur = {}
        cfg = self._handle_json()
        if os.path.isfile(cfg):
            try:
                cur = json.load(open(cfg, encoding='utf-8'))
            except Exception:
                cur = {}
        vars = {}
        grid = ttk.Frame(f); grid.pack(fill='x')
        ttk.Label(grid, text='子类型', width=10, anchor='w').grid(row=0, column=0, padx=(0, 8), pady=3)
        ttk.Label(grid, text='偏移名(name)', width=30, anchor='w').grid(row=0, column=1, padx=(0, 8))
        ttk.Label(grid, text='偏移值(value)', width=16, anchor='w').grid(row=0, column=2)
        for r, name in enumerate(('widget', 'image', 'font', 'lang'), start=1):
            dname, dvalue = _HANDLE_DEFAULT[name]
            item = cur.get(name, {})
            nv = tk.StringVar(value=item.get('name') or dname)
            vv = tk.StringVar(value=item.get('value') or dvalue)
            ttk.Label(grid, text=name, width=10, anchor='w').grid(row=r, column=0, padx=(0, 8), pady=4)
            ttk.Entry(grid, textvariable=nv, font=('Consolas', 9)).grid(row=r, column=1, sticky='ew', padx=(0, 8))
            ttk.Entry(grid, textvariable=vv, font=('Consolas', 9)).grid(row=r, column=2, sticky='ew')
            vars[name] = (nv, vv)
        grid.columnconfigure(1, weight=2)
        grid.columnconfigure(2, weight=1)

        def _save():
            data = {}
            for name, (nv, vv) in vars.items():
                data[name] = {'name': nv.get().strip(), 'value': vv.get().strip()}
            try:
                with open(cfg, 'w', encoding='utf-8') as fp:
                    json.dump(data, fp, ensure_ascii=False, indent=4)
                messagebox.showinfo('句柄偏移设置', '已保存: %s' % self._rel_log(cfg), parent=top)
            except Exception as e:
                messagebox.showwarning('句柄偏移设置', '保存失败:\n%r' % e, parent=top)

        def _reset():
            for name, (nv, vv) in vars.items():
                dname, dvalue = _HANDLE_DEFAULT[name]
                nv.set(dname); vv.set(dvalue)

        bar = ttk.Frame(f); bar.pack(fill='x', pady=(10, 0))
        ttk.Button(bar, text='保存', command=_save).pack(side='left')
        ttk.Button(bar, text='恢复默认', command=_reset).pack(side='left', padx=(8, 0))
        ttk.Button(bar, text='关闭', command=top.destroy).pack(side='right')
        ttk.Label(bar, text='打包时写入生成的头文件, 作为各子类型句柄枚举起始偏移.',
                  foreground='#888').pack(side='left', padx=(12, 0))
        top.protocol('WM_DELETE_WINDOW', top.destroy)

    def _apply_path(self, side, name, var):
        val = var.get().strip()
        if not val:
            return
        # 相对路径 -> 转绝对(相对 scui_ui)
        if not os.path.isabs(val):
            val = os.path.normpath(os.path.join(self.ui, val))
        if name == 'font_src':
            self.font_src_dir = val                      # font 源输入(独立于 in_abs)
            if hasattr(self, '_font_src_refresh'):
                self._font_src_refresh()
            return
        if side == 'in':
            self.in_abs[name] = val
        else:
            self.out_abs[name] = val
        if hasattr(self, 'path_status'):
            self.path_status.config(text='%s %s: %s' % (name, '输入' if side == 'in' else '输出', self._rel(val)))
        pv = self._pv.get((name, side))
        if pv is not None:
            pv.set('%s: %s' % ('输入' if side == 'in' else '输出', self._rel(val)))

    def _rel(self, p):
        try:
            return os.path.relpath(p, self.ui).replace('\\', '/')
        except Exception:
            return p

    def _rel_log(self, p):
        # 日志路径显示: scui_ui_res => 相对路径
        return 'scui_ui_res => %s' % self._rel(p)

    #--------------- 四个子界面(tab) ---------------
    def _build_notebook(self):
        import tkinter as tk
        from tkinter import ttk
        self.nb = ttk.Notebook(self.root)
        self.nb.pack(fill='both', expand=True, padx=12, pady=(0, 12))
        self.tabs = {}
        self._build_widget_tab()         # widget: 完整(左文件/中册子/右编辑)
        self._build_image_tab()          # image: 完整
        self._build_font_tab()           # font: 完整
        self._build_lang_tab()           # lang: 预览为主
        self._build_cwf_tab()            # cwf: 参考 widget
        self.nb.bind('<<NotebookTabChanged>>', self._on_tab)

    #--------------- widget 子界面(左: src文件 | 中: analyze字段册 | 右: 编辑副本+log) ---------------
    def _build_widget_tab(self):
        import tkinter as tk
        from tkinter import ttk, scrolledtext
        f = ttk.Frame(self.nb, padding=8)
        self.tabs['widget'] = f
        # 编辑状态(副本): path/分支/条目索引
        self.widget_path  = None        # 当前编辑文件(绝对)
        self.widget_data  = None        # json 副本对象
        self.widget_maker = False       # False=scene; True=maker(默认配置)
        self.widget_edit  = None        # 当前条目索引
        self._w_base     = None         # 加载时的基快照(用于未保存修改检测)

        # 顶部工具条: 操作按钮(左) + 执行打包(右) + 状态(左)
        top = ttk.Frame(f); top.pack(fill='x', pady=(2, 4))
        ttk.Button(top, text='新建场景', command=self._wid_new_scene).pack(side='left')
        ttk.Button(top, text='加载 maker json', command=self._wid_load_maker).pack(side='left', padx=(6, 0))
        ttk.Button(top, text='预览 json', command=self._wid_preview_json).pack(side='left', padx=(6, 0))
        ttk.Button(top, text='保存 json', command=self._wid_save).pack(side='left', padx=(6, 0))
        ttk.Button(top, text='执行 widget 打包', command=lambda: self._run('widget')).pack(side='right')
        self.widget_status = ttk.Label(top, text='未选择文件', foreground='#777')
        self.widget_status.pack(side='left', padx=(12, 0))

        # 主体: 左(scene) | 中(上 analyze 下 模板) | 右(编辑)
        hpan = ttk.Panedwindow(f, orient='horizontal'); hpan.pack(fill='both', expand=True, pady=(6, 0))

        # 左: scene 文件树(虚拟化 Canvas, 只绘可见行, 支持折叠)
        lf = ttk.LabelFrame(hpan, text=' scene布局集合 ', padding=(4, 4)); hpan.add(lf, weight=3)
        self.wcanv_tree = tk.Canvas(lf, highlightthickness=0, bg='#ffffff')
        self._wt_vsb = ttk.Scrollbar(lf, orient='vertical', command=self._wid_scroll_cmd)
        self.wcanv_tree.configure(yscrollcommand=self._wt_vsb.set)
        self.wcanv_tree.pack(side='left', fill='both', expand=True)
        self._wt_vsb.pack(side='right', fill='y')
        self.wcanv_tree.bind('<Configure>', lambda e: self._wid_draw())
        self.wcanv_tree.bind('<Enter>', lambda e: self.wcanv_tree.bind_all('<MouseWheel>', self._wid_tree_wheel))
        self.wcanv_tree.bind('<Leave>', lambda e: self.wcanv_tree.unbind_all('<MouseWheel>'))
        self.wcanv_tree.bind('<Button-1>', self._wid_tree_click)
        self._wdisp  = []              # 折叠后可见行([depth,name,path,kind,opened])
        self._wrowH  = 24              # 每行像素高
        self._wt_tree = []             # 缓存目录节点树
        self._wt_exp  = set()          # 已展开的目录

        # 中: 上 analyze 下 模板(垂直分栏)
        mf = ttk.Frame(hpan); hpan.add(mf, weight=3)
        vpan = ttk.Panedwindow(mf, orient='vertical'); vpan.pack(fill='both', expand=True)
        # 上: analyze 可配置字段册
        ab = ttk.LabelFrame(vpan, text=' analyze 可配置字段 ', padding=(4, 4)); vpan.add(ab, weight=3)
        self.wbook = ttk.Treeview(ab, columns=('slot',), show='tree headings', selectmode='browse')
        self.wbook.heading('#0', text='字段路径'); self.wbook.heading('slot', text='值槽')
        self.wbook.column('slot', width=56, anchor='e', stretch=False)
        bvs = ttk.Scrollbar(ab, orient='vertical', command=self.wbook.yview)
        self.wbook.configure(yscrollcommand=bvs.set)
        self.wbook.pack(side='left', fill='both', expand=True); bvs.pack(side='right', fill='y')
        self.wbook.bind('<<TreeviewSelect>>', self._wid_book_copy)
        # 下: 模板面板(T 按钮触发渲染)
        tb = ttk.LabelFrame(vpan, text=' 模板(T) ', padding=(4, 4)); vpan.add(tb, weight=2)
        self._tpl_canv = tk.Canvas(tb, highlightthickness=0)
        tplvs = ttk.Scrollbar(tb, orient='vertical', command=self._tpl_canv.yview)
        self._tpl_canv.configure(yscrollcommand=tplvs.set)
        self._tpl_box = ttk.Frame(self._tpl_canv)
        self._tpl_box_id = self._tpl_canv.create_window((0, 0), window=self._tpl_box, anchor='nw')
        self._tpl_box.bind('<Configure>',
            lambda e: self._tpl_canv.configure(scrollregion=self._tpl_canv.bbox('all')))
        self._tpl_canv.bind('<Configure>',
            lambda e: (self._tpl_canv.itemconfigure(self._tpl_box_id, width=e.width),
                       self._tpl_canv.configure(scrollregion=self._tpl_canv.bbox('all'))))
        self._tpl_canv.pack(side='left', fill='both', expand=True)
        tplvs.pack(side='right', fill='y')
        self._tpl_canv.bind('<Enter>', lambda e: self._tpl_canv.bind_all('<MouseWheel>', self._wid_tpl_wheel))
        self._tpl_canv.bind('<Leave>', lambda e: self._tpl_canv.unbind_all('<MouseWheel>'))
        self._tpl_hint = ttk.Label(self._tpl_box, text='点击 T 查看可用模板',
                                   foreground='#888', anchor='w')
        self._tpl_hint.pack(fill='x', padx=(2, 2), pady=(4, 2))
        self._w_templates = {}
        self._w_tpl_state  = {}

        # 右: 预览/编辑(json: 键|值两列就地编辑; c: 只读文本)
        rf = ttk.LabelFrame(hpan, text=' 预览 / 编辑(json 就地, .c 只读) ', padding=(4, 4)); hpan.add(rf, weight=4)
        self.widget_sub = tk.StringVar(value='json: 每行 key|value 直接修改; 标题行左侧按钮操作整个控件; 行尾按钮操作所在字段; 点 .c 只读')
        ttk.Label(rf, textvariable=self.widget_sub, foreground='#777').pack(anchor='w', padx=(2, 2))
        # 键值编辑(json): 滚动行列表, 每行 key|value 常驻输入框(自由编辑)
        self.wcanv = tk.Canvas(rf, highlightthickness=0)
        wsb = ttk.Scrollbar(rf, orient='vertical', command=self.wcanv.yview)
        self.wcanv.configure(yscrollcommand=wsb.set)
        self._wbox = ttk.Frame(self.wcanv)
        self._wbox_id = self.wcanv.create_window((0, 0), window=self._wbox, anchor='nw')
        self._wbox.bind('<Configure>', lambda e: self.wcanv.configure(scrollregion=self.wcanv.bbox('all')))
        self.wcanv.bind('<Configure>', lambda e: (self.wcanv.itemconfigure(self._wbox_id, width=e.width),
                                                  self.wcanv.configure(scrollregion=self.wcanv.bbox('all'))))
        self.wcanv.pack(side='left', fill='both', expand=True)
        wsb.pack(side='right', fill='y')
        self.wcanv.bind('<Enter>', lambda e: self.wcanv.bind_all('<MouseWheel>', self._wid_wheel))
        self.wcanv.bind('<Leave>', lambda e: self.wcanv.unbind_all('<MouseWheel>'))
        self._w_rows = []
        self._w_vars = []          # 保存StringVar引用防GC
        # 文本预览(c): 只读
        self.wctext = scrolledtext.ScrolledText(rf, wrap='char', font=('Consolas', 9), state='disabled')
        self._wid_switch(True)

        self.nb.add(f, text='widget  ')
        self._wid_analyze()
        self._wid_load_tree()
        self._wid_maker_guard()
        self._wid_load_maker(_silent=True)   # 打开默认加载 maker json

    # 右侧主体切换: True=json 行编辑; False=c 只读文本
    def _wid_switch(self, edit):
        if edit:
            self.wcanv.pack(side='left', fill='both', expand=True)
            self.wctext.pack_forget()
        else:
            self.wcanv.pack_forget()
            self.wctext.pack(fill='both', expand=True)

    # maker mode 守卫: 控件级 +-上下 按钮在渲染时按 widget_maker 条件省略(maker 满足唯一完备, 不可增减/移动控件)
    def _wid_maker_guard(self):
        pass

    # canvas 鼠标滚轮
    def _wid_wheel(self, e):
        self.wcanv.yview_scroll(-1 * (e.delta // 120), 'units')

    # 按钮回调统一安全包裹: 异常写入日志窗口(否则 --noconsole 静默吞掉, 界面直接空白)
    def _wid_guard(self, fn, *a):
        try:
            fn(*a)
        except Exception as e:
            import traceback
            msg = '[widget] 操作异常: %r\n%s' % (e, traceback.format_exc())
            self._wlog(msg)
            self.widget_status.config(text='操作异常, 详见日志窗口')

    def _wid_analyze(self):
        # 生成左窗格字段册(analyze 结果; tools 加入 sys.path 以便 import)
        tools = self.tools
        if tools not in sys.path:
            sys.path.insert(0, tools)
        try:
            import scui_pack_widget as pw
            book = pw.scui_widget_analyze_result(tools)
        except Exception as e:
            self._wlog('analyze 失败: %r' % e)
            return
        self.wbook.delete(*self.wbook.get_children())
        # analyze 全量控件类型(scui_xxx_maker_t), 供 maker json 加载时补齐缺失 class
        self._wid_class_types = list(book.get('class_maker', {}).keys())
        for wtype, maker in book.get('class_maker', {}).items():
            prefix = maker
            if prefix.startswith('scui_'):
                prefix = prefix[5:]
            if prefix.endswith('_maker_t'):
                prefix = prefix[:-len('_maker_t')]
            slots = book.get('path_slots', {}).get(maker, {})
            top = self.wbook.insert('', 'end', text=prefix, open=False, values=('',))
            # 按字段路径逐点分层折叠(widget.style.buffer -> widget/style/buffer), 默认折叠
            node_items = {}
            for path, slot in slots.items():
                segs = path.split('.')
                d = node_items
                for s in segs[:-1]:
                    n = d.get(s)
                    if n is None:
                        n = {'slot': None, 'kids': {}}
                        d[s] = n
                    d = n['kids']
                leaf = d.get(segs[-1])
                if leaf is None:
                    d[segs[-1]] = {'slot': slot, 'kids': {}}
                else:
                    leaf['slot'] = slot
            def _rend(parent, items):
                for text, node in items.items():
                    if node['kids']:
                        nid = self.wbook.insert(parent, 'end', text=text, open=False, values=('',))
                        _rend(nid, node['kids'])
                    else:
                        self.wbook.insert(parent, 'end', text=text, values=(node['slot'],))
            _rend(top, node_items)

    # 左窗格点选字段 -> 复制路径到剪贴板(粘贴到右侧 key 用)
    def _wid_book_copy(self, _ev=None):
        sel = self.wbook.selection()
        if not sel:
            return
        path = self.wbook.item(sel[0], 'text')
        if not path:
            return
        try:
            self.root.clipboard_clear()
            self.root.clipboard_append(path)
        except Exception as e:
            return
        self.widget_status.config(text='已复制字段: %s(可粘贴到右侧 key)' % path)

    def _wlog(self, s):
        self._append_log('widget', s + '\n')

    # 左: scene 文件树(虚拟化 Canvas + 折叠)
    def _wid_load_tree(self):
        self.wcanv_tree.delete('all')
        old_exp = getattr(self, '_wt_exp', set())
        root = self.in_abs['widget']
        if not os.path.isdir(root):
            root = self.ui
        self._wt_root = root
        self._wt_tree = self._wid_tree_scan(root)
        # 保留仍存在的展开目录(删除/新建后折叠状态不丢)
        self._wt_exp = {p for p in old_exp if os.path.isdir(p)}
        self._wid_reflatten()
        self._wid_load_templates()       # 文件树刷新即刷新模板缓存

    def _wid_tree_scan(self, path):
        """brief: 扫描目录成嵌套节点
        @param path 目录绝对路径
        @retval 节点列表[{name,path,kind,children}]
        """
        nodes = []
        try:
            items = sorted(os.listdir(path),
                key=lambda x: (not os.path.isdir(os.path.join(path, x)), x.lower()))
        except OSError:
            return nodes
        for e in items:
            full = os.path.join(path, e)
            if os.path.isdir(full):
                nodes.append({'name': e, 'path': full, 'kind': 'dir',
                              'children': self._wid_tree_scan(full)})
            elif e.lower().endswith('.json'):
                nodes.append({'name': e, 'path': full, 'kind': 'json', 'children': []})
            elif e.lower().endswith('.c'):
                nodes.append({'name': e, 'path': full, 'kind': 'c', 'children': []})
        return nodes

    def _wid_toggle_dir(self, path):
        if path in self._wt_exp:
            self._wt_exp.discard(path)
        else:
            self._wt_exp.add(path)
        self._wid_reflatten()

    def _wid_reflatten(self):
        # 保存当前可见首行索引, 展开/折叠后保持滚动位置
        old_n = len(self._wdisp)
        try:
            v0, _ = self.wcanv_tree.yview()
        except Exception:
            v0 = 0.0
        idx_top = int(v0 * old_n) if old_n else 0
        disp = []
        self._wid_tree_flatten(self._wt_tree, 0, disp)   # 最外围根路径不显示
        self._wdisp = disp
        n = len(disp)
        self.wcanv_tree.configure(scrollregion=(0, 0, 0, n * self._wrowH))
        if n:
            # 恢复滚动位置(按首行索引, 超界 clamp; 不满一屏自动归顶)
            self.wcanv_tree.yview_moveto(min(1.0, idx_top / n))
        self._wid_draw()

    def _wid_tree_flatten(self, nodes, depth, disp):
        for nd in nodes:
            opened = (nd['kind'] == 'dir' and nd['path'] in self._wt_exp)
            disp.append([depth, nd['name'], nd['path'], nd['kind'], opened])
            if opened:
                self._wid_tree_flatten(nd['children'], depth + 1, disp)

    # 滚动条 command: 先滚动后重绘(虚拟化无持久 item, 必须重绘)
    def _wid_scroll_cmd(self, *args):
        self.wcanv_tree.yview(*args)
        self._wid_draw()

    def _wid_tree_wheel(self, e):
        self.wcanv_tree.yview_scroll(-1 * (e.delta // 120), 'units')
        self._wid_draw()

    def _wid_draw(self):
        c = self.wcanv_tree
        # 滚动条可见性 + 内容不足一屏顶对齐不可滚动
        vh = c.winfo_height() or 200
        self._wid_needs_scroll = (len(self._wdisp) * self._wrowH > vh)
        if self._wid_needs_scroll:
            if not self._wt_vsb.winfo_ismapped():
                self._wt_vsb.pack(side='right', fill='y')
        else:
            if self._wt_vsb.winfo_ismapped():
                self._wt_vsb.pack_forget()
            c.yview_moveto(0)
        c.delete('all')
        n = len(self._wdisp)
        if n == 0:
            return
        v0, v1 = c.yview()
        i0 = max(0, int(v0 * n) - 2)
        i1 = min(n, int(v1 * n) + 2)
        cw = max(c.winfo_width(), 120)
        for idx in range(i0, i1):
            depth, name, path, kind, opened = self._wdisp[idx]
            y = idx * self._wrowH
            x = depth * 14
            # 左"-"柱: json 可点删除; dir/c 灰占位
            if kind == 'json':
                c.create_text(x + 7, y + self._wrowH // 2, text='−',
                              font=('Segoe UI Symbol', 10), fill='#333')
            else:
                c.create_text(x + 7, y + self._wrowH // 2, text='−',
                              font=('Segoe UI Symbol', 10), fill='#ccc')
            if kind == 'dir':
                c.create_text(x + 20, y + self._wrowH // 2,
                              text='▼' if opened else '▶', font=('Segoe UI Symbol', 8),
                              fill='#555')
            c.create_text(x + 36, y + self._wrowH // 2,
                          text='📁' if kind == 'dir' else '📄', font=('Segoe UI Emoji', 9))
            fg = '#2266cc' if kind == 'json' else ('#666' if kind == 'c' else '#333')
            c.create_text(x + 54, y + self._wrowH // 2, text=name,
                          font=('Consolas', 9, 'bold') if kind == 'dir' else ('Consolas', 9),
                          fill=fg, anchor='w')
            if kind == 'dir':                        # 右侧小"+" 新建场景
                c.create_rectangle(cw - 30, y + 3, cw - 2, y + self._wrowH - 3,
                                   outline='#cfcfcf', fill='#eeeeee')
                c.create_text(cw - 16, y + self._wrowH // 2, text='+',
                              font=('Segoe UI Symbol', 9))

    def _wid_tree_click(self, e):
        cy = self.wcanv_tree.canvasy(e.y)
        idx = int(cy // self._wrowH)
        if not (0 <= idx < len(self._wdisp)):
            return
        depth, name, path, kind, opened = self._wdisp[idx]
        cx = self.wcanv_tree.canvasx(e.x)
        cw = self.wcanv_tree.winfo_width() or 400
        x = depth * 14
        if cx >= cw - 34:                             # 右按钮区: 目录新建场景
            if kind == 'dir':
                self._wid_guard(self._wid_tree_new, path)
            return
        if x <= cx < x + 16:                          # 左"-"柱: json 删除
            if kind == 'json':
                self._wid_guard(self._wid_tree_del, path)
            return
        if kind == 'dir':                             # 主区: 目录折叠
            self._wid_toggle_dir(path)
        else:                                          # 文件: 打开
            self._wid_pick_file_path(path)

    # 点选文件加载(.json 可编辑, .c 只读)
    def _wid_pick_file_path(self, path):
        """brief: 点选文件加载
        @param path 文件绝对路径
        """
        if not (os.path.isfile(path) and path.lower().endswith(('.json', '.c'))):
            return
        if not self._wid_confirm_discard():      # 切走前, 未保存修改需确认丢弃
            return
        self.widget_edit = None
        if path.lower().endswith('.c'):
            self.widget_path = path; self.widget_data = None; self.widget_maker = False
            self.widget_status.config(text='%s (只读)' % os.path.basename(path))
            self.widget_sub.set('%s: 文件文本预览(只读, 不可修改)' % os.path.basename(path))
            self.wctext.configure(state='normal')
            self.wctext.delete('1.0', 'end')
            try:
                with open(path, 'r', encoding='utf-8', errors='replace') as fp:
                    self.wctext.insert('1.0', fp.read())
            except Exception as e:
                self.wctext.insert('1.0', '读取失败: %r' % e)
            self.wctext.configure(state='disabled')
            self._wid_switch(False)
            return
        try:
            with open(path, 'r', encoding='utf-8') as fp:
                self.widget_data = json.load(fp)
            self.widget_path = path; self.widget_maker = False
            self._w_base = json.loads(json.dumps(self.widget_data))   # 记录基快照
            self.widget_status.config(text='已加载: %s' % os.path.basename(path))
            self.widget_sub.set('%s: 每行 key/value 直接修改, 光标所在可增删' % os.path.basename(path))
            self._wid_switch(True)
            self._wid_maker_guard()
            self._wid_render_json(_top=True)
        except Exception as e:
            self.widget_status.config(text='加载失败: %r' % e)

    # 新建空场景: 从 json 起步, 生成 scui_ui_<name>.json + scui_ui_<name>.c 模板
    def _wid_new_scene(self):
        import tkinter as tk
        from tkinter import ttk, messagebox
        root = self.in_abs['widget']
        cats = []
        if os.path.isdir(root):
            cats = sorted(d for d in os.listdir(root) if os.path.isdir(os.path.join(root, d)))
        top = tk.Toplevel(self.root)
        top.title('新建空场景')
        top.geometry('440x210')
        top.transient(self.root)
        f = ttk.Frame(top, padding=(12, 10)); f.pack(fill='both', expand=True)
        ttk.Label(f, text='分类目录(scene 下)').pack(anchor='w')
        catvar = tk.StringVar(value=cats[0] if cats else '')
        ttk.Combobox(f, textvariable=catvar, values=cats, width=40).pack(fill='x', pady=(2, 8))
        ttk.Label(f, text='场景名(小写蛇形, 如 test_ui)').pack(anchor='w')
        namevar = tk.StringVar()
        ttk.Entry(f, textvariable=namevar, width=40).pack(fill='x', pady=(2, 8))
        note = ttk.Label(f, text='将生成 .json(空 window) 与 .c(事件模板), 作为新场景起点.',
                         foreground='#888')
        note.pack(anchor='w')
        bar = ttk.Frame(f); bar.pack(fill='x', pady=(10, 0))
        ttk.Button(bar, text='创建', command=lambda: self._do_new_scene(top, catvar, namevar)).pack(side='left')
        ttk.Button(bar, text='取消', command=top.destroy).pack(side='right')
        namevar.trace_add('write', lambda *a: note.after(1, lambda: self._wid_new_preview(note, catvar, namevar, root)))
        top.protocol('WM_DELETE_WINDOW', top.destroy)

    def _wid_new_preview(self, note, catvar, namevar, root):
        name = namevar.get().strip() if namevar.get() else ''
        if not re.match(r'^[A-Za-z0-9_]+$', name):
            note.configure(text='将生成 .json(空 window) 与 .c(事件模板), 作为新场景起点.')
            return
        cat = catvar.get().strip()
        base = os.path.join(root, cat, 'scui_ui_%s' % name.lower()) if cat else os.path.join(root, 'scui_ui_%s' % name.lower())
        note.configure(text='将生成: %s\n           %s' % (self._rel(base + '.json'), self._rel(base + '.c')))

    def _do_new_scene(self, top, catvar, namevar, dir_override=None):
        """brief: 生成场景模板文件
        @param top          对话框
        @param catvar       分类(可 None)
        @param namevar      场景名
        @param dir_override 目标目录(优先)
        """
        from tkinter import messagebox
        name = namevar.get().strip()
        if not re.match(r'^[A-Za-z0-9_]+$', name):
            messagebox.showwarning('新建场景', '场景名仅允许字母/数字/下划线', parent=top)
            return
        root = self.in_abs['widget']
        if dir_override is not None:
            cat_dir = dir_override
        else:
            cat = catvar.get().strip() if catvar else ''
            cat_dir = os.path.join(root, cat) if cat else root
        os.makedirs(cat_dir, exist_ok=True)
        name_l = name.lower()
        name_u = name.upper()
        jpath = os.path.join(cat_dir, 'scui_ui_%s.json' % name_l)
        cpath = os.path.join(cat_dir, 'scui_ui_%s.c' % name_l)
        if os.path.exists(jpath) or os.path.exists(cpath):
            if not messagebox.askyesno('新建场景', '已存在同名场景, 覆盖写?\n%s' % self._rel_log(jpath), parent=top):
                return
        scene = {
            'type': 'scene',
            'widget': [{
                'widget.type':              'scui_widget_type_window',
                'widget.style.sched_anima': 'true',
                'widget.style.indev_ptr':   'true',
                'widget.style.indev_key':   'true',
                'widget.style.indev_enc':   'true',
                'widget.clip.x':            '0',
                'widget.clip.y':            '0',
                'widget.clip.w':            '466',
                'widget.clip.h':            '466',
                'widget.myself':            'SCUI_UI_SCENE_' + name_u,
                'widget.event_cb':          'scui_ui_scene_%s_event_proc' % name_l,
                'widget.child_num':         '0',
                'widget.format':            'SCUI_PIXEL_CF_DEF_A',
            }],
        }
        crlf = lambda s: s.replace('\n', '\r\n')
        with open(jpath, 'w', encoding='utf-8', newline='') as fp:
            fp.write(crlf(json.dumps(scene, ensure_ascii=False, indent=4)))
        c_tmpl = (
            '/*实现目标:\n'
            ' *    窗口:%s\n'
            ' */\n'
            '\n'
            '#define SCUI_LOG_LOCAL_STATUS       1\n'
            '#define SCUI_LOG_LOCAL_LEVEL        2   /* 0:DEBUG,1:INFO,2:WARN,3:ERROR,4:NONE */\n'
            '\n'
            '#include "scui.h"\n'
            '\n'
            '/*@brief 控件事件响应回调\n'
            ' *@param event 事件\n'
            ' */\n'
            'void scui_ui_scene_%s_event_proc(scui_event_t *event)\n'
            '{\n'
            '    switch (event->type) {\n'
            '    case scui_event_create:\n'
            '        break;\n'
            '    case scui_event_destroy:\n'
            '        break;\n'
            '    case scui_event_focus_get:\n'
            '        break;\n'
            '    case scui_event_focus_lost:\n'
            '        break;\n'
            '    default:\n'
            '        break;\n'
            '    }\n'
            '}\n') % (name_l, name_l)
        with open(cpath, 'w', encoding='utf-8', newline='') as fp:
            fp.write(crlf(c_tmpl))
        top.destroy()
        self._wid_load_tree()                      # 刷新 src 树
        self.widget_data = scene
        self.widget_path = jpath; self.widget_maker = False
        self._w_base = json.loads(json.dumps(scene))
        self.widget_status.config(text='已新建: %s' % os.path.basename(jpath))
        self.widget_sub.set('%s: 每行 key/value 直接修改, 光标所在可增删' % os.path.basename(jpath))
        self._wid_switch(True)
        self._wid_maker_guard()
        self._wid_render_json(_top=True)
        self._append_log('widget', '[new] 已新建场景: %s\n' % self._rel_log(jpath))

    # 目录 + 按钮: 该目录下新建场景(轻量对话框, 仅输入名称)
    def _wid_tree_new(self, dir_path):
        """brief: 目录内新建场景
        @param dir_path 目标目录
        """
        import tkinter as tk
        from tkinter import ttk
        top = tk.Toplevel(self.root)
        top.title('新建场景于: %s' % os.path.basename(dir_path))
        top.geometry('420x180'); top.transient(self.root)
        f = ttk.Frame(top, padding=(12, 10)); f.pack(fill='both', expand=True)
        ttk.Label(f, text='场景名(小写蛇形, 如 test_ui)').pack(anchor='w')
        namevar = tk.StringVar()
        ttk.Entry(f, textvariable=namevar, width=40).pack(fill='x', pady=(2, 8))
        ttk.Label(f, text='将生成 .json(空 window) 与 .c(事件模板)',
                  foreground='#888').pack(anchor='w')
        bar = ttk.Frame(f); bar.pack(fill='x', pady=(10, 0))
        ttk.Button(bar, text='创建',
            command=lambda: self._do_new_scene(top, None, namevar,
                                               dir_override=dir_path))\
            .pack(side='left')
        ttk.Button(bar, text='取消', command=top.destroy).pack(side='right')
        top.protocol('WM_DELETE_WINDOW', top.destroy)

    # json − 按钮: 删除 .json + 同名 .c
    def _wid_tree_del(self, json_path):
        """brief: 删除场景 json+同名 c
        @param json_path json 绝对路径
        """
        from tkinter import messagebox
        if not os.path.isfile(json_path):
            return
        c_path = json_path[:-5] + '.c'              # .json -> .c
        files = [json_path]
        if os.path.isfile(c_path):
            files.append(c_path)
        msg_lines = '\n'.join(self._rel(p) for p in files)
        if not messagebox.askyesno('删除场景',
                '将删除:\n%s\n确认?' % msg_lines, parent=self.root):
            return
        if self.widget_path in files:               # 当前编辑文件正属删除范围, 先清空
            self.widget_path  = None
            self.widget_data  = None
            self.widget_maker = False
            self._w_base      = None
            self.widget_status.config(text='未选择文件')
            for w in self._wbox.winfo_children():
                w.destroy()
        for p in files:
            try:
                os.remove(p)
                self._wlog('[del] 已删除: %s' % self._rel_log(p))
            except OSError as e:
                self._wlog('[del] 失败 %s: %r' % (self._rel_log(p), e))
        self._wid_load_tree()                       # 刷新文件树 + 模板缓存

    # 平铺渲染: 每个条目 = 标题Label + 每字段一行 [key输入框][value], 之间空行
    # _top=True 强制滚动回顶部(加载/新建); 否则保留当前可视滚动位置(局部增删字段不跳动)
    def _wid_render_json(self, _top=False):
        import tkinter as tk
        from tkinter import ttk
        try:
            frac = 0.0 if _top else self.wcanv.yview()[0]
        except Exception:
            frac = 0.0
        for w in self._wbox.winfo_children():
            w.destroy()
        self._w_rows = []
        self._w_vars = []          # 保存StringVar引用防GC(否则Entry内容被清空)
        self._w_titles = {}          # 条目索引 -> 标题 Label(供整条高亮)
        self._w_entry_rows = {}      # 条目索引 -> 该条目所有行 Frame
        if self.widget_data is None:
            self.wcanv.yview_moveto(0)
            return
        items = self.widget_data.get('widget', [])
        for i, it in enumerate(items):
            if self.widget_maker:
                title = '#%d  %s' % (i, it.get('class', '?'))
                fields = it.setdefault('default', {})
            else:
                title = '#%d  %s' % (i, it.get('widget.myself') or it.get('widget.type', ''))
                fields = it
            head = tk.Frame(self._wbox, bg='#ffffff')
            head.pack(fill='x', pady=(6, 2))
            # 右侧: +(加字段, 最右) -> T(查模板, 紧贴其左)
            ttk.Button(head, text='+', width=1, command=lambda i=i: self._wid_guard(self._wid_add_field, i))\
                .pack(side='right', padx=(2, 0))
            ttk.Button(head, text='T', width=1, command=lambda i=i: self._wid_guard(self._wid_show_templates, i))\
                .pack(side='right', padx=(0, 1))
            if not self.widget_maker:               # 控件级小按钮(标题左侧, 仅 scene): +加/ -删/ ▲▼移动
                for txt, cb in (('+', lambda i=i: self._wid_guard(self._wid_add_ctl, i)),
                                ('-', lambda i=i: self._wid_guard(self._wid_del_ctl, i)),
                                ('▲', lambda i=i: self._wid_guard(self._wid_ctl_up, i)),
                                ('▼', lambda i=i: self._wid_guard(self._wid_ctl_dn, i))):
                    ttk.Button(head, text=txt, width=1, command=cb).pack(side='left', padx=(0, 1))
            tl = tk.Label(head, text=title, foreground='#2266cc', background='#ffffff',
                          font=('Consolas', 9, 'bold'), anchor='w')
            tl.pack(side='left', fill='x', expand=True, padx=(4, 0))
            tl.bind('<Button-1>', lambda e, i=i: self._wid_focus(i, None, None))
            head.bind('<Button-1>', lambda e, i=i: self._wid_focus(i, None, None))
            self._w_titles[i] = tl
            self._w_entry_rows[i] = []
            for k, v in fields.items():
                kvar = tk.StringVar(value=str(k))
                vvar = tk.StringVar(value=str(v))
                self._w_vars.extend([kvar, vvar])    # 保存引用防GC
                row = tk.Frame(self._wbox, bg='#ffffff'); row.pack(fill='x', padx=(2, 2))
                # maker下widget.type灰显锁定(与class绑定, 不可改不可删)
                locked = self.widget_maker and k == 'widget.type'
                st = 'disabled' if locked else 'normal'
                # 右侧按钮先pack(从右取位), 确保不被key/val挤压
                del_btn = ttk.Button(row, text='-', width=1, command=lambda i=i, k=k: self._wid_guard(self._wid_del_field_k, i, k))
                del_btn.pack(side='right', padx=(1, 0))
                ttk.Button(row, text='▼', width=1, command=lambda i=i, k=k: self._wid_guard(self._wid_field_move, i, k, 1))\
                    .pack(side='right', padx=(1, 0))
                ttk.Button(row, text='▲', width=1, command=lambda i=i, k=k: self._wid_guard(self._wid_field_move, i, k, -1))\
                    .pack(side='right', padx=(1, 0))
                if locked:
                    del_btn.configure(state='disabled')
                ke = ttk.Entry(row, textvariable=kvar, font=('Consolas', 9), state=st)
                ve = ttk.Entry(row, textvariable=vvar, font=('Consolas', 9), state=st)
                ke.pack(side='left', fill='x', expand=True, padx=(0, 8))
                ve.pack(side='left', fill='x', expand=True)
                ke.bind('<FocusIn>', lambda e, i=i, k=k, row=row: self._wid_focus(i, k, row))
                ve.bind('<FocusIn>', lambda e, i=i, k=k, row=row: self._wid_focus(i, k, row))
                row.bind('<Button-1>', lambda e, i=i, k=k, row=row: self._wid_focus(i, k, row))
                self._w_rows.append([i, k, ke, ve, row])      # [条目, 字段key, key框, value框, 行框]
                self._w_entry_rows[i].append(row)
            ttk.Label(self._wbox, text='').pack()
        self.wcanv.yview_moveto(frac)                 # 恢复滚动位置(加载新文件已置 0)

    # 光标定位 + 高亮: 点字段行高亮该行; 点标题整条高亮(空条目也可选中). 状态栏提示正在编辑对象
    def _wid_focus(self, i, key, row):
        self._w_active = (i, key)
        # 清除全部高亮(字段行 + 标题)
        for rows in getattr(self, '_w_entry_rows', {}).values():
            for _row in rows:
                if _row is not None:
                    _row.configure(bg='#ffffff')
        for _t in getattr(self, '_w_titles', {}).values():
            if _t is not None:
                _t.configure(bg='#ffffff')
        if row is not None:                          # 命中的具体字段行
            row.configure(bg='#fff3a3')
        elif i is not None:                          # 点标题: 整条高亮(含空条目)
            tl = getattr(self, '_w_titles', {}).get(i)
            if tl is not None:
                tl.configure(bg='#fff3a3')
            for _row in getattr(self, '_w_entry_rows', {}).get(i, []):
                if _row is not None:
                    _row.configure(bg='#fff3a3')
        # 状态栏: 提示当前操作对象
        if i is not None and self.widget_data and 0 <= i < len(self.widget_data.get('widget', [])):
            if key is not None:
                self.widget_status.config(text='正在编辑 条目#%d · 字段 %s' % (i, key))
            elif getattr(self, '_w_entry_rows', {}).get(i):
                self.widget_status.config(text='正在编辑 条目#%d(整条)' % i)
            else:
                self.widget_status.config(text='正在编辑 条目#%d(空条目, 可「添加字段」)' % i)

    # 取某条目的字段 dict
    def _wid_fields_of(self, item):
        it = self.widget_data['widget'][item]
        if self.widget_maker:
            return it.setdefault('default', {})
        return it

    # 把当前输入框文本回填到副本(改 key 重建保序; 空 key 行保留, 预览/保存时统一去重)
    def _wid_sync_back(self):
        if not self.widget_data:
            return
        groups = {}
        for i, k, ke, ve, row in self._w_rows:
            groups.setdefault(i, []).append((ke.get(), ve.get()))
        for i, pairs in groups.items():
            if i >= len(self.widget_data['widget']):
                continue
            d = {}
            for k, v in pairs:
                d[k] = v
            if self.widget_maker:
                self.widget_data['widget'][i]['default'] = d
            else:
                self.widget_data['widget'][i] = d

    # 递归删除空 key 占位行, 返回清理后的新结构(不改动当前副本; 预览/保存时去重)
    def _wid_clean_copy(self, obj):
        if isinstance(obj, dict):
            return {k: self._wid_clean_copy(v) for k, v in obj.items() if k != ''}
        if isinstance(obj, list):
            return [self._wid_clean_copy(x) for x in obj]
        return obj

    # 写某条目字段 dict
    def _wid_set_fields(self, it, newd):
        if self.widget_maker:
            self.widget_data['widget'][it]['default'] = newd
        else:
            self.widget_data['widget'][it] = newd

    # 新控件默认字段
    def _wid_new_control(self):
        b = {'widget.type': 'scui_widget_type_window', 'widget.clip.w': 'SCUI_HOR_RES',
             'widget.clip.h': 'SCUI_VER_RES', 'widget.myself': 'SCUI_UI_SCENE_'}
        return {'class': 'scui_widget_type_window', 'default': b} if self.widget_maker else b

    # 未保存修改检测(先把输入框回填再与基快照比对)
    def _wid_dirty(self):
        if self.widget_data is None or self._w_base is None:
            return False
        self._wid_sync_back()
        a = json.dumps(self.widget_data, ensure_ascii=False, sort_keys=True)
        b = json.dumps(self._w_base, ensure_ascii=False, sort_keys=True)
        return a != b

    # 切换前确认丢弃修改(是->继续; 否->取消本次切换)
    def _wid_confirm_discard(self):
        if not self._wid_dirty():
            return True
        from tkinter import messagebox
        return messagebox.askyesno('未保存的修改', '当前 json 有未保存的修改，确定丢弃吗？', parent=self.root)

    # 控件级 + : 当前控件 i 下面加一个新控件, 焦点到达新控件
    def _wid_add_ctl(self, i):
        from tkinter import messagebox
        if not (self.widget_data is not None and self.widget_path):
            messagebox.showwarning('widget', '请先在中间选择一个 .json 文件')
            return
        if self.widget_maker:
            return
        self._wid_sync_back()
        items = self.widget_data.setdefault('widget', [])
        items.insert(i + 1, self._wid_new_control())
        self._wid_render_json()
        self._wid_focus(i + 1, None, None)
        self._wlog('已添加控件(默认字段, 未保存)')

    # 控件级 - : 移除当前控件, 焦点到达附近(先上后下)
    def _wid_del_ctl(self, i):
        from tkinter import messagebox
        if not (self.widget_data is not None and self.widget_path):
            messagebox.showwarning('widget', '请先在中间选择一个 .json 文件')
            return
        if self.widget_maker:
            return
        self._wid_sync_back()
        items = self.widget_data.get('widget', [])
        if i >= len(items):
            return
        del items[i]
        self._wid_render_json()
        if items:                                   # 聚焦附近: 优先上面 i-1, 否则下面(原 i 处)
            tgt = (i - 1) if i > 0 else 0
            self._wid_focus(tgt, None, None)
        else:
            self._w_active = None
        self._wlog('已删除控件(未保存)')

    # 控件上移 / 下移
    def _wid_ctl_move(self, i, delta):
        if self.widget_maker:
            return
        self._wid_sync_back()
        items = self.widget_data.get('widget', [])
        j = i + delta
        if i < 0 or i >= len(items) or j < 0 or j >= len(items):
            return
        items[i], items[j] = items[j], items[i]
        self._wid_render_json()
        self._wid_focus(j, None, None)

    def _wid_ctl_up(self, i):
        self._wid_ctl_move(i, -1)

    def _wid_ctl_dn(self, i):
        self._wid_ctl_move(i, 1)

    # 字段级 + (标题最右): 该控件末尾追加一个空白字段, 焦点落到该空白行
    def _wid_add_field(self, i):
        from tkinter import messagebox
        if not (self.widget_data is not None and self.widget_path):
            messagebox.showwarning('widget', '请先在中间选择一个 .json 文件')
            return
        self._wid_sync_back()
        if i >= len(self.widget_data.get('widget', [])):
            return
        fields = self._wid_fields_of(i)
        newd = dict(fields)
        newd[''] = ''                               # 追加到末尾空白字段
        self._wid_set_fields(i, newd)
        self._wid_render_json()
        row = next((_row for _i, _k, _ke, _ve, _row in self._w_rows if _i == i and _k == ''), None)
        self._wid_focus(i, '', row)
        self._wlog('已添加字段到控件末尾(空行, 未保存)')

    # 字段级 - : 去除当前字段(maker 下 widget.type 不可删); 删空则保留空控件
    def _wid_del_field_k(self, i, fkey):
        from tkinter import messagebox
        if not (self.widget_data is not None and self.widget_path):
            messagebox.showwarning('widget', '请先在中间选择一个 .json 文件')
            return
        self._wid_sync_back()
        if i >= len(self.widget_data.get('widget', [])):
            return
        fields = self._wid_fields_of(i)
        if fkey not in fields:
            return
        if self.widget_maker and fkey == 'widget.type':
            self._wlog('maker json 的 widget.type 与 class 绑定, 不可删除(保留占位)')
            return
        newd = {k: v for k, v in fields.items() if k != fkey}
        self._wid_set_fields(i, newd)
        self._wid_render_json()
        self._wid_focus(i, None, None)              # 聚焦到该控件
        self._wlog('已删除字段 %s(未保存)' % fkey)

    # 字段上移 / 下移(按字段 key 定位并重排保序)
    def _wid_field_move(self, i, fkey, delta):
        if not (self.widget_data is not None and self.widget_path):
            return
        self._wid_sync_back()
        if i >= len(self.widget_data.get('widget', [])):
            return
        fields = self._wid_fields_of(i)
        ks = list(fields.keys())
        if fkey not in ks:
            return
        j = ks.index(fkey)
        nj = j + delta
        if nj < 0 or nj >= len(ks):
            return
        ks[j], ks[nj] = ks[nj], ks[j]
        newd = {k: fields[k] for k in ks}
        self._wid_set_fields(i, newd)
        self._wid_render_json()
        row = next((_row for _i, _k, _ke, _ve, _row in self._w_rows if _i == i and _k == fkey), None)
        self._wid_focus(i, fkey, row)

    #--------------- 控件模板 ---------------

    def _wid_load_templates(self):
        """brief: 扫描 scene json 抽取模板(内存缓存)
        @retval 无
        """
        root = self.in_abs['widget']
        if not os.path.isdir(root):
            self._w_templates = {}
            return
        templates = {}
        for dp, dns, fns in os.walk(root):
            for fn in fns:
                if not fn.lower().endswith('.json'):
                    continue
                full = os.path.join(dp, fn)
                try:
                    with open(full, 'r', encoding='utf-8') as fp:
                        data = json.load(fp)
                except Exception:
                    continue
                if not isinstance(data, dict) or data.get('type') != 'scene':
                    continue
                for w in data.get('widget', []):
                    if not isinstance(w, dict):
                        continue
                    wt = w.get('widget.type')
                    name = w.get('widget.myself')
                    if not wt or not name:
                        continue
                    fields = {k: v for k, v in w.items()
                              if k not in ('widget.type', 'widget.myself')}
                    templates.setdefault(wt, []).append(
                        {'name': name, 'fields': fields})
        self._w_templates = {wt: sorted(items, key=lambda x: x['name'])
                             for wt, items in sorted(templates.items())}

    def _wid_tpl_wheel(self, e):
        self._tpl_canv.yview_scroll(-1 * (e.delta // 120), 'units')

    def _wid_show_templates(self, i):
        """brief: 显示控件可用模板
        @param i 控件索引
        """
        self._wid_render_templates(i)
        self.widget_status.config(text='模板已加载到下方面板(点 T 应用)')

    def _wid_render_templates(self, control_idx):
        """brief: 渲染模板面板
        @param control_idx 控件索引
        """
        import tkinter as tk
        from tkinter import ttk
        for w in self._tpl_box.winfo_children():
            w.destroy()
        self._w_tpl_state = {'control_idx': control_idx, 'expanded': set()}
        if not self.widget_data or control_idx is None:
            ttk.Label(self._tpl_box, text='请先选择一个控件',
                      foreground='#888').pack(anchor='w', padx=(2, 2))
            return
        items = self.widget_data.get('widget', [])
        if control_idx >= len(items):
            return
        it = items[control_idx]
        wt = it.get('class') if self.widget_maker else it.get('widget.type')
        if not wt:
            ttk.Label(self._tpl_box, text='当前控件无 widget.type',
                      foreground='#888').pack(anchor='w', padx=(2, 2))
            return
        tlist = self._w_templates.get(wt, [])
        if not tlist:
            ttk.Label(self._tpl_box, text='尚不支持(%s)' % wt,
                      foreground='#888').pack(anchor='w', padx=(2, 2))
            return
        for t in tlist:
            name = t['name']
            row = tk.Frame(self._tpl_box, bg='#ffffff')
            row.pack(fill='x', padx=(2, 2), pady=(1, 1))
            ttk.Button(row, text='T', width=1,
                command=lambda n=name, i=control_idx:
                    self._wid_guard(self._wid_apply_template, i, n))\
                .pack(side='left', padx=(0, 2))
            btn = ttk.Button(row, text='?', width=1,
                command=lambda n=name: self._wid_guard(self._wid_tpl_toggle, n))
            btn.pack(side='right', padx=(2, 0))
            tk.Label(row, text=name, bg='#ffffff', anchor='w',
                     font=('Consolas', 9)).pack(side='left',
                        fill='x', expand=True, padx=(2, 4))
            detail = tk.Frame(self._tpl_box, bg='#f6f6f6')
            for k, v in t['fields'].items():
                dr = tk.Frame(detail, bg='#f6f6f6'); dr.pack(fill='x',
                                                           padx=(14, 2))
                tk.Label(dr, text=k, bg='#f6f6f6', font=('Consolas', 9),
                         foreground='#666').pack(side='left')
                tk.Label(dr, text=str(v), bg='#f6f6f6',
                         font=('Consolas', 9)).pack(side='left', padx=(6, 0))
            self._w_tpl_state['row_'    + name] = row
            self._w_tpl_state['btn_'    + name] = btn
            self._w_tpl_state['detail_' + name] = detail

    def _wid_tpl_toggle(self, name):
        """brief: 展开/折叠模板详情
        @param name 模板名
        """
        st = self._w_tpl_state
        detail = st.get('detail_' + name)
        btn    = st.get('btn_'    + name)
        row    = st.get('row_'    + name)
        if detail is None or btn is None or row is None:
            return
        if detail.winfo_ismapped():
            detail.pack_forget()
            btn.configure(text='?')
            st['expanded'].discard(name)
        else:
            detail.pack(in_=self._tpl_box, after=row, fill='x',
                        padx=(14, 2), pady=(0, 2))
            btn.configure(text='-')
            st['expanded'].add(name)

    def _wid_apply_template(self, control_idx, tpl_name):
        """brief: 应用模板到控件
        @param control_idx 控件索引
        @param tpl_name    模板名
        """
        if not self.widget_data:
            return
        items = self.widget_data.get('widget', [])
        if control_idx >= len(items):
            return
        it = items[control_idx]
        wt = it.get('class') if self.widget_maker else it.get('widget.type')
        tlist = self._w_templates.get(wt, [])
        tpl = next((t for t in tlist if t['name'] == tpl_name), None)
        if tpl is None:
            self._wlog('[tpl] 未找到模板 %s' % tpl_name)
            return
        new_fields = dict(tpl['fields'])          # 保留 type/myself, 其余覆盖
        if self.widget_maker:
            old_def = it.get('default', {})
            new_fields['widget.type'] = old_def.get('widget.type', wt)
            if 'widget.myself' in old_def:
                new_fields['widget.myself'] = old_def['widget.myself']
            it['default'] = new_fields
        else:
            new_fields['widget.type'] = it.get('widget.type', wt)
            if 'widget.myself' in it:
                new_fields['widget.myself'] = it['widget.myself']
            items[control_idx] = new_fields
        self._wid_render_json()
        self._wid_focus(control_idx, None, None)
        self._wlog('[tpl] 已应用模板: %s (未保存)' % tpl_name)

    # 预览 json: 弹窗显示修改后的完整 json
    def _wid_preview_json(self):
        import tkinter as tk
        from tkinter import scrolledtext
        self._wid_sync_back()
        top = tk.Toplevel(self.root)
        top.title('json 预览(修改后) - %s' % (os.path.basename(self.widget_path) if self.widget_path else '未命名'))
        top.geometry('680x560')
        t = scrolledtext.ScrolledText(top, wrap='none', font=('Consolas', 9), state='normal')
        t.pack(fill='both', expand=True, padx=8, pady=8)
        try:
            # 预览清除空白占位行(不改动当前编辑副本, 编辑时仍保留空行自由编辑)
            s = json.dumps(self._wid_clean_copy(self.widget_data), ensure_ascii=False, indent=4)
        except Exception as e:
            s = '%r' % e
        t.insert('1.0', s)
        t.configure(state='disabled')

    # maker json 加载(out 下默认配置, 可编辑保存; _silent=自动加载时不弹窗)
    def _wid_load_maker(self, _silent=False):
        from tkinter import messagebox
        path = os.path.join(self.out_abs['widget'], 'scui_ui_maker.json')
        if not os.path.exists(path):
            if not _silent:
                messagebox.showwarning('widget', '未找到 maker 默认配置:\n%s' % path)
            return
        try:
            if not _silent and not self._wid_confirm_discard():   # 手动加载前, 未保存修改需确认
                return
            with open(path, 'r', encoding='utf-8') as fp:
                self.widget_data = json.load(fp)
            # 补齐缺失 class: 以 analyze 全量类型为准, class 唯一且完备
            types = getattr(self, '_wid_class_types', None)
            if types:
                by_class = {}
                for w in self.widget_data.get('widget', []):
                    by_class.setdefault(w.get('class'), w)
                full = []
                for wt in types:
                    if wt in by_class:
                        full.append(by_class[wt])          # 保留已有条目的完整 default
                    else:
                        full.append({'class': wt, 'default': {'widget.type': wt}})
                self.widget_data['widget'] = full
            self.widget_path = path; self.widget_maker = True
            self._w_base = json.loads(json.dumps(self.widget_data))   # 记录基快照
            self.widget_status.config(text='已加载 maker json(默认配置): %s' % os.path.basename(path))
            self.widget_sub.set('maker json: key/value 直接修改')
            self._wid_switch(True)
            self._wid_maker_guard()
            self._wid_render_json(_top=True)
        except Exception as e:
            self.widget_status.config(text='加载失败: %r' % e)

    # 保存(json 可写; 先回填输入框, 保存才写文件)
    def _wid_save(self):
        if self.widget_data is None or not self.widget_path:
            self._wlog('[save] 无待保存内容')
            return
        if self.widget_path.lower().endswith('.c'):
            self._wlog('[save] .c 只读, 不允许修改写入')
            return
        try:
            self._wid_sync_back()
            # 保存时清除空白占位行(编辑时保留空行自由编辑, 落盘前统一去重)
            self.widget_data = self._wid_clean_copy(self.widget_data)
            with open(self.widget_path, 'w', encoding='utf-8') as fp:
                json.dump(self.widget_data, fp, ensure_ascii=False, indent=4)
            self._w_base = json.loads(json.dumps(self.widget_data))   # 保存后重置基快照
            self._wid_render_json()                     # 按清理后数据重绘(空行消失)
            self.widget_status.config(text='已保存: %s' % os.path.basename(self.widget_path))
            self._wlog('[save] 已写入: %s' % self._rel_log(self.widget_path))
        except Exception as e:
            self.widget_status.config(text='保存失败: %r' % e)
            self._wlog('[save] 失败: %r' % e)

#--------------- lang 子界面(左:语言列 | 中:句柄行列表 | 右:行详情) ---------------
    def _build_lang_tab(self):
        import tkinter as tk
        from tkinter import ttk, scrolledtext
        f = ttk.Frame(self.nb, padding=8)
        self.tabs['lang'] = f
        self.lang_data = None     # (languages, rows)
        self.lang_sel = None

        # 顶部
        top = ttk.Frame(f); top.pack(fill='x', pady=(2, 4))
        ttk.Button(top, text='执行 lang 打包', command=lambda: self._run('lang')).pack(side='right')

        # 主体: 三栏(占满, 不再内嵌 LOG)
        hp = ttk.Panedwindow(f, orient='horizontal'); hp.pack(fill='both', expand=True, pady=(6, 0))

        # 左: 语言列表
        lf = ttk.LabelFrame(hp, text=' language ', padding=(4, 4)); hp.add(lf, weight=1)
        self.ltree = ttk.Treeview(lf, show='tree', selectmode='browse')
        lvs = ttk.Scrollbar(lf, orient='vertical', command=self.ltree.yview)
        self.ltree.configure(yscrollcommand=lvs.set)
        self.ltree.pack(side='left', fill='both', expand=True); lvs.pack(side='right', fill='y')

        # 中: 句柄行列表(zh 提示, 带省略号)
        mf = ttk.LabelFrame(hp, text=' 句柄行(zh 预览) ', padding=(4, 4)); hp.add(mf, weight=3)
        self.mtree = ttk.Treeview(mf, columns=('idx',), show='tree headings', selectmode='browse')
        self.mtree.heading('#0', text='句柄行'); self.mtree.heading('idx', text='#')
        self.mtree.column('idx', width=40, anchor='e', stretch=False)
        mvs = ttk.Scrollbar(mf, orient='vertical', command=self.mtree.yview)
        self.mtree.configure(yscrollcommand=mvs.set)
        self.mtree.pack(side='left', fill='both', expand=True); mvs.pack(side='right', fill='y')
        self.mtree.bind('<<TreeviewSelect>>', self._lang_pick)

        # 右: 行详情(各语言内容)
        rf = ttk.LabelFrame(hp, text=' 行详情 ', padding=(4, 4)); hp.add(rf, weight=3)
        self.rtext = scrolledtext.ScrolledText(rf, wrap='word', font=('Consolas', 9), state='disabled')
        self.rtext.pack(fill='both', expand=True)

        self.nb.add(f, text='lang  ')
        self._lang_load()

    def _lang_load(self):
        self.ltree.delete(*self.ltree.get_children())
        self.mtree.delete(*self.mtree.get_children())
        self.rtext.configure(state='normal'); self.rtext.delete('1.0', 'end'); self.rtext.configure(state='disabled')
        self.lang_data = None; self.lang_sel = None

        src = self.in_abs['lang']
        cfg = os.path.join(src, 'scui_lang_parser.json')
        if not os.path.exists(cfg):
            self._append_log('lang', '未找到配置: %s\n' % self._rel_log(cfg))
            return
        try:
            with open(cfg, 'r', encoding='utf-8') as fp:
                j = json.load(fp)
            langs = j.get('language', [])
            xlsx_path = os.path.join(src, j.get('xlsx', ''))
            sheet_name = j.get('sheet', '')
            if not os.path.exists(xlsx_path):
                self._append_log('lang', '未找到 xlsx: %s\n' % self._rel_log(xlsx_path))
                return
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else '' for c in row])
            wb.close()
            self.lang_data = (langs, rows)
            # 左: 语言列表
            for i, l in enumerate(langs):
                self.ltree.insert('', 'end', text=l, iid='lang%d' % i)
            # 中: 句柄行(第 0 列为 zh, 带省略号)
            for i, row in enumerate(rows):
                zh = row[0] if row else ''
                show = zh if len(zh) <= 24 else zh[:22] + '…'
                self.mtree.insert('', 'end', iid='r%d' % i, text=show, values=(i,))
            self._append_log('lang', '已加载 %d 行, %d 语言\n' % (len(rows), len(langs)))
        except Exception as e:
            self._append_log('lang', '加载失败: %r\n' % e)

    def _lang_pick(self, _ev=None):
        sel = self.mtree.selection()
        if not sel or self.lang_data is None:
            return
        idx = int(sel[0][1:])
        langs, rows = self.lang_data
        if idx >= len(rows):
            return
        row = rows[idx]
        self.rtext.configure(state='normal')
        self.rtext.delete('1.0', 'end')
        self.rtext.insert('end', '句柄号: SCUI_LANG_0X%04x\n' % idx)
        self.rtext.insert('end', '-' * 40 + '\n')
        for ci, lang in enumerate(langs):
            val = row[ci] if ci < len(row) else ''
            self.rtext.insert('end', '%s: %s\n' % (lang, val))
        self.rtext.configure(state='disabled')

    #--------------- cwf 子界面(左:协议类型+注释 | 中:cwf浏览 | 右:json编辑) ---------------
    def _build_cwf_tab(self):
        import tkinter as tk
        from tkinter import ttk, scrolledtext
        f = ttk.Frame(self.nb, padding=8)
        self.tabs['cwf'] = f
        self.cwf_data = None
        self.cwf_path = None
        self.cwf_edit = None
        self._cwf_base = None
        self._cwf_rows = []
        self._cwf_active = None

        # 顶部工具条
        top = ttk.Frame(f); top.pack(fill='x', pady=(2, 4))
        ttk.Button(top, text='预览 json', command=self._cwf_preview).pack(side='left')
        ttk.Button(top, text='保存 json', command=self._cwf_save).pack(side='left', padx=(6, 0))
        ttk.Button(top, text='执行 cwf 打包', command=lambda: self._run('cwf')).pack(side='right')
        self.cwf_status = ttk.Label(top, text='未选择', foreground='#777')
        self.cwf_status.pack(side='left', padx=(12, 0))

        # 主体: 三栏(占满, 不再内嵌 LOG)
        hp = ttk.Panedwindow(f, orient='horizontal'); hp.pack(fill='both', expand=True, pady=(6, 0))

        # 左: 协议类型 + 注释
        lf = ttk.LabelFrame(hp, text=' cwf 协议类型 ', padding=(4, 4)); hp.add(lf, weight=3)
        self.cbook = ttk.Treeview(lf, columns=('anno',), show='tree headings', selectmode='browse')
        self.cbook.heading('#0', text='type'); self.cbook.heading('anno', text='annotation')
        self.cbook.column('anno', width=120, stretch=False)
        cvs = ttk.Scrollbar(lf, orient='vertical', command=self.cbook.yview)
        self.cbook.configure(yscrollcommand=cvs.set)
        self.cbook.pack(side='left', fill='both', expand=True); cvs.pack(side='right', fill='y')

        # 中: cwf 浏览(满足: 有 image.7z + 同名 .json)
        mf = ttk.LabelFrame(hp, text=' cwf 列表 ', padding=(4, 4)); hp.add(mf, weight=3)
        self.ctree = ttk.Treeview(mf, show='tree', selectmode='browse')
        ctvs = ttk.Scrollbar(mf, orient='vertical', command=self.ctree.yview)
        self.ctree.configure(yscrollcommand=ctvs.set)
        self.ctree.pack(side='left', fill='both', expand=True); ctvs.pack(side='right', fill='y')
        self.ctree.bind('<<TreeviewSelect>>', self._cwf_pick)

        # 右: json 编辑(layout 行列表, 每行 key|value 常驻输入框)
        rf = ttk.LabelFrame(hp, text=' 编辑(json 就地) ', padding=(4, 4)); hp.add(rf, weight=4)
        erow = ttk.Frame(rf); erow.pack(fill='x')
        for t, c in (('添加类型', self._cwf_add_type), ('删除类型', self._cwf_del_type),
                     ('添加字段', self._cwf_add_field), ('删除字段', self._cwf_del_field)):
            ttk.Button(erow, text=t, command=c).pack(side='left', padx=(0, 6))
        self.cwf_sub = tk.StringVar(value='中间选择一个 cwf: 各 layout 行 key|value 直接修改')
        ttk.Label(rf, textvariable=self.cwf_sub, foreground='#777').pack(anchor='w', padx=(2, 2))
        self.ccanv = tk.Canvas(rf, highlightthickness=0)
        csb = ttk.Scrollbar(rf, orient='vertical', command=self.ccanv.yview)
        self.ccanv.configure(yscrollcommand=csb.set)
        self._cbox = ttk.Frame(self.ccanv)
        self._cbox_id = self.ccanv.create_window((0, 0), window=self._cbox, anchor='nw')
        self._cbox.bind('<Configure>', lambda e: self.ccanv.configure(scrollregion=self.ccanv.bbox('all')))
        self.ccanv.bind('<Configure>', lambda e: (self.ccanv.itemconfigure(self._cbox_id, width=e.width),
                                                   self.ccanv.configure(scrollregion=self.ccanv.bbox('all'))))
        self.ccanv.pack(side='left', fill='both', expand=True); csb.pack(side='right', fill='y')
        self.ccanv.bind('<Enter>', lambda e: self.ccanv.bind_all('<MouseWheel>', self._cwf_wheel))
        self.ccanv.bind('<Leave>', lambda e: self.ccanv.unbind_all('<MouseWheel>'))

        self.nb.add(f, text='cwf  ')
        self._cwf_load_proto()
        self._cwf_load_list()

    def _cwf_wheel(self, e):
        self.ccanv.yview_scroll(-1 * (e.delta // 120), 'units')

    # 加载协议 json(类型 + 注释)
    def _cwf_load_proto(self):
        self.cbook.delete(*self.cbook.get_children())
        proto = os.path.join(self.tools, 'scui_pack_cwf.json')
        if not os.path.exists(proto):
            self._append_log('cwf', '未找到协议: %s\n' % self._rel_log(proto))
            return
        try:
            with open(proto, 'r', encoding='utf-8') as fp:
                j = json.load(fp)
            for item in j.get('scui_cwf_json_type', []):
                key = item.get('key', '?')
                annos = [v for k, v in item.items() if k == 'annotation']
                anno = ' ; '.join(annos) if annos else ''
                self.cbook.insert('', 'end', text=key, values=(anno,))
        except Exception as e:
            self._append_log('cwf', '协议加载失败: %r\n' % e)

    # 加载 cwf 列表(满足: 有 image.7z + 同名 .json)
    def _cwf_load_list(self):
        self.ctree.delete(*self.ctree.get_children())
        root = self.in_abs['cwf']
        if not os.path.isdir(root):
            self._append_log('cwf', 'cwf 路径不存在: %s\n' % self._rel_log(root))
            return
        for name in sorted(os.listdir(root)):
            d = os.path.join(root, name)
            if not os.path.isdir(d):
                continue
            has_7z = os.path.exists(os.path.join(d, 'image.7z'))
            has_json = os.path.exists(os.path.join(d, name + '.json'))
            if has_7z and has_json:
                self.ctree.insert('', 'end', text=name, iid=d)

    # 选中 cwf -> 加载 json
    def _cwf_pick(self, _ev=None):
        sel = self.ctree.selection()
        if not sel:
            return
        path = sel[0]
        if not os.path.isdir(path):
            return
        name = os.path.basename(path)
        json_path = os.path.join(path, name + '.json')
        if not os.path.exists(json_path):
            return
        if not self._cwf_confirm_discard():
            return
        try:
            with open(json_path, 'r', encoding='utf-8') as fp:
                self.cwf_data = json.load(fp)
            self.cwf_path = json_path
            self._cwf_base = json.loads(json.dumps(self.cwf_data))
            self.cwf_status.config(text='已加载: %s' % name)
            self._cwf_render()
        except Exception as e:
            self.cwf_status.config(text='加载失败: %r' % e)

    # 渲染 layout 行(每个条目 = 标题 + key|value 行)
    def _cwf_render(self):
        import tkinter as tk
        from tkinter import ttk
        for w in self._cbox.winfo_children():
            w.destroy()
        self._cwf_rows = []
        if self.cwf_data is None:
            return
        items = self.cwf_data.get('layout', [])
        for i, it in enumerate(items):
            title = '#%d  %s' % (i, it.get('type', '?'))
            tl = ttk.Label(self._cbox, text=title, foreground='#2266cc', font=('Consolas', 9, 'bold'))
            tl.pack(fill='x', anchor='w', pady=(6, 2))
            tl.bind('<Button-1>', lambda e, i=i: self._cwf_active_set(i, None))
            for f, (k, v) in enumerate(it.items()):
                kvar = tk.StringVar(value=str(k))
                vvar = tk.StringVar(value=str(v))
                row = ttk.Frame(self._cbox); row.pack(fill='x', padx=(2, 2))
                ke = ttk.Entry(row, textvariable=kvar, font=('Consolas', 9))
                ve = ttk.Entry(row, textvariable=vvar, font=('Consolas', 9))
                ke.pack(side='left', fill='x', expand=True, padx=(0, 8))
                ve.pack(side='left', fill='x', expand=True)
                ke.bind('<FocusIn>', lambda e, i=i, f=f: self._cwf_active_set(i, f))
                ve.bind('<FocusIn>', lambda e, i=i, f=f: self._cwf_active_set(i, f))
                self._cwf_rows.append([i, f, kvar, vvar])
            ttk.Label(self._cbox, text='').pack()

    def _cwf_active_set(self, i, f):
        self._cwf_active = (i, f)

    def _cwf_fields_of(self, item):
        return self.cwf_data['layout'][item]

    def _cwf_sync_back(self):
        if not self.cwf_data:
            return
        groups = {}
        for i, f, ke, ve in self._cwf_rows:
            groups.setdefault(i, []).append((ke.get(), ve.get()))
        for i, pairs in groups.items():
            if i >= len(self.cwf_data['layout']):
                continue
            d = {}
            for k, v in pairs:
                if k:
                    d[k] = v
            self.cwf_data['layout'][i] = d

    def _cwf_dirty(self):
        if self.cwf_data is None or self._cwf_base is None:
            return False
        self._cwf_sync_back()
        a = json.dumps(self.cwf_data, ensure_ascii=False, sort_keys=True)
        b = json.dumps(self._cwf_base, ensure_ascii=False, sort_keys=True)
        return a != b

    def _cwf_confirm_discard(self):
        if not self._cwf_dirty():
            return True
        from tkinter import messagebox
        return messagebox.askyesno('未保存的修改', '当前 cwf json 有未保存的修改，确定丢弃吗？', parent=self.root)

    def _cwf_new_type(self):
        return {'type': 'scui_cwf_json_type_img_simple', 'x': 0, 'y': 0, 'image_src': []}

    def _cwf_add_type(self):
        from tkinter import messagebox
        if self.cwf_data is None:
            messagebox.showwarning('cwf', '请先在中间选择一个 cwf')
            return
        self._cwf_sync_back()
        items = self.cwf_data.setdefault('layout', [])
        pos = self._cwf_active[0] if self._cwf_active else None
        if pos is None:
            self._append_log('cwf', '请定位操作位置\n')
            return
        items.insert(pos + 1, self._cwf_new_type())
        self._cwf_render()
        self._append_log('cwf', '已添加类型(未保存)\n')

    def _cwf_del_type(self):
        from tkinter import messagebox
        if self.cwf_data is None:
            messagebox.showwarning('cwf', '请先在中间选择一个 cwf')
            return
        self._cwf_sync_back()
        pos = self._cwf_active[0] if self._cwf_active else None
        if pos is None:
            self._append_log('cwf', '请定位操作位置\n')
            return
        items = self.cwf_data.get('layout', [])
        if pos < len(items):
            del items[pos]
            self._cwf_render()
            self._append_log('cwf', '已删除类型(未保存)\n')

    def _cwf_add_field(self):
        from tkinter import messagebox
        if self.cwf_data is None:
            messagebox.showwarning('cwf', '请先在中间选择一个 cwf')
            return
        self._cwf_sync_back()
        act = self._cwf_active
        pos = act[0] if act else None
        if pos is None:
            self._append_log('cwf', '请定位操作位置\n')
            return
        items = self.cwf_data.get('layout', [])
        if pos >= len(items):
            return
        f = act[1] if act else None
        fields = self._cwf_fields_of(pos)
        newd = {}
        for j, (k, v) in enumerate(fields.items()):
            newd[k] = v
            if f is not None and j == f:
                newd[''] = ''
        if f is None:
            newd[''] = ''
        self.cwf_data['layout'][pos] = newd
        self._cwf_render()
        self._append_log('cwf', '已添加字段(未保存)\n')

    def _cwf_del_field(self):
        from tkinter import messagebox
        if self.cwf_data is None:
            messagebox.showwarning('cwf', '请先在中间选择一个 cwf')
            return
        self._cwf_sync_back()
        act = self._cwf_active
        if act is None or act[0] is None or act[1] is None:
            self._append_log('cwf', '请定位操作位置(点 key/value 输入框)\n')
            return
        pos, f = act
        items = self.cwf_data.get('layout', [])
        if pos >= len(items):
            return
        fields = self._cwf_fields_of(pos)
        ks = list(fields.keys())
        if f >= len(ks):
            return
        kdel = ks[f]
        newd = {k: v for k, v in fields.items() if k != kdel}
        if not newd:
            del items[pos]
            self._append_log('cwf', '已删除字段 %s, 类型已空故删除(未保存)\n' % kdel)
        else:
            self.cwf_data['layout'][pos] = newd
            self._append_log('cwf', '已删除字段 %s(未保存)\n' % kdel)
        self._cwf_render()

    def _cwf_preview(self):
        import tkinter as tk
        from tkinter import scrolledtext
        self._cwf_sync_back()
        top = tk.Toplevel(self.root)
        top.title('cwf json 预览 - %s' % (os.path.basename(self.cwf_path) if self.cwf_path else '未命名'))
        top.geometry('680x560')
        t = scrolledtext.ScrolledText(top, wrap='none', font=('Consolas', 9), state='normal')
        t.pack(fill='both', expand=True, padx=8, pady=8)
        try:
            s = json.dumps(self.cwf_data, ensure_ascii=False, indent=4)
        except Exception as e:
            s = '%r' % e
        t.insert('1.0', s)
        t.configure(state='disabled')

    def _cwf_save(self):
        if self.cwf_data is None or not self.cwf_path:
            self._append_log('cwf', '[save] 无待保存内容\n')
            return
        try:
            self._cwf_sync_back()
            with open(self.cwf_path, 'w', encoding='utf-8') as fp:
                json.dump(self.cwf_data, fp, ensure_ascii=False, indent=4)
            self._cwf_base = json.loads(json.dumps(self.cwf_data))
            self.cwf_status.config(text='已保存: %s' % os.path.basename(self.cwf_path))
            self._append_log('cwf', '[save] 已写入: %s\n' % self._rel_log(self.cwf_path))
        except Exception as e:
            self.cwf_status.config(text='保存失败: %r' % e)
            self._append_log('cwf', '[save] 失败: %r\n' % e)

    def _build_font_tab(self):
        import tkinter as tk
        from tkinter import ttk
        f = ttk.Frame(self.nb, padding=8)
        self.tabs['font'] = f
        self.font_status = tk.StringVar(value='未加载')

        # 顶部工具条: json 三键 + 执行打包(走 font_bin->font_out)
        bar = ttk.Frame(f); bar.pack(fill='x', pady=(2, 4))
        for t, c in (('加载 json', lambda: self._font_load(False)),
                     ('预览 json', self._font_preview),
                     ('保存 json', self._font_save)):
            ttk.Button(bar, text=t, command=c).pack(side='left')
        ttk.Button(bar, text='执行 font 打包', command=lambda: self._run('font')).pack(side='right')
        ttk.Button(bar, text='刷新 font 源', command=self._font_src_refresh).pack(side='right')
        ttk.Button(bar, text='lv_font_conv 简介', command=self._font_conv_info).pack(side='right')
        self._font_conv_lbl = ttk.Label(bar, text='检测 lv_font_conv…', foreground='#888')
        self._font_conv_lbl.pack(side='right', padx=(4, 0))
        ttk.Label(bar, textvariable=self.font_status, foreground='#777').pack(side='left', padx=(12, 0))

        # 主体: 上 = bin 整合(固定, 保证"语言/字库"按钮列显示) / 下 = bin 生成(占剩余)
        # ---------- 上半: bin 整合 font_bin->font_out ----------
        top_fr = ttk.LabelFrame(f, text=' bin 整合 (font_bin → font_out) ',
                                padding=(4, 4))
        top_fr.configure(height=245)          # 固定高度: 刚好完全显示"语言/字库"六按钮, 不挤压下半
        top_fr.pack_propagate(False)
        top_fr.pack(side='top', fill='x', pady=(4, 0))
        hp = ttk.Panedwindow(top_fr, orient='horizontal'); hp.pack(fill='both', expand=True, pady=(2, 0))
        lf = ttk.LabelFrame(hp, text=' 字库(json) ', padding=(4, 4)); hp.add(lf, weight=1)   # 三列各 1/3
        self.ftree = ttk.Treeview(lf, columns=('size',), show='tree headings', selectmode='browse')
        self.ftree.heading('#0', text='字库条目'); self.ftree.heading('size', text='字号')
        self.ftree.column('size', width=60, anchor='e', stretch=False)
        fvs = ttk.Scrollbar(lf, orient='vertical', command=self.ftree.yview)
        self.ftree.configure(yscrollcommand=fvs.set)
        self.ftree.pack(side='left', fill='both', expand=True); fvs.pack(side='right', fill='y')
        self.ftree.bind('<<TreeviewSelect>>', self._on_font_sel)
        rf = ttk.Frame(hp); hp.add(rf, weight=1)
        self.font_json = None
        self.font_sel = None       # (lang_idx, item_idx)
        ops = ttk.LabelFrame(rf, text='语言 / 字库'); ops.pack(side='left', fill='y', padx=(0, 8))
        for t, c in (('添加语言', self._font_addlang), ('删除语言', self._font_dellang),
                     ('修改语言', self._font_modlang), ('添加字库', self._font_add),
                     ('删除字库', self._font_del), ('修改字库', self._font_modlib)):
            ttk.Button(ops, text=t, command=c).pack(fill='x', pady=2)
        ef = ttk.LabelFrame(rf, text=' 编辑(json 特性) '); ef.pack(side='left', fill='both', expand=True)
        self._build_font_editor(ef)
        # range 辅助列(bin 整合右侧, 空间大; 单行输出 + 多行输入)
        rast_fr = ttk.LabelFrame(hp, text=' range 辅助 ', padding=(4, 4)); hp.add(rast_fr, weight=1)
        ttk.Label(rast_fr, text='输出(单行)').pack(anchor='w')
        self._aux_out_var = tk.StringVar()
        self._aux_out = ttk.Entry(rast_fr, textvariable=self._aux_out_var, state='readonly',
                                  font=('Consolas', 9))
        self._aux_out.pack(fill='x', pady=(1, 4))
        self._aux_out.bind('<Double-Button-1>', lambda e: self._font_aux_copy())
        ttk.Label(rast_fr, text='输入(多行: 字符 / 0x.. / 十进制 / 区间); 输出框双击复制').pack(anchor='w')
        self._aux_in = tk.Text(rast_fr, font=('Consolas', 9), height=10)
        self._aux_in.pack(fill='both', expand=True)
        ab = ttk.Frame(rast_fr); ab.pack(fill='x', pady=(2, 0))
        ttk.Button(ab, text='生成 range', command=self._font_aux_gen).pack(side='left')

        # ---------- 下半: bin 生成 font_src->font_bin (外源 lv_font_conv) ----------
        bot_fr = ttk.LabelFrame(f, text=' bin 生成 (font_src → font_bin, 外源 lv_font_conv) ',
                                padding=(4, 4)); bot_fr.pack(side='bottom', fill='both', expand=True, pady=(0, 4))
        htop = ttk.Panedwindow(bot_fr, orient='horizontal'); htop.pack(fill='both', expand=True)
        # 列1: font 源(字库源文件)
        fsrc_fr = ttk.LabelFrame(htop, text=' font 源 ', padding=(4, 4)); htop.add(fsrc_fr, weight=1)
        self.font_src_dir = os.path.join(self.ui, 'font_src')
        self._font_src_files = []
        self.ftsrc = ttk.Treeview(fsrc_fr, columns=('kb',), show='tree headings', selectmode='browse')
        self.ftsrc.heading('#0', text='文件'); self.ftsrc.heading('kb', text='大小')
        self.ftsrc.column('kb', width=64, anchor='e', stretch=False)
        fsv = ttk.Scrollbar(fsrc_fr, orient='vertical', command=self.ftsrc.yview)
        self.ftsrc.configure(yscrollcommand=fsv.set)
        self.ftsrc.pack(side='left', fill='both', expand=True); fsv.pack(side='right', fill='y')
        # 列2: font.bin 任务
        tf_fr = ttk.LabelFrame(htop, text=' font.bin 任务 ', padding=(4, 4)); htop.add(tf_fr, weight=2)
        row0 = ttk.Frame(tf_fr); row0.pack(fill='x')
        self._font_area = ttk.Frame(tf_fr); self._font_area.pack(fill='both', expand=True, pady=(2, 0))
        self._font_tasks = self._font_task_default()
        self._font_task_build(self._font_area)
        tbar = ttk.Frame(tf_fr); tbar.pack(fill='x', pady=(2, 0))
        ttk.Button(tbar, text='添加', command=self._font_task_add).pack(side='left')
        ttk.Button(tbar, text='删除末项', command=self._font_task_del).pack(side='left', padx=(4, 0))
        ttk.Button(tbar, text='检测 lv_font_conv', command=self._font_conv_check).pack(side='left', padx=(4, 0))
        self._font_conv_btn = ttk.Button(tbar, text='批量生成 bin', command=self._font_conv_batch)
        self._font_conv_btn.pack(side='right')

        self.nb.add(f, text='font  ')
        self._font_load(False)
        self._font_src_refresh()
        self._font_conv_check()

    #--------------- font 前半段: font_src -> font_bin (外源 lv_font_conv) ---------------
    def _font_tf_wheel(self, e):
        # 内容不满一屏时不滚动, 避免滚出顶部留白
        vh = self._tf_canv.winfo_height() or 200
        bh = self._tf_canv.bbox('all')[3]
        if bh <= vh:
            self._tf_canv.yview_moveto(0)
            return
        self._tf_canv.yview_scroll(-1 * (e.delta // 120), 'units')

    def _font_src_refresh(self):
        """扫描 font_src 下发字库源(ttf/woff), 填左树 + 刷新任务表字体下拉"""
        self.ftsrc.delete(*self.ftsrc.get_children())
        self._font_src_files = []
        if not os.path.isdir(self.font_src_dir):
            return
        for e in sorted(os.listdir(self.font_src_dir)):
            if not e.lower().endswith(('.ttf', '.woff', '.woff2', '.otf')):
                continue
            full = os.path.join(self.font_src_dir, e)
            try:
                kb = (os.path.getsize(full) + 1023) // 1024
            except OSError:
                kb = 0
            self.ftsrc.insert('', 'end', text=e, values=('%dK' % kb,))
            self._font_src_files.append(e)
        if hasattr(self, '_font_area'):
            self._font_task_build(self._font_area)   # 刷新任务行字体下拉选项

    def _font_task_default(self):
        """默认任务批次(对齐 font_src/lv_font_conv.py 实际执行, 可编辑/增删)"""
        sz = '8,12,16,20,24,32,40,48,56,64,72,80,88'
        t = [
            # 类型: conv(生成) / copy(拷贝字库); bin 填前缀, size 逗号分隔生成多条 _size.bin
            {'t': 'conv', 'font': 'font_zh_en.ttf',   'range': '-r 0x20-0x7f',              'size': sz, 'bpp': '8', 'bin': 'font_ascii',  'mode': 'diy'},
            {'t': 'conv', 'font': 'font_symbol.woff', 'range': '-r ' + _FONT_SYMBOL_RANGE, 'size': sz, 'bpp': '8', 'bin': 'font_symbol', 'mode': 'ord'},
            {'t': 'conv', 'font': 'font_en_2.ttf',    'range': '-r 0x00-0x7f',              'size': '32,36', 'bpp': '8', 'bin': 'font_en',  'mode': 'ord'},
            {'t': 'conv', 'font': 'font_zh_2.ttf',    'range': _FONT_EU_RANGE,              'size': '32,36', 'bpp': '4', 'bin': 'font_eu',  'mode': 'diy'},
            {'t': 'conv', 'font': 'font_zh_1.ttf',    'range': _FONT_CJK_RANGE,             'size': '32,36', 'bpp': '2', 'bin': 'font_cjk', 'mode': 'diy'},
            {'t': 'copy', 'font': 'font_tinyTTF.ttf', 'range': '', 'size': '', 'bpp': '',   'bin': 'font_tinyTTF.ttf', 'mode': 'copy'},
        ]
        return t

    def _font_task_build(self, area):
        """重建任务区: 始终显示任务表, 不可用时整表置灰禁用"""
        import tkinter as tk
        from tkinter import ttk
        for w in area.winfo_children():
            w.destroy()
        self._font_w = []
        ok = self._font_conv_effective()
        cols = (('字体', 12), ('range', 24), ('size', 8), ('bpp', 6), ('bin前缀', 12), ('模式', 6))
        # 表头固定(Label 可读); 字体/模式为下拉列需加宽补偿箭头, Entry 列 +1 补偿内边距
        head = ttk.Frame(area); head.pack(fill='x', padx=(2, 2), pady=(1, 1))
        for i, (txt, w) in enumerate(cols):
            w = w + 4 if i in (0, 5) else w + 1
            ttk.Label(head, text=txt, foreground='#333' if ok else '#b0b0b0', width=w, anchor='w')\
                .pack(side='left', padx=(0, 2))
        ttk.Label(head, text='', width=7).pack(side='right')
        # 滚动任务表(只有任务行, 不含表头)
        ttc = ttk.Frame(area); ttc.pack(fill='both', expand=True)
        self._tf_canv = tk.Canvas(ttc, highlightthickness=0, bg='#ffffff')
        tvs = ttk.Scrollbar(ttc, orient='vertical', command=self._tf_canv.yview)
        self._tf_canv.configure(yscrollcommand=tvs.set)
        self._tf_box = ttk.Frame(self._tf_canv)
        self._tf_id = self._tf_canv.create_window((0, 0), window=self._tf_box, anchor='nw')
        def _tf_reconfigure(e):
            # 滚动区域调整 + 滚动条可见性判断
            self._tf_canv.configure(scrollregion=self._tf_canv.bbox('all'))
            vh = self._tf_canv.winfo_height() or 200
            bh = self._tf_canv.bbox('all')[3]
            needs_scroll = bh > vh
            if needs_scroll:
                if not tvs.winfo_ismapped():
                    tvs.pack(side='right', fill='y')
            else:
                if tvs.winfo_ismapped():
                    tvs.pack_forget()
                self._tf_canv.yview_moveto(0)
        self._tf_box.bind('<Configure>', _tf_reconfigure)
        self._tf_canv.bind('<Configure>',
            lambda e: (self._tf_canv.itemconfigure(self._tf_id, width=e.width),
                       _tf_reconfigure(e)))
        self._tf_canv.pack(side='left', fill='both', expand=True); tvs.pack(side='right', fill='y')
        self._tf_canv.bind('<Enter>', lambda e: self._tf_canv.bind_all('<MouseWheel>', self._font_tf_wheel))
        self._tf_canv.bind('<Leave>', lambda e: self._tf_canv.unbind_all('<MouseWheel>'))
        srcs = self._font_src_files or ['font_zh_en.ttf', 'font_symbol.woff',
                                        'font_en_2.ttf', 'font_zh_1.ttf', 'font_zh_2.ttf', 'font_tinyTTF.ttf']
        st = 'normal' if ok else 'disabled'
        stro = 'normal' if ok else 'disabled'
        for rows in self._font_tasks:
            row = tk.Frame(self._tf_box, bg='#ffffff'); row.pack(fill='x', padx=(2, 2), pady=1)
            fvar = tk.StringVar(value=rows['font'])
            rvar = tk.StringVar(value=rows['range'])
            svar = tk.StringVar(value=rows['size'])
            bvar = tk.StringVar(value=rows['bpp'])
            nvar = tk.StringVar(value=rows['bin'])
            mvar = tk.StringVar(value=rows['mode'])
            ttk.Combobox(row, textvariable=fvar, values=srcs, width=12,
                         state='readonly' if (not ok or rows['t'] == 'copy') else 'normal')\
                .pack(side='left', padx=(0, 2))
            ttk.Entry(row, textvariable=rvar, state=st, width=24).pack(side='left', padx=(0, 2))
            ttk.Entry(row, textvariable=svar, state=st, width=8).pack(side='left', padx=(0, 2))
            ttk.Entry(row, textvariable=bvar, state=st, width=6).pack(side='left', padx=(0, 2))
            ttk.Entry(row, textvariable=nvar, state=st, width=12).pack(side='left', padx=(0, 2))
            ttk.Combobox(row, textvariable=mvar, values=('ord', 'diy', 'copy'),
                         width=6, state='readonly' if ok else 'disabled').pack(side='left', padx=(0, 4))
            ttk.Button(row, text='▶生成', width=6, state=stro,
                       command=lambda i=len(self._font_w): self._font_conv_one(i)).pack(side='right')
            self._font_w.append([fvar, rvar, svar, bvar, nvar, mvar])

    def _font_task_add(self):
        # 默认追加一份复制当前首批? 追加一个空 conv 任务(字体取首个源)
        src = self._font_src_files[0] if self._font_src_files else 'font_zh_en.ttf'
        self._font_tasks.append({'t': 'conv', 'font': src, 'range': '-r 0x20-0x7f',
                                 'size': '16', 'bpp': '8', 'bin': 'font_new_%d' % len(self._font_tasks),
                                 'mode': 'diy'})
        self._font_task_build(self._font_area)

    def _font_task_del(self):
        if self._font_tasks:
            self._font_tasks.pop()
        self._font_task_build(self._font_area)

    def _font_task_row_read(self, i):
        """读取任务行 vars -> dict"""
        if not (0 <= i < len(self._font_w)):
            return None
        fvar, rvar, svar, bvar, nvar, mvar = self._font_w[i]
        return {'font': fvar.get().strip(), 'range': rvar.get().strip(),
                'size': svar.get().strip(), 'bpp': bvar.get().strip(),
                'bin': nvar.get().strip(), 'mode': mvar.get().strip()}

    def _font_aux_gen(self):
        """range 辅助: 输入字符/码点/范围 -> 合并去重 -r 表达式"""
        import re
        text = self._aux_in.get('1.0', 'end').strip()
        if not text:
            return
        slots = set()
        for token in re.split(r'[\s,;，；]+', text):
            token = token.strip()
            if not token:
                continue
            # 区间: 0x..-0x.. / X-X / 字符-字符
            m = re.match(r'^(0x[0-9a-fA-F]+|\d+)\s*[-–]\s*(0x[0-9a-fA-F]+|\d+)$', token)
            if m:
                lo = int(m.group(1), 16) if m.group(1).startswith('0x') else int(m.group(1))
                hi = int(m.group(2), 16) if m.group(2).startswith('0x') else int(m.group(2))
                slots.update(range(lo, hi + 1))
                continue
            # 单点 0x 或 十进制
            if token.startswith('0x'):
                slots.add(int(token, 16))
            elif token.isdigit():
                slots.add(int(token))
            else:
                # 多字符每个
                for ch in token:
                    slots.add(ord(ch))
        if not slots:
            return
        sorted_ = sorted(slots)
        # 合并连续区间
        parts = []
        lo = sorted_[0]
        hi = sorted_[0]
        for cp in sorted_[1:]:
            if cp == hi + 1:
                hi = cp
            else:
                parts.append('0x%X' % lo if lo == hi else '0x%X-0x%X' % (lo, hi))
                lo = hi = cp
        parts.append('0x%X' % lo if lo == hi else '0x%X-0x%X' % (lo, hi))
        expr = ' '.join('-r ' + p for p in parts)
        self._aux_out_var.set(expr)
        self._aux_copy.configure(state='normal')

    def _font_aux_copy(self):
        """复制 range 到剪贴板"""
        txt = self._aux_out_var.get().strip()
        if txt:
            self.root.clipboard_clear()
            self.root.clipboard_append(txt)
            self._append_log('font', '[range] 已复制到剪贴板\n')

    def _font_conv_effective(self):
        """可用状态: 按真实检测结果 lv_font_conv 是否可用"""
        return getattr(self, '_font_conv_ok', False)

    def _font_conv_check(self):
        """cmd 检测 lv_font_conv 是否可用(用 shell=True 适配 Windows)"""
        import subprocess, tempfile
        ok = True
        try:
            r = subprocess.run('lv_font_conv -h', shell=True, capture_output=True, timeout=5)
            ok = (r.returncode == 0)
        except Exception:
            ok = False
        self._font_conv_ok = ok
        if ok:
            self._font_conv_lbl.config(text='lv_font_conv 可用 ✓', foreground='#0a0')
            self._font_conv_btn.configure(state='normal')
        else:
            self._font_conv_lbl.config(text='lv_font_conv 未安装/不可用', foreground='#c00')
            self._font_conv_btn.configure(state='disabled')
        if hasattr(self, '_font_area'):
            self._font_task_build(self._font_area)   # 切换显示任务表/提示

    def _font_conv_info(self):
        """lv_font_conv 简介弹窗: 任务表使用说明 + lv_font_conv.txt 安装构建指引"""
        import tkinter as tk
        from tkinter import scrolledtext
        top = tk.Toplevel(self.root); top.title('lv_font_conv 简介'); top.geometry('680x460')
        top.transient(self.root)
        txt = scrolledtext.ScrolledText(top, wrap='word', font=('Consolas', 9), state='disabled')
        txt.pack(fill='both', expand=True, padx=6, pady=6)
        txt.configure(state='normal')
        # 固定显示 lv_font_conv.txt 前 80 行
        p = os.path.join(self.font_src_dir, 'lv_font_conv.txt')
        if os.path.isfile(p):
            try:
                buf = ''.join(open(p, encoding='utf-8').readlines()[:80])
            except Exception:
                buf = '(lv_font_conv.txt 读取失败)'
        else:
            buf = '(lv_font_conv.txt 不存在)'
        txt.insert('1.0', buf)
        txt.configure(state='disabled')
        top.grab_set()

    def _font_conv_batch(self):
        for i in range(len(self._font_w)):
            self._font_conv_one(i)

    def _font_conv_one(self, i):
        """执行第 i 个任务: 走 lv_font_conv(subprocess) 或拷贝字库.
        size 逗号分隔生成多条; bin 填前缀自动拼 _size.bin.
        """
        import shutil, subprocess
        t = self._font_task_row_read(i)
        if not t or not t['font'] or not t['bin']:
            self._append_log('font', '[font] 任务%d 缺 字体/bin 名, 跳过\n' % i)
            return
        bin_dir = self.in_abs['font']                     # font_bin
        os.makedirs(bin_dir, exist_ok=True)
        src_font = os.path.join(self.font_src_dir, t['font'])
        is_copy = (t['mode'] == 'copy') or (not t['size'] and not t['range'])
        if is_copy:
            # 拷贝: bin 列是完整文件名(如 font_tinyTTF.ttf)或补 .bin
            out = t['bin'] if '.' in os.path.basename(t['bin']) else t['bin'] + '.bin'
            out_path = os.path.join(bin_dir, out)
            try:
                shutil.copy2(src_font, out_path)
                self._append_log('font', '[font] copy: %s -> %s\n' % (_rel_log(src_font), out_path))
            except OSError as e:
                self._append_log('font', '[font] copy 失败 %s: %r\n' % (t['font'], e))
            return
        if not self._font_conv_effective():
            self._append_log('font', '[font] lv_font_conv 不可用, 跳过 %s\n' % t['bin'])
            return
        sizes = [s.strip() for s in t['size'].split(',') if s.strip()]
        if not sizes:
            self._append_log('font', '[font] 任务%d 无 size, 跳过\n' % i)
            return
        mode = t['mode'] if t['mode'] in ('ord', 'diy') else 'diy'
        for sz in sizes:
            out = '%s_%s.bin' % (t['bin'], sz)            # bin 前缀自动拼 _size.bin
            out_path = os.path.join(bin_dir, out)
            cmd = (_FONT_CMD_ORD if mode == 'ord' else _FONT_CMD_DIY).format(
                t['font'], t['range'], sz, t['bpp'], out_path)
            self._append_log('font', '[font] $ %s\n' % cmd)
            try:
                r = subprocess.run(cmd, shell=True, cwd=self.font_src_dir,
                                   capture_output=True, timeout=600)
                err = r.stderr.decode('utf-8', 'replace') if r.stderr else ''
                if err:
                    self._append_log('font', '[font] %s\n' % err)
                self._append_log('font', '[font] 完成(%d): %s\n' % (r.returncode, out))
            except Exception as e:
                self._append_log('font', '[font] 执行异常: %r\n' % e)

    def _build_font_editor(self, rf):
        import tkinter as tk
        from tkinter import ttk
        self._fe = {}
        grid = ttk.Frame(rf); grid.pack(fill='x')
        ttk.Label(grid, text='lang').grid(row=0, column=0, sticky='w', padx=(0, 8), pady=3)
        self._fe['lang'] = tk.Entry(grid, width=16, state='readonly')   # 用 tk.Entry, readonly 才生效
        self._fe['lang'].grid(row=0, column=1, sticky='ew', pady=3)
        ttk.Label(grid, text='name').grid(row=1, column=0, sticky='w', padx=(0, 8), pady=3)
        self._fe['name'] = ttk.Entry(grid, width=30)
        self._fe['name'].grid(row=1, column=1, sticky='ew', pady=3)
        ttk.Label(grid, text='size').grid(row=2, column=0, sticky='w', padx=(0, 8), pady=3)
        self._fe['size'] = ttk.Entry(grid, width=10)
        self._fe['size'].grid(row=2, column=1, sticky='w', pady=3)
        ttk.Label(grid, text='base_line_ext').grid(row=3, column=0, sticky='w', padx=(0, 8), pady=3)
        self._fe['base'] = ttk.Entry(grid, width=10)
        self._fe['base'].grid(row=3, column=1, sticky='w', pady=3)
        ttk.Label(grid, text='line_height_ext').grid(row=4, column=0, sticky='w', padx=(0, 8), pady=3)
        self._fe['line'] = ttk.Entry(grid, width=10)
        self._fe['line'].grid(row=4, column=1, sticky='w', pady=3)
        grid.columnconfigure(1, weight=1)

    def _font_json_path(self):
        return os.path.join(self.out_abs['font'], 'scui_font_package.json')

    def _font_rebuild(self):
        """从内存副本重建左树(保留已展开的语言分组, 保证 iid 与数组下标同步)."""
        openli = set()
        if hasattr(self, '_font_lang_iid'):
            for top, li in list(self._font_lang_iid.items()):
                try:
                    if self.ftree.item(top, 'open'):
                        openli.add(li)
                except Exception:
                    pass
        self._font_lang_iid = {}
        self.ftree.delete(*self.ftree.get_children())
        for li, item in enumerate(self.font_json.get('info', [])):
            lang = item.get('lang', '?')
            top = self.ftree.insert('', 'end', text=lang, values=('',),
                                    open=(li in openli))
            self._font_lang_iid[top] = li
            sizes = item.get('size', [])
            for ei, name in enumerate(item.get('name', [])):
                sz = sizes[ei] if ei < len(sizes) else 0
                self.ftree.insert(top, 'end', text=name, values=(sz,),
                                  iid='f%d_%d' % (li, ei))

    def _font_normalize(self):
        """归一化: 保证每个条目的 name/size/ext 长度对齐, 缺失 ext 补默认 0."""
        info = self.font_json.setdefault('info', [])
        for item in info:
            n = len(item.setdefault('name', []))
            item.setdefault('size', [0] * n)
            if len(item['size']) < n:
                item['size'] += [0] * (n - len(item['size']))
            ext = item.setdefault('ext', [None] * n)
            for i in range(n):
                if not isinstance(ext[i], dict):
                    ext[i] = {'base_line_ext': 0, 'line_height_ext': 0}

    def _font_load(self, _notify):
        import tkinter as tk
        p = self._font_json_path()
        try:
            with open(p, 'r', encoding='utf-8') as fp:
                self.font_json = json.load(fp)
            self._font_normalize()
            self._font_rebuild()
            self.font_status.set('已加载: %s' % os.path.basename(p))
        except Exception as e:
            self.font_status.set('加载失败: %r' % e)

    def _on_font_sel(self, _ev=None):
        # 支持直接传 iid(添加条目后填充); 否则从当前 selection 取
        if isinstance(_ev, str):
            iid = _ev
        else:
            sel = self.ftree.selection()
            if not sel:
                return
            iid = sel[0]
        # 顶层 lang 分组: 记录 li, 供添加字号使用
        if iid in getattr(self, '_font_lang_iid', {}):
            self.font_sel = (self._font_lang_iid[iid], -1)
            for k in self._fe:
                self._fe[k].configure(state='normal'); self._fe[k].delete(0, 'end')
                self._fe[k].configure(state='readonly' if k == 'lang' else 'normal')
            self.font_status.set('已选择分组 %s(可点「添加字号」)' % self.font_json['info'][self._font_lang_iid[iid]]['lang'])
            return
        if not iid.startswith('f'):
            return
        li, ei = map(int, iid[1:].split('_'))
        self.font_sel = (li, ei)
        item = self.font_json['info'][li]
        name = item['name'][ei]
        size = item['size'][ei] if ei < len(item['size']) else 0
        self._fe['lang'].configure(state='normal'); self._fe['lang'].delete(0, 'end'); self._fe['lang'].insert(0, item['lang']); self._fe['lang'].configure(state='readonly')
        self._fe['name'].delete(0, 'end'); self._fe['name'].insert(0, name)
        self._fe['size'].delete(0, 'end'); self._fe['size'].insert(0, str(size))
        # ext(可选)
        base = line = 0
        ext = item.get('ext')
        if ext and ei < len(ext) and isinstance(ext[ei], dict):
            base = ext[ei].get('base_line_ext', 0)
            line = ext[ei].get('line_height_ext', 0)
        self._fe['base'].delete(0, 'end'); self._fe['base'].insert(0, str(base))
        self._fe['line'].delete(0, 'end'); self._fe['line'].insert(0, str(line))

    def _font_prune(self):
        """合规性: 移除没有字库条目(空 name)的语言分组, 并从 lang 白名单剔除未被引用的语言."""
        from tkinter import messagebox
        removed = []
        info = self.font_json.get('info', [])
        new_info = []
        for item in info:
            names = [n for n in item.get('name', []) if str(n).strip()]
            if names:
                item['name'] = names
                if len(item['size']) != len(names):
                    item['size'] = [0] * len(names)
                new_info.append(item)
            else:
                removed.append(item.get('lang', '?'))
        self.font_json['info'] = new_info
        used = set(i.get('lang') for i in new_info)
        lang = self.font_json.setdefault('lang', [])
        self.font_json['lang'] = [l for l in lang if l in used]
        if removed:
            messagebox.showinfo('保存 json', '已移除无效语言分组(未含字库条目): %s' % ', '.join(removed), parent=self.root)
        return removed

    def _font_save(self):
        if self.font_json is None:
            return
        if self.font_sel:
            li, ei = self.font_sel
            item = self.font_json['info'][li]
            item['name'][ei] = self._fe['name'].get().strip()
            try:
                item['size'][ei] = int(self._fe['size'].get().strip())
            except Exception:
                pass
            # ext(可选, 空=0)
            def _to_int(s):
                try:
                    return int(str(s).strip())
                except Exception:
                    return 0
            b = _to_int(self._fe['base'].get()); l = _to_int(self._fe['line'].get())
            item.setdefault('ext', [None] * len(item['name']))
            while ei >= len(item['ext']):
                item['ext'].append(None)
            item['ext'][ei] = {'base_line_ext': b, 'line_height_ext': l}
        # 合规性: 去掉只加语言未加字库的无效分组
        self._font_prune()
        try:
            with open(self._font_json_path(), 'w', encoding='utf-8') as fp:
                json.dump(self.font_json, fp, ensure_ascii=False, indent=4)
            # 若当前选中了条目, 就地同步树行文本(不重建, 保持展开)
            if self.font_sel:
                li, ei = self.font_sel
                if li < len(self.font_json['info']):
                    it = self.font_json['info'][li]
                    if ei < len(it['name']):
                        self.ftree.item('f%d_%d' % (li, ei), text=it['name'][ei],
                                        values=(it['size'][ei] if ei < len(it['size']) else 0,))
            self.font_status.set('已保存')
        except Exception as e:
            self.font_status.set('保存失败: %r' % e)

    def _font_add(self):
        from tkinter import messagebox
        if self.font_json is None:
            messagebox.showwarning('font', '尚未加载 json，请先点「加载 json」')
            return
        sel = self.ftree.selection()
        if not sel:
            messagebox.showwarning('font', '请先点选左侧一个语言分组或某个条目，再添加')
            return
        iid = sel[0]
        if iid.startswith('f'):
            li = int(iid[1:].split('_')[0])   # 点的是某 bin -> 所属分组
        else:
            li = self._font_lang_iid.get(iid)
            if li is None:
                messagebox.showwarning('font', '请先点选左侧一个语言分组')
                return
        item = self.font_json['info'][li]
        lang = item['lang']
        item.setdefault('name', [])
        item.setdefault('size', [])
        base = 'font_%s_new_%d.bin' % (lang, len(item['name']))
        item['name'].append(base)
        item['size'].append(0)
        # 就地加数组 + 重建树(iid 与数组同步, 且保留展开), 再选中新条目
        self._font_rebuild()
        new_iid = 'f%d_%d' % (li, len(item['name']) - 1)
        top_iid = next((k for k, v in self._font_lang_iid.items() if v == li), None)
        if top_iid is not None:
            self.ftree.item(top_iid, open=True)
        self.ftree.selection_set(new_iid)
        self._on_font_sel(new_iid)
        self.font_status.set('已添加(未保存，点「保存 json」同步)')

    def _font_addlang(self):
        from tkinter import simpledialog, messagebox
        if self.font_json is None:
            messagebox.showwarning('font', '尚未加载 json，请先点「加载 json」')
            return
        lang = simpledialog.askstring('添加语言', '新语言名(如 en2):', parent=self.root)
        if not lang:
            return
        lang = lang.strip()
        if not lang:
            return
        self.font_json.setdefault('lang', [])
        if lang not in self.font_json['lang']:
            self.font_json['lang'].append(lang)
        self.font_json.setdefault('info', [])
        # 同名已存在则直接选中
        for li, item in enumerate(self.font_json['info']):
            if item.get('lang') == lang:
                self._font_rebuild()
                top = next((k for k, v in self._font_lang_iid.items() if v == li), None)
                if top is not None:
                    self.ftree.item(top, open=True)
                    self.ftree.selection_set(top)
                self.font_status.set('语言 %s 已存在，已选中' % lang)
                return
        self.font_json['info'].append({'lang': lang, 'name': [], 'size': []})
        self._font_rebuild()
        li = len(self.font_json['info']) - 1
        top = next((k for k, v in self._font_lang_iid.items() if v == li), None)
        if top is not None:
            self.ftree.item(top, open=True)
            self.ftree.selection_set(top)
        self.font_status.set('已添加语言 %s(可再点「添加字号」补条目)' % lang)

    def _font_modlang(self):
        from tkinter import simpledialog, messagebox
        if self.font_json is None:
            messagebox.showwarning('font', '尚未加载 json，请先点「加载 json」')
            return
        sel = self.ftree.selection()
        if not sel:
            messagebox.showwarning('font', '请先选中一个语言分组')
            return
        iid = sel[0]
        if iid.startswith('f'):
            li = int(iid[1:].split('_')[0])
        else:
            li = self._font_lang_iid.get(iid)
            if li is None:
                messagebox.showwarning('font', '请先选中一个语言分组')
                return
        item = self.font_json['info'][li]
        old = item.get('lang', '')
        new = simpledialog.askstring('修改语言', '新语言名:', initialvalue=old, parent=self.root)
        if not new:
            return
        new = new.strip()
        if not new:
            return
        item['lang'] = new
        L = self.font_json.setdefault('lang', [])
        if old in L and not any(i.get('lang') == old for i in self.font_json['info']):
            L.remove(old)
        if new not in L:
            L.append(new)
        self._font_rebuild()
        top = next((k for k, v in self._font_lang_iid.items() if v == li), None)
        if top is not None:
            self.ftree.item(top, open=True)
            self.ftree.selection_set(top)
        self.font_status.set('语言已改为 %s(未保存)' % new)

    def _font_dellang(self):
        from tkinter import messagebox
        if self.font_json is None:
            messagebox.showwarning('font', '尚未加载 json，请先点「加载 json」')
            return
        sel = self.ftree.selection()
        if not sel:
            messagebox.showwarning('font', '请先选中一个语言分组')
            return
        iid = sel[0]
        if iid.startswith('f'):
            li = int(iid[1:].split('_')[0])
        else:
            li = self._font_lang_iid.get(iid)
            if li is None:
                messagebox.showwarning('font', '请先选中一个语言分组')
                return
        item = self.font_json['info'][li]
        lang = item.get('lang', '?')
        if not messagebox.askyesno('删除语言', '确认删除整个语言分组 [%s] 及其全部字库条目?' % lang, parent=self.root):
            return
        del self.font_json['info'][li]
        L = self.font_json.setdefault('lang', [])
        if lang in L and not any(i.get('lang') == lang for i in self.font_json['info']):
            L.remove(lang)
        self.ftree.selection_remove(self.ftree.selection())
        self.font_sel = None
        self._font_rebuild()
        self.font_status.set('已删除语言 %s(未保存)' % lang)

    def _font_modlib(self):
        from tkinter import messagebox
        if self.font_json is None:
            messagebox.showwarning('font', '尚未加载 json，请先点「加载 json」')
            return
        sel = self.ftree.selection()
        if not sel or not sel[0].startswith('f'):
            messagebox.showwarning('font', '请先在左侧选中一个具体字库条目(.bin 那一行)')
            return
        iid = sel[0]
        li, ei = map(int, iid[1:].split('_'))
        item = self.font_json['info'][li]
        self.font_sel = (li, ei)
        name = self._fe['name'].get().strip()
        if name:
            item['name'][ei] = name
        try:
            item['size'][ei] = int(self._fe['size'].get().strip())
        except Exception:
            pass
        def _to_int(s):
            try:
                return int(str(s).strip())
            except Exception:
                return 0
        b = _to_int(self._fe['base'].get()); l = _to_int(self._fe['line'].get())
        item.setdefault('ext', [None] * len(item['name']))
        while ei >= len(item['ext']):
            item['ext'].append(None)
        item['ext'][ei] = {'base_line_ext': b, 'line_height_ext': l}
        # 就地同步树行文本
        self.ftree.item('f%d_%d' % (li, ei), text=item['name'][ei],
                        values=(item['size'][ei],))
        self.font_status.set('已修改字库(未写文件，点「保存 json」同步)')

    def _font_del(self):
        from tkinter import messagebox
        if self.font_json is None:
            messagebox.showwarning('font', '尚未加载 json，请先点「加载 json」')
            return
        sel = self.ftree.selection()
        if not sel:
            messagebox.showwarning('font', '请先在左侧选中一个具体的字库条目(.bin 那一行)')
            return
        iid = sel[0]
        if not iid.startswith('f'):
            messagebox.showwarning('font', '请选中具体的字库条目(展开语言分组后点 .bin 那一行)')
            return
        li, ei = map(int, iid[1:].split('_'))
        item = self.font_json['info'][li]
        if ei < len(item['name']):
            del item['name'][ei]
        if ei < len(item['size']):
            del item['size'][ei]
        if ei < len(item.get('ext', [])):
            del item['ext'][ei]
        # 就地删数组 + 重建树(iid 与数组同步, 且保留展开)
        self.ftree.selection_remove(iid)
        self.font_sel = None
        self._font_rebuild()
        self.font_status.set('已删除(未保存，点「保存 json」同步)')

    def _font_preview(self):
        import tkinter as tk
        from tkinter import scrolledtext
        if self.font_json is None:
            return
        top = tk.Toplevel(self.root)
        top.title('json 预览(副本，未保存的修改在此可见)')
        top.geometry('720x560')
        t = scrolledtext.ScrolledText(top, wrap='none', font=('Consolas', 9), state='normal')
        t.pack(fill='both', expand=True, padx=8, pady=8)
        try:
            s = json.dumps(self.font_json, ensure_ascii=False, indent=4)
        except Exception as e:
            s = '%r' % e
        t.insert('1.0', s)
        t.configure(state='disabled')

    #--------------- image 子界面 ---------------
    def _build_image_tab(self):
        import tkinter as tk
        from tkinter import ttk
        f = ttk.Frame(self.nb, padding=8)
        self.tabs['image'] = f

        # row0: 顶部固定(项目名 + 执行)
        top = ttk.Frame(f); top.pack(fill='x', pady=(2, 4))
        pr = ttk.Frame(top); pr.pack(fill='x')
        ttk.Label(pr, text='image 项目名称:').pack(side='left')
        self.proj_var = tk.StringVar(value='prj')
        ttk.Entry(pr, textvariable=self.proj_var, width=16).pack(side='left', padx=(6, 10))
        ttk.Button(pr, text='执行 image 打包', command=lambda: self._run('image')).pack(side='right')

        # 主体: 水平分栏(左|右)
        imh = ttk.Panedwindow(f, orient='horizontal'); imh.pack(fill='both', expand=True, pady=(4, 0))
        lv = ttk.Frame(imh); imh.add(lv, weight=4)
        rv = ttk.Panedwindow(imh, orient='vertical'); imh.add(rv, weight=3)

        # 左上: 图形配置(固定顶部自然高, 不被下方挤压)
        gf = ttk.LabelFrame(lv, text=' 图形配置 ', padding=(6, 4)); gf.pack(side='top', fill='x')
        self._img_cfg_vars = {}
        gcfg, _ = self._img_cfg_load()
        # 提示语置顶(首复选框上方)
        ttk.Label(gf, text='保存后下次打包生效(scui_pack_tools.json)',
                  foreground='#888').pack(anchor='w', padx=(2, 0), pady=(0, 2))
        ggrid = ttk.Frame(gf); ggrid.pack(fill='x')
        for r, (key, typ, lab, opts) in enumerate(_IMG_CFG_DEFS):
            if typ == 'bool':
                var = tk.BooleanVar(value=bool(gcfg.get(key)))
                ttk.Checkbutton(ggrid, text=lab, variable=var)\
                    .grid(row=r, column=0, sticky='w', padx=(2, 0), pady=2)
            else:
                ttk.Label(ggrid, text=lab).grid(row=r, column=0, sticky='w', padx=(2, 0), pady=2)
                var = tk.StringVar(value=str(gcfg.get(key)))
                ttk.Combobox(ggrid, textvariable=var, values=opts, width=6, state='readonly')\
                    .grid(row=r, column=1, sticky='e', padx=(8, 2), pady=2)
            self._img_cfg_vars[key] = var
        gbar = ttk.Frame(gf); gbar.pack(fill='x')
        ttk.Button(gbar, text='保存', command=self._img_cfg_save_cb).pack(side='left')
        ttk.Button(gbar, text='恢复默认', command=self._img_cfg_reset).pack(side='left', padx=(8, 0))

        # 左下: 图片资源(占剩余空间, 不去挤图形配置)
        lf = ttk.LabelFrame(lv, text=' 图片资源 ', padding=(4, 4)); lf.pack(side='bottom', fill='both', expand=True)
        self.it_canv = tk.Canvas(lf, highlightthickness=0, bg='#ffffff')
        self._img_vsb = ttk.Scrollbar(lf, orient='vertical', command=self._img_scroll_cmd)
        self.it_canv.configure(yscrollcommand=self._img_vsb.set)
        self.it_canv.pack(side='left', fill='both', expand=True)
        self._img_vsb.pack(side='right', fill='y')
        self._i_sel  = None          # 当前选中图片(绝对路径, 高亮)
        self._disp   = []            # 折叠后可见行([depth,name,path,kind,opened])
        self._img_root = None        # 图片源根目录
        self._img_exp = set()        # 已展开的目录(绝对路径)
        self._i_rowH = 22            # 每行像素高(虚拟化绘制)
        self.it_canv.bind('<Configure>', lambda e: self._img_draw())
        self.it_canv.bind('<Enter>', lambda e: self.it_canv.bind_all('<MouseWheel>', self._img_wheel))
        self.it_canv.bind('<Leave>', lambda e: self.it_canv.unbind_all('<MouseWheel>'))
        self.it_canv.bind('<Button-1>', self._img_canvas_click)
        # 延迟到主循环后再填充(大资源集不阻塞首帧窗口)
        self.root.after(80, self._img_reload)

        # 右上: 图片配置(选中图片: 不参与编译/dither/index; 更新配置=重命名)
        icf = ttk.LabelFrame(rv, text=' 图片配置 ', padding=(6, 4)); rv.add(icf, weight=2)
        self._i_sel_path = None
        self.i_sel_label = ttk.Label(icf, text='未选择图片(选中左侧图片生效)', foreground='#888')
        self.i_sel_label.pack(anchor='w', padx=(2, 0))
        self._i_exc = tk.BooleanVar(value=False)
        self._i_dit = tk.BooleanVar(value=False)
        self._i_idx = tk.BooleanVar(value=False)
        ttk.Checkbutton(icf, text='不参与编译(# 前缀)', variable=self._i_exc)\
            .pack(anchor='w', padx=(2, 0), pady=(4, 0))
        ttk.Checkbutton(icf, text='dither(.dit 尾缀)', variable=self._i_dit,
                        command=self._img_dit_on).pack(anchor='w', padx=(2, 0))
        ttk.Checkbutton(icf, text='index(.idx 尾缀)', variable=self._i_idx,
                        command=self._img_idx_on).pack(anchor='w', padx=(2, 0))
        ibar = ttk.Frame(icf); ibar.pack(fill='x', pady=(8, 0))
        ttk.Button(ibar, text='更新配置', command=self._img_iic_apply).pack(side='left')
        ttk.Label(ibar, text='更新配置≈重命名该图片文件', foreground='#888').pack(side='left', padx=(8, 0))

        # 右下: 图片详情(预览 + 解析字段)
        df = ttk.LabelFrame(rv, text=' 图片详情 ', padding=(6, 4)); rv.add(df, weight=3)
        dbody = ttk.Frame(df); dbody.pack(fill='both', expand=True)
        self.preview = ttk.Label(dbody, text='选择左侧图片查看预览', anchor='center',
                                 width=26, relief='groove')
        self.preview.pack(side='left', fill='y', padx=(0, 6))
        self.field_text = tk.Text(dbody, font=('Consolas', 9), state='disabled', wrap='none')
        self.field_text.pack(side='left', fill='both', expand=True)

        self.nb.add(f, text='image  ')

    #--------------- 图形配置持久化(scui_pack_tools.json) ---------------
    def _cfg_json(self):
        return os.path.join(self.tools, 'scui_pack_tools.json')

    def _img_cfg_load(self):
        g = dict(_IMAGE_CFG_DEFAULT)
        try:
            j = json.load(open(self._cfg_json(), encoding='utf-8')) if os.path.isfile(self._cfg_json()) else {}
        except Exception:
            j = {}
        g.update(j.get('image_config', {}))
        return g, j

    def _img_cfg_save(self, g):
        cfg = self._cfg_json()
        try:
            j = json.load(open(cfg, encoding='utf-8')) if os.path.isfile(cfg) else {}
        except Exception:
            j = {}
        j['image_config'] = g
        with open(cfg, 'w', encoding='utf-8', newline='') as fp:
            fp.write(json.dumps(j, ensure_ascii=False, indent=4).replace('\n', '\r\n'))

    def _img_cfg_save_cb(self):
        g = {}
        for key, typ, lab, opts in _IMG_CFG_DEFS:
            v = self._img_cfg_vars[key].get()
            g[key] = bool(v) if typ == 'bool' else int(v or _IMAGE_CFG_DEFAULT[key])
        try:
            self._img_cfg_save(g)
            self._append_log('image', '[cfg] 图形配置已保存: %s\n' % self._rel_log(self._cfg_json()))
        except Exception as e:
            self._wlog('[cfg] 保存图形配置失败: %r' % e)

    def _img_cfg_reset(self):
        for key, typ, lab, opts in _IMG_CFG_DEFS:
            v = _IMAGE_CFG_DEFAULT[key]
            if typ == 'bool':
                self._img_cfg_vars[key].set(bool(v))
            else:
                self._img_cfg_vars[key].set(str(v))

    #--------------- 图片资源(Canvas 行列表) ---------------
    def _img_wheel(self, e):
        self.it_canv.yview_scroll(-1 * (e.delta // 120), 'units')
        self._img_draw()

    # 滚动条 command: 先滚动后重绘(虚拟化没有持久 item, 必须重绘才刷新)
    def _img_scroll_cmd(self, *args):
        self.it_canv.yview(*args)
        self._img_draw()

    def _img_reload(self):
        self.it_canv.delete('all')
        old_exp = getattr(self, '_img_exp', set())
        self._img_root = self.in_abs['image'] if os.path.isdir(self.in_abs['image']) else self.ui
        self._img_tree = self._img_scan_nodes(self._img_root)
        # 保留仍存在的展开目录(重命名/增删后折叠状态不丢)
        self._img_exp = {p for p in old_exp if os.path.isdir(p)} | {self._img_root}
        self._img_reflatten()
        # 若选中图片已不存在(被重命名/删除), 清空配置
        if self._i_sel and not os.path.exists(self._i_sel):
            self._i_sel = None
            self._img_clear_config()

    def _img_scan_nodes(self, path):
        """brief: 扫描目录成嵌套节点
        @param path 目录绝对路径
        @retval 节点列表[{name,path,kind,children}]
        """
        nodes = []
        try:
            items = sorted(os.listdir(path),
                key=lambda x: (not os.path.isdir(os.path.join(path, x)), x.lower()))
        except OSError:
            return nodes
        for e in items:
            full = os.path.join(path, e)
            if os.path.isdir(full):
                nodes.append({'name': e, 'path': full, 'kind': 'dir',
                              'children': self._img_scan_nodes(full)})
            elif e.lower().endswith(_IMG_EXT):
                nodes.append({'name': e, 'path': full, 'kind': 'img', 'children': []})
        return nodes

    def _img_reflatten(self):
        """依据展开状态生成显示序列(虚拟化 + 支持折叠)"""
        # 保存当前可见首行索引, 展开/折叠后保持滚动位置
        old_n = len(self._disp)
        try:
            v0, _ = self.it_canv.yview()
        except Exception:
            v0 = 0.0
        idx_top = int(v0 * old_n) if old_n else 0
        disp = []
        self._img_flatten(self._img_tree, 0, disp)
        self._disp = disp
        n = len(disp)
        self.it_canv.configure(scrollregion=(0, 0, 0, n * self._i_rowH))
        if n:
            # 恢复滚动位置(按首行索引, 超界 clamp; 不满一屏自动归顶)
            self.it_canv.yview_moveto(min(1.0, idx_top / n))
        self._img_draw()

    def _img_flatten(self, nodes, depth, disp):
        for nd in nodes:
            opened = (nd['kind'] == 'dir' and nd['path'] in self._img_exp)
            disp.append([depth, nd['name'], nd['path'], nd['kind'], opened])
            if opened:
                self._img_flatten(nd['children'], depth + 1, disp)

    def _img_toggle_dir(self, path):
        if path in self._img_exp:
            self._img_exp.discard(path)
        else:
            self._img_exp.add(path)
        self._img_reflatten()

    # 虚拟化绘制: 只画可见区间行, 大资源集不一次性建全部控件
    def _img_draw(self):
        c = self.it_canv
        # 滚动条可见性 + 内容不足一屏顶对齐不可滚动
        vh = c.winfo_height() or 200
        self._img_needs_scroll = (len(self._disp) * self._i_rowH > vh)
        if self._img_needs_scroll:
            if not self._img_vsb.winfo_ismapped():
                self._img_vsb.pack(side='right', fill='y')
        else:
            if self._img_vsb.winfo_ismapped():
                self._img_vsb.pack_forget()
            c.yview_moveto(0)
        c.delete('all')
        n = len(self._disp)
        if n == 0:
            return
        v0, v1 = c.yview()
        i0 = max(0, int(v0 * n) - 2)
        i1 = min(n, int(v1 * n) + 2)
        cw = max(c.winfo_width(), 120)
        sel = self._i_sel
        btn_x0 = cw - 42
        for idx in range(i0, i1):
            depth, name, path, kind, opened = self._disp[idx]
            y = idx * self._i_rowH
            x = depth * 14
            exc = name.startswith('#')
            if path == sel:
                c.create_rectangle(0, y, cw, y + self._i_rowH, fill='#ffffcc', outline='')
            if kind == 'dir':
                c.create_text(x + 4, y + self._i_rowH // 2,
                              text='▼' if opened else '▶', font=('Segoe UI Symbol', 8),
                              fill='#555')
            fg = '#999' if exc else ('#2266cc' if kind == 'img' else '#333')
            c.create_text(x + 30, y + self._i_rowH // 2,
                          text='📁' if kind == 'dir' else '🖼',
                          font=('Segoe UI Emoji', 9))
            c.create_text(x + 50, y + self._i_rowH // 2, text=name,
                          font=('Consolas', 9, 'bold') if kind == 'dir' else ('Consolas', 9),
                          fill=fg, anchor='w')
            # 右侧小"启用/禁用"按钮
            c.create_rectangle(btn_x0, y + 3, cw - 2, y + self._i_rowH - 3,
                               outline='#cfcfcf', fill='#eeeeee')
            c.create_text(btn_x0 + (cw - 2 - btn_x0) // 2, y + self._i_rowH // 2,
                          text='启用' if exc else '禁用', font=('Microsoft YaHei UI', 8))

    # Canvas 点击: 命中右侧按钮区则 启用/禁用; 目录则折叠; 图片则选中
    def _img_canvas_click(self, e):
        cy = self.it_canv.canvasy(e.y)     # 转画布逻辑坐标(滚动偏移)
        idx = int(cy // self._i_rowH)
        if not (0 <= idx < len(self._disp)):
            return
        depth, name, path, kind, opened = self._disp[idx]
        cx = self.it_canv.canvasx(e.x)
        in_btn = cx >= (self.it_canv.winfo_width() or 400) - 46
        if in_btn:
            self._img_toggle_exclude(path)
        elif kind == 'dir':
            self._img_toggle_dir(path)
        elif kind == 'img':
            self._img_pick(path)

    # 目录/图片 # 前缀 启用/禁用(=重命名)
    def _img_toggle_exclude(self, path):
        from tkinter import messagebox
        base = os.path.basename(path)
        new = ('#' + base) if not base.startswith('#') else base[1:]
        np_ = os.path.join(os.path.dirname(path), new)
        if np_ != path:
            if os.path.exists(np_):
                messagebox.showwarning('启用/禁用', '目标已存在: %s' % new, parent=self.root); return
            try:
                os.rename(path, np_)
                self._wlog('[img] %s -> %s' % (base, new))
            except OSError as e:
                messagebox.showwarning('启用/禁用', '重命名失败: %r' % e, parent=self.root); return
        self._img_reload()
        # 所选图片整体失效则清空配置
        if self._i_sel and not os.path.exists(self._i_sel):
            self._i_sel = None
            self._img_clear_config()

    # 点图: 高亮 + 载入图片配置 + 详情
    def _img_pick(self, path):
        self._i_sel = path
        self._img_draw()
        self._img_iic_load(path)
        self._show_img(path)

    #--------------- 图片配置(选中图片) ---------------
    def _img_dit_on(self):
        if self._i_dit.get():
            self._i_idx.set(False)

    def _img_idx_on(self):
        if self._i_idx.get():
            self._i_dit.set(False)

    def _img_iic_load(self, path):
        stem, ext, exc, idx, dit = _img_mkbar_split(os.path.basename(path))
        self._i_sel_path = path
        self._i_exc.set(exc)
        self._i_idx.set(idx)
        self._i_dit.set(dit)
        self.i_sel_label.config(text='已选图片: %s' % os.path.basename(path))

    def _img_iic_apply(self):
        from tkinter import messagebox
        if not (self._i_sel_path and os.path.isfile(self._i_sel_path)):
            messagebox.showinfo('更新配置', '请先在左侧选择一张图片', parent=self.root); return
        d = os.path.dirname(self._i_sel_path)
        base = os.path.basename(self._i_sel_path)
        stem, ext, exc, idx, dit = _img_mkbar_split(base)
        newexc = self._i_exc.get()
        newidx = self._i_idx.get()
        newdit = self._i_dit.get()
        if newidx and newdit:          # 互斥: index 优先
            newdit = False
            self._i_dit.set(False)
        newname = _img_mkbar_join(stem, ext, newexc, newidx, newdit)
        if newname == base:
            messagebox.showinfo('更新配置', '配置无变化', parent=self.root); return
        np_ = os.path.join(d, newname)
        if os.path.exists(np_):
            messagebox.showwarning('更新配置', '目标文件已存在: %s' % newname, parent=self.root); return
        try:
            os.rename(self._i_sel_path, np_)
        except OSError as e:
            messagebox.showwarning('更新配置', '重命名失败: %r' % e, parent=self.root); return
        self._wlog('[img] %s -> %s' % (base, newname))
        self._i_sel = np_
        self._img_reload()
        self._img_iic_load(np_)
        self._show_img(np_)

    def _img_clear_config(self):
        self._i_sel_path = None
        self.i_sel_label.config(text='未选择图片(选中左侧图片生效)')
        self.preview.configure(image='', text='')
        self.field_text.configure(state='normal')
        self.field_text.delete('1.0', 'end')
        self.field_text.configure(state='disabled')

    def _show_img(self, path):
        import tkinter as tk
        from PIL import Image, ImageTk
        preview_err = None
        # 预览图无条件显示
        try:
            im = Image.open(path)
            im.load()
            im.thumbnail((180, 180))
            ph = ImageTk.PhotoImage(im)
            self._imgs.append(ph)
            self.preview.configure(image=ph, text='')
        except Exception as e:
            preview_err = '%r' % e
            self.preview.configure(image='', text=os.path.basename(path))

        # 构造 file_tag(与打包后端一致的无尾缀句柄: 去扩展名/去.idx/.dit/去点)
        file_tag = _img_handle_tag(path, self.in_abs['image'])

        lines = []
        if preview_err:
            lines.append('【预览失败】 %s' % preview_err)
            lines.append('')
        # 匹配 .h 句柄信息
        h_info = self._image_from_h(file_tag)
        if h_info:
            lines.append('【句柄信息（scui_image_parser.h）】')
            lines.append('    %-18s = %s;' % ('handle_name', h_info['name']))
            lines.append('    %-18s = %s;' % ('handle_value', h_info['value']))
            lines.append('    %-18s = %s;' % ('pixel.width', h_info['width']))
            lines.append('    %-18s = %s;' % ('pixel.height', h_info['height']))
            lines.append('    %-18s = %s;' % ('pixel.size_bin', h_info['size_bin']))
            lines.append('    %-18s = %s;' % ('pixel.size_mem', h_info['size_mem']))
            lines.append('    %-18s = %s;' % ('com_pct', h_info['com_pct']))
            lines.append('')
        # 匹配 .c 结构体信息
        c_info = self._image_from_c(file_tag)
        if c_info:
            lines.append('【资源信息（scui_image_parser.c）】')
            for k, v in c_info.items():
                lines.append('    %-18s = %s;' % (k, v))
            lines.append('')
        # 都没匹配到
        if not h_info and not c_info:
            lines.append('【未匹配到打包记录 — 请先执行 image 打包】')
            lines.append('    (file_tag: %s)' % file_tag)

        self.field_text.configure(state='normal')
        self.field_text.delete('1.0', 'end')
        self.field_text.insert('1.0', '\n'.join(lines))
        self.field_text.configure(state='disabled')

    def _image_from_h(self, file_tag):
        """从 scui_image_parser.h 匹配句柄信息(file_tag 为去点后的相对路径)."""
        out_h = os.path.join(self.out_abs.get('image', ''), 'scui_image_parser.h')
        if not os.path.exists(out_h):
            return None
        try:
            with open(out_h, 'r', encoding='utf-8', errors='replace') as hf:
                content = hf.read()
            # 枚举项以 file_tag 结尾: scui_image_<prefix><file_tag>, // 0xNNNN
            re_enum = r'(scui_image_[A-Za-z0-9_]*%s)\s*,\s*//\s*(0x[0-9a-fA-F]+)' % re.escape(file_tag)
            m = re.search(re_enum, content)
            if not m:
                return None
            name = m.group(1)
            value = m.group(2)
            # 压缩注释行也以同后缀结尾: //< w,h,size_bin,size_mem,com_pct> scui_image_<prefix><file_tag>
            re_comment = r'//<\s*([^>]+)>\s*scui_image_[A-Za-z0-9_]*%s\b' % re.escape(file_tag)
            m2 = re.search(re_comment, content)
            if not m2:
                return None
            # 注释里为十六进制径 w,h,size_bin,size_mem(0x...)
            nums = re.findall(r'0x([0-9a-fA-F]+)', m2.group(1))
            ws = [int(x, 16) for x in nums]
            if len(ws) < 4:
                return None
            w, h, bin_, mem = ws[0], ws[1], ws[2], ws[3]
            return {
                'name': name,
                'value': value,
                'width': w,
                'height': h,
                'size_bin': bin_,
                'size_mem': mem,
                'com_pct': '%.4f' % (float(bin_) / mem) if mem else 1.0,
            }
        except OSError:
            return None

    def _image_from_c(self, file_tag):
        """从 scui_image_parser.c 匹配 scui_image_t 结构体."""
        out_c = os.path.join(self.out_abs.get('image', ''), 'scui_image_parser.c')
        if not os.path.exists(out_c):
            return None
        try:
            with open(out_c, 'r', encoding='utf-8', errors='replace') as cf:
                content = cf.read()
            # 结构体名以 file_tag 结尾(无 scui_image_ 前缀): static const scui_image_t <prefix><file_tag>
            re_struct = r'static\s+const\s+scui_image_t\s+([A-Za-z0-9_]*%s)\s*=\s*\{([^}]+)\}' % re.escape(file_tag)
            m = re.search(re_struct, content, re.DOTALL)
            if not m:
                return None
            body = m.group(2)   # group(1)=结构体名, group(2)={...}体
            info = {}
            # 解析各字段(脚本生成以 ','/';' 结尾)
            for line in body.split('\n'):
                line = line.strip()
                if '=' in line and line[-1] in (';', ','):
                    k, v = line.split('=', 1)
                    k = k.strip().lstrip('.')
                    v = v.strip().rstrip(';').rstrip(',').strip()
                    info[k] = v
            return info
        except OSError:
            return None

    #--------------- 执行打包(子线程, UI 不卡) ---------------
    def _run(self, name):
        if getattr(self, '_busy', False):
            return
        self._busy = True
        self._status('执行中…(%s)' % name)
        src = self.in_abs[name]
        dst = self.out_abs[name]
        proj = getattr(self, 'proj_var', None)
        proj = proj.get() if proj else 'prj'
        # 打包前清空全局日志缓冲, 并弹出共享日志窗口
        self._log_text = ''
        if self._log_win is not None:
            t = getattr(self._log_win, '_text', None)
            if t is not None:
                try:
                    t.configure(state='normal')
                    t.delete('1.0', 'end')
                    t.configure(state='disabled')
                except Exception:
                    pass
        self._log_show()
        self._append_log(name, '==== start: %s ====\n' % name)

        self._logq = queue.Queue()
        wire = _TabLog(self._logq)

        def worker():
            so, se = sys.stdout, sys.stderr
            si = sys.stdin
            sys.stdout = wire
            sys.stderr = wire
            sys.stdin = _NoStdin()          # 无控制台 exe 下避免 input() 阻塞
            try:
                _do_task(name, self.ui, src, dst, proj)
            except Exception:
                import traceback as _tb
                print('[pack] 脚本运行异常:\n%s' % _tb.format_exc())
            finally:
                sys.stdout, sys.stderr, sys.stdin = so, se, si
                self._logq.put('__DONE__')

        threading.Thread(target=worker, daemon=True).start()

    def _append_log(self, name, txt):
        # 所有子界面日志共用一份全局内存缓冲 + 全局日志窗口
        del name
        self._log_text += txt
        if self._log_win is not None:
            t = getattr(self._log_win, '_text', None)
            if t is not None:
                try:
                    t.configure(state='normal')
                    t.insert('end', txt)
                    t.see('end')
                    t.configure(state='disabled')
                except Exception:
                    pass

    def _log_show(self):
        # 打开/聚焦全局日志窗口(与菜单「日志」共用一个窗口)
        import tkinter as tk
        if self._log_win is not None:
            try:
                self._log_win.deiconify()
                self._log_win.lift()
                return
            except Exception:
                self._log_win = None
        self._log_win = tk.Toplevel(self.root)
        self._log_win.title('运行日志(全局共享)')
        self._log_win.geometry('820x560')
        from tkinter import scrolledtext
        t = scrolledtext.ScrolledText(self._log_win, wrap='word', font=('Consolas', 9), state='disabled')
        t.pack(fill='both', expand=True, padx=8, pady=8)
        self._log_win._text = t
        static = self._log_text if self._log_text.strip() else '(暂无日志, 请先在任意子界面执行打包)'
        t.configure(state='normal')
        t.insert('1.0', static)
        t.see('end')
        t.configure(state='disabled')
        self._log_win.protocol('WM_DELETE_WINDOW', self._log_hide)

    def _log_hide(self):
        if self._log_win is not None:
            self._log_win.withdraw()

    def _poll(self):
        if self._logq is not None:
            # 每帧限流 + 整批合并: 只做一次 insert + 一次 see,
            # 避免海量日志逐条刷新(每条 see/重布局)拖垮主线程导致卡顿
            chunks = []
            n = 0
            while n < 4000:
                try:
                    txt = self._logq.get_nowait()
                except queue.Empty:
                    break
                n += 1
                if txt == '__DONE__':
                    self._flush_log(chunks)
                    chunks = []
                    self._busy = False
                    self._status('完成')
                    continue
                chunks.append(txt)
            self._flush_log(chunks)
        self.root.after(120, self._poll)

    def _flush_log(self, chunks):
        # 批量写入全局日志缓冲, 并一次性刷新日志窗口(running in main thread)
        if not chunks:
            return
        joined = ''.join(chunks)
        self._log_text += joined
        t = getattr(self._log_win, '_text', None) if self._log_win is not None else None
        if t is None:
            return
        try:
            t.configure(state='normal')
            t.insert('end', joined)
            t.see('end')
            t.configure(state='disabled')
        except Exception:
            pass

    def _status(self, s):
        self.root.title('scui资源工具 - %s' % s)

    def _on_tab(self, ev):
        idx = self.nb.index(self.nb.select())
        names = ['widget', 'image', 'font', 'lang', 'cwf']
        self._current_name = names[idx] if idx < len(names) else 'widget'

    def _about(self):
        from tkinter import messagebox
        messagebox.showinfo('scui资源工具',
            '作者：Agent\n'
            '版本：Ver 0.0.1\n'
            '状态：持续开发中...\n'
            '类型：widget/image/font/lang/cwf\n'
            '来源：scui/tools & scui/plugs')

    # 工具「日志」菜单 -> 打开/聚焦全局共享日志窗口
    def _open_logs(self):
        self._log_show()

class _TabLog(object):
    """把 print 写入的文本重定向进指定队列(保留换行), 由 GUI 日志泵写日志区."""
    def __init__(self, q):
        self.q = q
    def write(self, s):
        if s:
            self.q.put(s)
    def flush(self):
        pass

class _NoStdin(object):
    """无控制台环境下替换 stdin, 使 input() 立即抛 EOFError 而非阻塞."""
    def read(self, *a, **k):
        raise EOFError
    def readline(self, *a, **k):
        raise EOFError
    def readlines(self, *a, **k):
        raise EOFError
    def __iter__(self):
        return self
    def __next__(self):
        raise StopIteration

#============================================================
# 入口
#============================================================
def _single_instance():
    """brief: GUI 单实例锁(命名互斥体), 已存在实例返回 False
    @retval True=可启动, False=已有实例
    """
    try:
        import ctypes
        from ctypes import wintypes
        k32 = ctypes.WinDLL('kernel32', use_last_error=True)
        k32.CreateMutexW.restype = wintypes.HANDLE
        k32.CreateMutexW.argtypes = [wintypes.LPVOID, wintypes.BOOL, wintypes.LPCWSTR]
        k32.CreateMutexW(None, False, 'Global\\scui_pack_tools_singleinstance')
        return ctypes.get_last_error() != 183      # ERROR_ALREADY_EXISTS
    except Exception:
        return True                                # 非 Windows/异常则放行

def main():
    # 无控制台(exe)下 stdout/stderr 为 None, 替换为 devnull 避免意外崩溃
    if sys.stdout is None:
        sys.stdout = open(os.devnull, 'w')
    if sys.stderr is None:
        sys.stderr = open(os.devnull, 'w')
    # 输出统一 UTF-8, 兼容中文日志(Windows 控制台默认 gbk)
    for _s in (sys.stdout, sys.stderr):
        try:
            _s.reconfigure(encoding='utf-8', errors='replace')
        except Exception:
            pass

    # 轻量手动解析: 首个位置参数作为类型(须为已知类型); 未知参数一律忽略, 回退 GUI
    argv = sys.argv[1:]
    typ = ui = src = dst = proj = ''
    i = 0
    while i < len(argv):
        a = argv[i]
        if a == '--ui' and i + 1 < len(argv):
            ui = argv[i + 1]; i += 2
        elif a == '--src' and i + 1 < len(argv):
            src = argv[i + 1]; i += 2
        elif a == '--dst' and i + 1 < len(argv):
            dst = argv[i + 1]; i += 2
        elif a == '--proj' and i + 1 < len(argv):
            proj = argv[i + 1]; i += 2
        elif a.startswith('--'):
            i += 1                            # 未知选项忽略
        else:
            if typ == '' and a in _TASK:
                typ = a                       # 仅首个位置参数可为已知类型
            i += 1                            # 其它位置参数(如误拖入的路径)忽略

    ui_root = os.path.normpath(ui or os.getcwd())
    if typ:
        sdef, ddef = _DEFAULT[typ]
        src = src or os.path.join(ui_root, sdef)
        dst = dst or os.path.join(ui_root, ddef)
        return _do_task(typ, ui_root, src, dst, proj)
    # GUI 单实例: 已有实例运行则提示并退出
    if not _single_instance():
        try:
            import ctypes
            ctypes.windll.user32.MessageBoxW(
                None, 'scui_pack_tools 已在运行，仅允许一个实例。', '提示', 0x40)
        except Exception:
            pass
        return 0
    return _launch_gui(ui_root)

if __name__ == '__main__':
    sys.exit(main())