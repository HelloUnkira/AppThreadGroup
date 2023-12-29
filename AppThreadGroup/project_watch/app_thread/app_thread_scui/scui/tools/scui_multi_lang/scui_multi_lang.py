import os
import os.path
import json
import openpyxl
import ctypes
# -- coding: utf-8 --**


# for row in xlsx_sheet.rows: # 获取每一行的数据
#     for data in row:        # 获取每一行中单元格的数据
#         print(data.value)  # 打印单元格的值
# print(xlsx_sheet.cell(1, 1).value)


# 编写集成化源文件
def encode_scui_multi_lang_c(file, xlsx_sheet, sheet_row, sheet_col):
    # 写点简要的说明
    file.write('/*本地静态的字符串表\n')
    file.write(' *通过scui_multi_lang.py生成\n */\n\n')
    file.write('#include "app_ext_lib.h"\n')
    file.write('#include "app_sys_lib.h"\n')
    file.write('#include "scui.h"\n\n')
    # 提取所有外源依赖
    file.write('static const char * scui_multi_lang_table[%d][%d] = {\n' % (sheet_row, sheet_col))
    for item in xlsx_sheet.rows:
        file.write('\t{')
        for data in item:
            c_array = str([c for c in str(data.value).encode('utf-8')])
            c_array = c_array.replace('[', '{')
            c_array = c_array.replace(']', ', 0}')
            c_array = c_array.replace('{', '(char []){')
            file.write('{0}'.format("%s," % c_array))    # 多国语这里不能很好的对齐,因为不同语言的空格宽度不一致
        file.write('},\n')
    file.write('};\n\n')
    file.write('static uint32_t scui_multi_lang_type = 0;\n\n')
    # 编写固化访问函数
    file.write('/*@brief 设置搜索语言\n')
    file.write(' *@param index语言编号(0~n-1)\n */\n')
    file.write('void scui_multi_lang_type_config(uint32_t type)\n')
    file.write('{\n\tscui_multi_lang_type = type;\n}\n\n')
    file.write('/*@brief 获得多国语字符串\n')
    file.write(' *@param index字符串编号(0~n-1)\n */\n')
    file.write('const char * scui_multi_lang_str_find(uint32_t index)\n')
    file.write('{\n')
    file.write('\tif (index < SCUI_MULTI_LANG_NUM)\n')
    file.write('\t\treturn scui_multi_lang_table[index][scui_multi_lang_type];\n')
    file.write('\treturn NULL;\n')
    file.write('}\n')


# 编写集成化头文件
def encode_scui_multi_lang_h(file, xlsx_sheet, sheet_row, sheet_col):
    file.write('#ifndef SCUI_MULTI_LANG_H\n')
    file.write('#define SCUI_MULTI_LANG_H\n\n')
    # 写点简要的说明
    file.write('/*本地静态的字符串表\n')
    file.write(' *通过scui_multi_lang.py生成\n */\n\n')
    # 编写头部索引
    file.write('typedef enum {\n')
    for i, item in enumerate(xlsx_sheet.rows):
        c_str = str(item[0].value).replace('\n', '\\n')
        file.write('\tSCUI_MULTI_LANG_0X%04x,\t/* %s */\n' % (i, c_str))
    file.write('\tSCUI_MULTI_LANG_NUM,\n')
    file.write('} scui_multi_lang_t;\n\n')
    # 编写固化访问函数接口
    file.write('/*@brief 设置搜索语言\n')
    file.write(' *@param index语言编号(0~n-1)\n */\n')
    file.write('void scui_multi_lang_type_config(uint32_t type);\n\n')
    file.write('/*@brief 获得多国语字符串\n')
    file.write(' *@param index字符串编号(0~n-1)\n */\n')
    file.write('const char * scui_multi_lang_str_find(uint32_t index);\n\n')
    file.write('#endif\n')


# 启用集成化资源总表生成
def encode_scui_multi_lang():
    # json转Python字符串并转标准字典
    json_file = open('scui_multi_lang.json', 'r')
    json_dict = json.loads(json_file.read())
    json_file.close()
    # 仅支持批量化字符串表管理
    if json_dict['type'] != 'scui_multi_lang':
        return
    # 从标准字典获取工作簿名字和单元
    xlsx_file = openpyxl.load_workbook(json_dict['xlsx'], read_only=True)
    xlsx_sheet = xlsx_file[json_dict['sheet']]
    sheet_row = xlsx_sheet.max_row
    sheet_col = xlsx_sheet.max_column
    # 开启三个文件
    file_encode = 'utf-8'
    scui_multi_lang_h = open('scui_multi_lang.h', mode='w', encoding=file_encode)
    scui_multi_lang_c = open('scui_multi_lang.c', mode='w', encoding=file_encode)
    # 解析
    encode_scui_multi_lang_h(scui_multi_lang_h, xlsx_sheet, sheet_row, sheet_col)
    encode_scui_multi_lang_c(scui_multi_lang_c, xlsx_sheet, sheet_row, sheet_col)
    # 关闭三个文件
    scui_multi_lang_h.close()
    scui_multi_lang_c.close()
    # 关闭工作簿
    xlsx_file.close()


if __name__ == '__main__':
    try:
        encode_scui_multi_lang()
    except Exception as e:
        print
    print('scui multi lang finish')
    input('请按任意键退出...')
