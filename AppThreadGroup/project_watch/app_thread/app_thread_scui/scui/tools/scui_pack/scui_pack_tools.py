# 资源打包统一前端(GUI + 命令行)
# 位于 scui/tools/scui_pack/; 引用相对父路径 scui/tools 取用其它打包工具
# 支持: image / font / lang / cwf 四类
# 无参       -> 启动图形界面(顶部固定信息 + 四个子界面)
# 带参型     -> 命令行模式: python scui_pack_tools.py <type> [--src P] [--dst P] [--proj N]
# 说明: GUI 运行时把控制台 stdout/stderr 打印重定向输出到子界面 log 区;
#       运行产生的临时/缓存一律落在 ui 环境, 不在 tools 内引入 cache 与子文件
import os
import re
import json
import sys
import importlib
import shutil
import queue
import threading

#============================================================
# 路径定位(基于 ui 路径推导, 兼容 dev 与 PyInstaller 单文件 exe)
#   ui 目录(scui_ui_res)的父级为 app_thread_scui
#   -> scui/tools(打包脚本)  scui/plugs(cwf 脚本)
#   注意: 不能用 __file__ 推导(exe 运行时指向临时解包目录)
#============================================================
def _dirs(ui):
    ui   = os.path.normpath(ui)
    app  = os.path.dirname(ui)                      # app_thread_scui
    scui = os.path.join(app, 'scui')
    tools= os.path.join(scui, 'tools')
    plugs= os.path.join(scui, 'plugs')
    return ui, app, scui, tools, plugs

def _ui_root(ui):
    return os.path.normpath(ui or os.getcwd())

#============================================================
# 任务定义
#============================================================
_TASK = {
    'widget': {'module': 'scui_widget_parser', 'entry': 'scui_widget_parser',
               'as_main': ['scui_widget_parser.py'], 'dir': 'tools'},
    'image' : {'module': 'scui_image_parser', 'entry': 'scui_image_parser',
               'as_main': ['scui_image_parser.py'], 'dir': 'tools'},
    'font'  : {'module': 'scui_font_package', 'entry': 'scui_font_package',
               'as_main': ['scui_font_package.py'], 'dir': 'tools'},
    'lang'  : {'module': 'scui_lang_parser', 'entry': 'encode_scui_lang_parser',
               'as_main': ['scui_lang_parser.py'], 'dir': 'tools'},
    'cwf'   : {'module': 'scui_cwf_json_parser', 'entry': 'scui_cwf_json_parser',
               'as_main': ['scui_cwf_json_parser.py'], 'dir': 'plugs'},
}

# ui 默认子路径(src, dst)
_DEFAULT = {
    'widget': ('scene_src', 'scene_out'),
    'image' : ('image_src', 'image_out'),
    'font'  : ('font_bin',  'font_out'),
    'lang'  : ('lang_src',  'lang_out'),
    'cwf'   : ('cwf_json',  'cwf_json/bin'),
}

_IMG_EXT = ('.bmp', '.jpg', '.jpeg', '.png', '.gif', '.lottie.json', '.mp4')

#============================================================
# 脚本模块调用(注入 argv, 直接 import 复用主函数)
#============================================================
def _run_module(task, argv, ui):
    ui, app, scui, tools, plugs = _dirs(ui)
    script_dir = tools if task['dir'] == 'tools' else plugs
    if script_dir not in sys.path:
        sys.path.insert(0, script_dir)
    argv_old = sys.argv
    sys.argv = [task['as_main'][0]] + argv
    try:
        mod   = importlib.import_module(task['module'])
        if task['module'] == 'scui_widget_parser':
            mod.SCUI_WIDGET_TOOLS = tools     # 注入真实 tools 目录(打包环境下 __file__ 指向临时解包目录)
        entry = getattr(mod, task['entry'])
        entry()
        return 0
    except ImportError as e:
        print('[pack] 缺模块: %s' % e)
        print('[pack] 请确保依赖已安装(image需 pillow/lz4; image.7z解压需 py7zr)')
        return 1
    except Exception as e:
        print('[pack] 执行异常: %r' % e)
        return 1
    finally:
        sys.argv = argv_old
        shutil.rmtree(os.path.join(tools, '__pycache__'), ignore_errors=True)
        shutil.rmtree(os.path.join(plugs, '__pycache__'), ignore_errors=True)

def _image_src_unpack(src, tag):
    if os.path.isdir(src):
        return src, None
    if src.lower().endswith('.7z'):
        tmp = os.path.join(os.path.dirname(src), tag + '.src_tmp')
        try:
            import py7zr
            shutil.rmtree(tmp, ignore_errors=True)
            with py7zr.SevenZipFile(src, 'r') as z:
                z.extractall(tmp)
            return tmp, tmp
        except ImportError:
            print('[pack] 解压 image.7z 需要 py7zr: pip3 install py7zr')
            return None, None
    return None, None

def _do_task(name, ui, src, dst, proj):
    task   = _TASK[name]
    src    = os.path.normpath(src)
    dst    = os.path.normpath(dst)

    print()
    print('[pack] type : %s' % name)
    print('[pack] src  : %s' % src)
    print('[pack] dst  : %s' % dst)

    tmp = None
    if name == 'image':
        src, tmp = _image_src_unpack(src, 'image')
        if src is None:
            return 1
        argv = [src, dst, proj or 'prj']
    elif name == 'widget':
        argv = [src, dst, os.path.join(dst, 'scui_ui_maker.json')]
    elif name == 'cwf':
        return _do_cwf_task(ui, src, dst)
    else:
        argv = [src, dst]

    if not os.path.exists(src):
        print('[pack] src not exist: %s' % src)
        return 1
    if not os.path.isdir(dst):
        os.makedirs(dst, exist_ok=True)

    print('[pack] argv : %s' % argv)
    ret = _run_module(task, argv, ui)
    if tmp:
        shutil.rmtree(tmp, ignore_errors=True)
    print()
    print('[pack] %s 执行完成: %s' % (name, 'OK' if ret == 0 else 'FAIL'))
    return ret

#============================================================
# cwf 多步打包(解压7z -> image parser -> cwf parser -> 清理)
#============================================================
def _do_cwf_task(ui, src, dst):
    ui_dir, app, scui, tools, plugs = _dirs(ui)
    img_task = _TASK['image']
    cwf_task = _TASK['cwf']
    if not os.path.isdir(src):
        print('[pack] cwf src not exist: %s' % src)
        return 1
    if not os.path.isdir(dst):
        os.makedirs(dst, exist_ok=True)

    # 收集所有 cwf 目录(有 image.7z + 同名 .json)
    cwf_list = []
    for name in sorted(os.listdir(src)):
        d = os.path.join(src, name)
        if not os.path.isdir(d):
            continue
        if os.path.exists(os.path.join(d, 'image.7z')) and \
           os.path.exists(os.path.join(d, name + '.json')):
            cwf_list.append(name)

    if not cwf_list:
        print('[pack] cwf: 未找到有效的 cwf 目录(需含 image.7z + 同名 .json)')
        return 1

    print('[pack] cwf list: %s' % ', '.join(cwf_list))
    print()

    # 确保工具路径在 sys.path
    if tools not in sys.path:
        sys.path.insert(0, tools)
    if plugs not in sys.path:
        sys.path.insert(0, plugs)

    failed = []
    for wf in cwf_list:
        wf_dir = os.path.join(src, wf)
        print('=========== [%s] start ============' % wf)
        # step1: 解压 image.7z -> cwf 目录(7z 内含 image/ 子目录, 不能再嵌套)
        img_dir = os.path.join(wf_dir, 'image')
        shutil.rmtree(img_dir, ignore_errors=True)
        try:
            import py7zr
            with py7zr.SevenZipFile(os.path.join(wf_dir, 'image.7z'), 'r') as z:
                z.extractall(wf_dir)                    # 解压到 cwf 目录, 7z 自带 image/
            print('[%s] unzip image.7z -> OK' % wf)
        except Exception as e:
            print('[%s] unzip fail: %r' % (wf, e))
            failed.append(wf)
            continue

        # step2: image parser (image/ -> .)
        print('[%s] step1 image parser...' % wf)
        argv_old = sys.argv
        sys.argv = ['scui_image_parser.py', img_dir, wf_dir, 'cwf']
        try:
            mod = importlib.import_module(img_task['module'])
            mod.scui_image_parser()
        except Exception as e:
            print('[%s] image parser fail: %r' % (wf, e))
            failed.append(wf)
        finally:
            sys.argv = argv_old

        # step3: cwf json parser (. -> bin/)
        print('[%s] step2 cwf json parser...' % wf)
        sys.argv = ['scui_cwf_json_parser.py', wf_dir, dst]
        try:
            mod = importlib.import_module(cwf_task['module'])
            mod.scui_cwf_json_parser()
        except Exception as e:
            print('[%s] cwf parser fail: %r' % (wf, e))
            failed.append(wf)
        finally:
            sys.argv = argv_old

        # step4: 清理临时文件
        shutil.rmtree(img_dir, ignore_errors=True)
        shutil.rmtree(os.path.join(wf_dir, 'image_array'), ignore_errors=True)
        for f in ('scui_image_parser.h', 'scui_image_parser.c',
                  'scui_image_parser.bin'):
            p = os.path.join(wf_dir, f)
            if os.path.exists(p):
                os.remove(p)
        # cwf parser 临时文件(相对 CWD)
        for f in ('scui_cwf_json_parser.tmp.json.bin',
                  'scui_cwf_json_parser.tmp.image_info.bin',
                  'scui_cwf_json_parser.tmp.image_data.bin'):
            if os.path.isfile(f):
                os.remove(f)
        prog = os.path.join(dst, wf + '_json.prog')
        if os.path.exists(prog):
            os.remove(prog)
        # 清理 __pycache__
        shutil.rmtree(os.path.join(tools, '__pycache__'), ignore_errors=True)
        shutil.rmtree(os.path.join(plugs, '__pycache__'), ignore_errors=True)
        print('=========== [%s] finish ============' % wf)
        print()

    if failed:
        print('[pack] cwf FAILED: %s' % ', '.join(failed))
        return 1
    print('[pack] cwf 执行完成: OK')
    return 0

#============================================================
# 图像信息探测(对齐 scui_image_t 结构体字段)
#============================================================
def _img_type_of(path):
    lo = path.lower()
    if lo.endswith('.gif'):           return 'scui_image_type_gif'
    if lo.endswith('.lottie.json'):   return 'scui_image_type_lottie'
    if lo.endswith('.mp4'):           return 'scui_image_type_mp4'
    if lo.endswith(('.jpg', '.jpeg')):return 'scui_image_type_jpg'
    if lo.endswith('.png'):           return 'scui_image_type_png'
    return 'scui_image_type_bmp'

def _image_probe(path):
    info = {'path': path, 'name': os.path.basename(path)}
    info['type'] = _img_type_of(path)
    # size
    w = h = 0
    try:
        import PIL.Image as _img
        im = _img.open(path); w, h = im.size; im.close()
    except Exception:
        pass
    info['pixel.width']  = w
    info['pixel.height'] = h
    info['pixel.size_bin'] = os.path.getsize(path)
    # format & 每像素内存字节
    fmt = 'scui_pixel_cf_none'; pb = 0
    if info['type'] == 'scui_image_type_bmp': fmt, pb = 'scui_pixel_cf_bmp565', 2
    elif info['type'] == 'scui_image_type_jpg': fmt, pb = 'scui_pixel_cf_bmp565', 2
    elif info['type'] == 'scui_image_type_png': fmt, pb = 'scui_pixel_cf_bmp8565', 3
    info['pixel.format'] = fmt
    info['pixel.size_mem'] = w * h * pb
    info['com_pct'] = (float(info['pixel.size_bin']) / info['pixel.size_mem']) if info['pixel.size_mem'] else 1.0
    return info

#============================================================
# GUI
#============================================================
def _launch_gui(ui):
    try:
        import tkinter as tk
    except Exception as e:
        print('[pack] 无图形环境, 请使用命令行模式: %s' % e)
        return 1
    app = PackApp(ui)
    app.root.mainloop()
    return 0

class PackApp(object):
    def __init__(self, ui):
        import tkinter as tk
        from tkinter import ttk
        self.ui  = ui
        self.ui, self.app, self.scui, self.tools, self.plugs = _dirs(ui)
        self._imgs = []          # 保持 PhotoImage 引用
        self._pv  = {}           # (name,side) -> StringVar
        self._current_name = 'image'
        self._logq = None
        self._log_win = None    # 全局共享日志窗口(Toplevel)
        self._log_text = ''     # 日志内存缓冲(所有子界面共用)

        self.in_abs = {}; self.out_abs = {}
        for k, (si, so) in _DEFAULT.items():
            self.in_abs[k]  = os.path.join(ui, si)
            self.out_abs[k] = os.path.join(ui, so)

        self.root = tk.Tk()
        self.root.title('scui资源工具 - 正在开发中...')
        self.root.geometry('980x720')
        self.root.minsize(860, 620)

        menubar = tk.Menu(self.root)
        m_set = tk.Menu(menubar, tearoff=0)
        m_set.add_command(label='路径…', command=self._open_paths)
        menubar.add_cascade(label='设置', menu=m_set)
        m_log = tk.Menu(menubar, tearoff=0)
        m_log.add_command(label='打开 out/err 浏览…', command=self._open_logs)
        menubar.add_cascade(label='日志', menu=m_log)
        m_help = tk.Menu(menubar, tearoff=0)
        m_help.add_command(label='关于', command=self._about)
        menubar.add_cascade(label='帮助', menu=m_help)
        self.root.config(menu=menubar)

        self._build_header()
        self._build_notebook()

        self.root.protocol('WM_DELETE_WINDOW', self.root.destroy)
        self.root.after(120, self._poll)

    #--------------- 顶部信息条(只读, 无按钮) ---------------
    def _build_header(self):
        from tkinter import ttk
        bar = ttk.Frame(self.root, padding=(12, 6))
        bar.pack(fill='x')
        ttk.Label(bar, text='scui_ui: %s' % self.ui, foreground='#777').pack(side='left')
        self.path_status = ttk.Label(bar, text='', foreground='#888')
        self.path_status.pack(side='left', padx=(14, 0))

    #--------------- 路径设置子弹窗(显示所有路径, 前两项只读, 其余可改) ---------------
    def _open_paths(self):
        import tkinter as tk
        from tkinter import ttk, messagebox
        try:
            self._open_paths_dialog()
        except Exception as e:
            messagebox.showwarning('路径设置', '打开路径设置失败:\n%r' % e, parent=self.root)

    def _open_paths_dialog(self):
        import tkinter as tk
        from tkinter import ttk, filedialog
        top = tk.Toplevel(self.root)
        top.title('路径设置')
        top.geometry('620x600')
        top.transient(self.root)
        f = ttk.Frame(top, padding=(12, 10)); f.pack(fill='both', expand=True)
        ttk.Label(f, text='scui 路径: %s' % self.scui, font=('Consolas', 9)).pack(anchor='w')
        ttk.Label(f, text='scui_ui 路径: %s' % self.ui, font=('Consolas', 9)).pack(anchor='w', pady=(2, 8))

        grid = ttk.Frame(f); grid.pack(fill='x')
        varman = {}
        def _browse(entryvar):
            abs_p = filedialog.askdirectory(title='选择路径', initialdir=entryvar.get() or self.ui)
            if abs_p:
                entryvar.set(self._rel(abs_p))     # 显示为相对
        r = 0
        for name in ('widget', 'image', 'font', 'lang', 'cwf'):
            for side, lab in (('in', '输入'), ('out', '输出')):
                cur = self.in_abs[name] if side == 'in' else self.out_abs[name]
                ttk.Label(grid, text='%s %s:' % (name, lab), width=14, anchor='e')\
                    .grid(row=r, column=0, padx=(0, 8), pady=3)
                var = tk.StringVar(value=self._rel(cur))   # 子类型显示相对路径
                e = ttk.Entry(grid, textvariable=var, font=('Consolas', 9))
                e.grid(row=r, column=1, sticky='ew', pady=3)
                ttk.Button(grid, text='浏览…', width=7, command=lambda v=var: _browse(v))\
                    .grid(row=r, column=2, padx=(8, 0))
                varman[(side, name)] = var
                r += 1
        grid.columnconfigure(1, weight=1)

        note = ttk.Label(f, text='以下为相对 scui_ui 路径, 可修改；scui / scui_ui 为绝对路径且固定。修改即时生效。',
                         foreground='#888').pack(anchor='w', pady=(8, 4))
        ttk.Button(f, text='关闭', command=top.destroy).pack(side='right')
        for (side, name), var in varman.items():
            var.trace_add('write', lambda *a, s=side, n=name, v=var: self._apply_path(s, n, v))
        top.protocol('WM_DELETE_WINDOW', top.destroy)

    def _apply_path(self, side, name, var):
        val = var.get().strip()
        if not val:
            return
        # 相对路径 -> 转绝对(相对 scui_ui)
        if not os.path.isabs(val):
            val = os.path.normpath(os.path.join(self.ui, val))
        if side == 'in':
            self.in_abs[name] = val
        else:
            self.out_abs[name] = val
        if hasattr(self, 'path_status'):
            self.path_status.config(text='%s %s: %s' % (name, '输入' if side == 'in' else '输出', self._rel(val)))
        pv = self._pv.get((name, side))
        if pv is not None:
            pv.set('%s: %s' % ('输入' if side == 'in' else '输出', self._rel(val)))

    def _rel(self, p):
        try:
            return os.path.relpath(p, self.ui).replace('\\', '/')
        except Exception:
            return p

    #--------------- 四个子界面(tab) ---------------
    def _build_notebook(self):
        import tkinter as tk
        from tkinter import ttk
        self.nb = ttk.Notebook(self.root)
        self.nb.pack(fill='both', expand=True, padx=12, pady=(0, 12))
        self.tabs = {}
        self._build_widget_tab()         # widget: 完整(左文件/中册子/右编辑)
        self._build_image_tab()          # image: 完整
        self._build_font_tab()           # font: 完整
        self._build_lang_tab()           # lang: 预览为主
        self._build_cwf_tab()            # cwf: 参考 widget
        self.nb.bind('<<NotebookTabChanged>>', self._on_tab)

    #--------------- widget 子界面(左: src文件 | 中: analyze字段册 | 右: 编辑副本+log) ---------------
    def _build_widget_tab(self):
        import tkinter as tk
        from tkinter import ttk, scrolledtext
        f = ttk.Frame(self.nb, padding=8)
        self.tabs['widget'] = f
        # 编辑状态(副本): path/分支/条目索引
        self.widget_path  = None        # 当前编辑文件(绝对)
        self.widget_data  = None        # json 副本对象
        self.widget_maker = False       # False=scene; True=maker(默认配置)
        self.widget_edit  = None        # 当前条目索引
        self._w_base     = None         # 加载时的基快照(用于未保存修改检测)

        # 顶部工具条: 操作按钮(左) + 执行打包(右) + 状态(左)
        top = ttk.Frame(f); top.pack(fill='x', pady=(2, 4))
        ttk.Button(top, text='加载 maker json', command=self._wid_load_maker).pack(side='left')
        ttk.Button(top, text='预览 json', command=self._wid_preview_json).pack(side='left', padx=(6, 0))
        ttk.Button(top, text='保存 json', command=self._wid_save).pack(side='left', padx=(6, 0))
        ttk.Button(top, text='执行 widget 打包', command=lambda: self._run('widget')).pack(side='right')
        self.widget_status = ttk.Label(top, text='未选择文件', foreground='#777')
        self.widget_status.pack(side='left', padx=(12, 0))

        # 主体: 左中右三栏(占满, 不再内嵌 LOG)
        hpan = ttk.Panedwindow(f, orient='horizontal'); hpan.pack(fill='both', expand=True, pady=(6, 0))

        # 左: analyze 可配置字段册(类型窗格: window/custom/scroll...)
        lf = ttk.LabelFrame(hpan, text=' analyze 可配置字段 ', padding=(4, 4)); hpan.add(lf, weight=3)
        self.wbook = ttk.Treeview(lf, columns=('slot',), show='tree headings', selectmode='browse')
        self.wbook.heading('#0', text='字段路径'); self.wbook.heading('slot', text='值槽')
        self.wbook.column('slot', width=56, anchor='e', stretch=False)
        bvs = ttk.Scrollbar(lf, orient='vertical', command=self.wbook.yview)
        self.wbook.configure(yscrollcommand=bvs.set)
        self.wbook.pack(side='left', fill='both', expand=True); bvs.pack(side='right', fill='y')
        self.wbook.bind('<<TreeviewSelect>>', self._wid_book_copy)

        # 中: src 文件树(.json / .c)
        mf = ttk.LabelFrame(hpan, text=' src 文件(scene) ', padding=(4, 4)); hpan.add(mf, weight=3)
        self.wtree = ttk.Treeview(mf, show='tree', selectmode='browse')
        wvs = ttk.Scrollbar(mf, orient='vertical', command=self.wtree.yview)
        self.wtree.configure(yscrollcommand=wvs.set)
        self.wtree.pack(side='left', fill='both', expand=True); wvs.pack(side='right', fill='y')
        self.wtree.bind('<<TreeviewSelect>>', self._wid_pick_file)

        # 右: 预览/编辑(json: 键|值两列就地编辑; c: 只读文本)
        rf = ttk.LabelFrame(hpan, text=' 预览 / 编辑(json 就地, .c 只读) ', padding=(4, 4)); hpan.add(rf, weight=4)
        erow = ttk.Frame(rf); erow.pack(fill='x')
        for t, c in (('添加控件', self._wid_add_control), ('删除控件', self._wid_del_control),
                     ('添加字段', self._wid_add_field), ('删除字段', self._wid_del_field)):
            ttk.Button(erow, text=t, command=c).pack(side='left', padx=(0, 6))
        self.widget_sub = tk.StringVar(value='json: 每行 key|value 直接点击修改; 点 .c 只读; 增删用按钮; 光标所在即操作对象')
        ttk.Label(rf, textvariable=self.widget_sub, foreground='#777').pack(anchor='w', padx=(2, 2))
        # 键值编辑(json): 滚动行列表, 每行 key|value 常驻输入框(自由编辑)
        self.wcanv = tk.Canvas(rf, highlightthickness=0)
        wsb = ttk.Scrollbar(rf, orient='vertical', command=self.wcanv.yview)
        self.wcanv.configure(yscrollcommand=wsb.set)
        self._wbox = ttk.Frame(self.wcanv)
        self._wbox_id = self.wcanv.create_window((0, 0), window=self._wbox, anchor='nw')
        self._wbox.bind('<Configure>', lambda e: self.wcanv.configure(scrollregion=self.wcanv.bbox('all')))
        self.wcanv.bind('<Configure>', lambda e: (self.wcanv.itemconfigure(self._wbox_id, width=e.width),
                                                  self.wcanv.configure(scrollregion=self.wcanv.bbox('all'))))
        self.wcanv.pack(side='left', fill='both', expand=True)
        wsb.pack(side='right', fill='y')
        self.wcanv.bind('<Enter>', lambda e: self.wcanv.bind_all('<MouseWheel>', self._wid_wheel))
        self.wcanv.bind('<Leave>', lambda e: self.wcanv.unbind_all('<MouseWheel>'))
        self._w_rows = []
        # 文本预览(c): 只读
        self.wctext = scrolledtext.ScrolledText(rf, wrap='char', font=('Consolas', 9), state='disabled')
        self._wid_switch(True)

        self.nb.add(f, text='widget  ')
        self._wid_analyze()
        self._wid_load_tree()
        self._wid_load_maker(_silent=True)   # 打开默认加载 maker json

    # 右侧主体切换: True=json 行编辑; False=c 只读文本
    def _wid_switch(self, edit):
        if edit:
            self.wcanv.pack(side='left', fill='both', expand=True)
            self.wctext.pack_forget()
        else:
            self.wcanv.pack_forget()
            self.wctext.pack(fill='both', expand=True)

    # canvas 鼠标滚轮
    def _wid_wheel(self, e):
        self.wcanv.yview_scroll(-1 * (e.delta // 120), 'units')

    # 记录光标所在对象: (item, field序号|None)
    def _wid_active(self, i, f):
        self._w_active = (i, f)

    def _wid_analyze(self):
        # 生成左窗格字段册(analyze 结果; tools 加入 sys.path 以便 import)
        tools = self.tools
        if tools not in sys.path:
            sys.path.insert(0, tools)
        try:
            import scui_widget_analyze as an
            book = an.scui_widget_analyze_result(tools)
        except Exception as e:
            self._wlog('analyze 失败: %r' % e)
            return
        self.wbook.delete(*self.wbook.get_children())
        for wtype, maker in book.get('class_maker', {}).items():
            prefix = maker
            if prefix.startswith('scui_'):
                prefix = prefix[5:]
            if prefix.endswith('_maker_t'):
                prefix = prefix[:-len('_maker_t')]
            slots = book.get('path_slots', {}).get(maker, {})
            top = self.wbook.insert('', 'end', text=prefix, open=False, values=('',))
            # 按字段路径逐点分层折叠(widget.style.buffer -> widget/style/buffer), 默认折叠
            node_items = {}
            for path, slot in slots.items():
                segs = path.split('.')
                d = node_items
                for s in segs[:-1]:
                    n = d.get(s)
                    if n is None:
                        n = {'slot': None, 'kids': {}}
                        d[s] = n
                    d = n['kids']
                leaf = d.get(segs[-1])
                if leaf is None:
                    d[segs[-1]] = {'slot': slot, 'kids': {}}
                else:
                    leaf['slot'] = slot
            def _rend(parent, items):
                for text, node in items.items():
                    if node['kids']:
                        nid = self.wbook.insert(parent, 'end', text=text, open=False, values=('',))
                        _rend(nid, node['kids'])
                    else:
                        self.wbook.insert(parent, 'end', text=text, values=(node['slot'],))
            _rend(top, node_items)

    # 左窗格点选字段 -> 复制路径到剪贴板(粘贴到右侧 key 用)
    def _wid_book_copy(self, _ev=None):
        sel = self.wbook.selection()
        if not sel:
            return
        path = self.wbook.item(sel[0], 'text')
        if not path:
            return
        try:
            self.root.clipboard_clear()
            self.root.clipboard_append(path)
        except Exception as e:
            return
        self.widget_status.config(text='已复制字段: %s(可粘贴到右侧 key)' % path)

    def _wlog(self, s):
        self._append_log('widget', s + '\n')

    # 左: src 文件树(递归展开全部, .json/.c 可点)
    def _wid_load_tree(self):
        self.wtree.delete(*self.wtree.get_children())
        root = self.in_abs['widget']
        if not os.path.isdir(root):
            root = self.ui
        rnode = self.wtree.insert('', 'end', text=os.path.basename(root) or root, open=True, iid='rdir')
        self._wid_scan_dir(rnode, root)

    def _wid_scan_dir(self, parent, path):
        try:
            entries = sorted(os.listdir(path),
                             key=lambda x: (not os.path.isdir(os.path.join(path, x)), x.lower()))
        except OSError:
            return
        for e in entries:
            full = os.path.join(path, e)
            if os.path.isdir(full):
                node = self.wtree.insert(parent, 'end', text=e, iid=full, open=True)
                self._wid_scan_dir(node, full)
            elif e.lower().endswith(('.json', '.c')):
                self.wtree.insert(parent, 'end', text=e, iid=full)

    # 中: 点选文件
    def _wid_pick_file(self, _ev=None):
        sel = self.wtree.selection()
        if not sel:
            return
        path = sel[0]
        if not (os.path.isfile(path) and path.lower().endswith(('.json', '.c'))):
            return
        if not self._wid_confirm_discard():      # 切走前, 未保存修改需确认丢弃
            return
        self.widget_edit = None
        if path.lower().endswith('.c'):
            self.widget_path = path; self.widget_data = None; self.widget_maker = False
            self.widget_status.config(text='%s (只读)' % os.path.basename(path))
            self.widget_sub.set('%s: 文件文本预览(只读, 不可修改)' % os.path.basename(path))
            self.wctext.configure(state='normal')
            self.wctext.delete('1.0', 'end')
            try:
                with open(path, 'r', encoding='utf-8', errors='replace') as fp:
                    self.wctext.insert('1.0', fp.read())
            except Exception as e:
                self.wctext.insert('1.0', '读取失败: %r' % e)
            self.wctext.configure(state='disabled')
            self._wid_switch(False)
            return
        try:
            with open(path, 'r', encoding='utf-8') as fp:
                self.widget_data = json.load(fp)
            self.widget_path = path; self.widget_maker = False
            self._w_base = json.loads(json.dumps(self.widget_data))   # 记录基快照
            self.widget_status.config(text='已加载: %s' % os.path.basename(path))
            self.widget_sub.set('%s: 每行 key/value 直接修改, 光标所在可增删' % os.path.basename(path))
            self._wid_switch(True)
            self._wid_render_json()
        except Exception as e:
            self.widget_status.config(text='加载失败: %r' % e)

    # 平铺渲染: 每个条目 = 标题Label + 每字段一行 [key输入框][value输入框], 之间空行
    def _wid_render_json(self):
        import tkinter as tk
        from tkinter import ttk
        for w in self._wbox.winfo_children():
            w.destroy()
        self._w_rows = []
        if self.widget_data is None:
            return
        items = self.widget_data.get('widget', [])
        for i, it in enumerate(items):
            if self.widget_maker:
                title = '#%d  %s' % (i, it.get('class', '?'))
                fields = it.setdefault('default', {})
            else:
                title = '#%d  %s' % (i, it.get('widget.myself') or it.get('widget.type', ''))
                fields = it
            tl = ttk.Label(self._wbox, text=title, foreground='#2266cc', font=('Consolas', 9, 'bold'))
            tl.pack(fill='x', anchor='w', pady=(6, 2))
            tl.bind('<Button-1>', lambda e, i=i: self._wid_active(i, None))
            for f, (k, v) in enumerate(fields.items()):
                kvar = tk.StringVar(value=str(k))
                vvar = tk.StringVar(value=str(v))
                row = ttk.Frame(self._wbox); row.pack(fill='x', padx=(2, 2))
                ke = ttk.Entry(row, textvariable=kvar, font=('Consolas', 9))
                ve = ttk.Entry(row, textvariable=vvar, font=('Consolas', 9))
                ke.pack(side='left', fill='x', expand=True, padx=(0, 8))
                ve.pack(side='left', fill='x', expand=True)
                ke.bind('<FocusIn>', lambda e, i=i, f=f: self._wid_active(i, f))
                ve.bind('<FocusIn>', lambda e, i=i, f=f: self._wid_active(i, f))
                self._w_rows.append([i, f, kvar, vvar])
            ttk.Label(self._wbox, text='').pack()

    # 取某条目的字段 dict
    def _wid_fields_of(self, item):
        it = self.widget_data['widget'][item]
        if self.widget_maker:
            return it.setdefault('default', {})
        return it

    # 把当前输入框文本回填到副本(改 key 重建保序; 空 key 行视为未定义跳过)
    def _wid_sync_back(self):
        if not self.widget_data:
            return
        groups = {}
        for i, f, ke, ve in self._w_rows:
            groups.setdefault(i, []).append((ke.get(), ve.get()))
        for i, pairs in groups.items():
            if i >= len(self.widget_data['widget']):
                continue
            d = {}
            for k, v in pairs:
                if k:
                    d[k] = v
            if self.widget_maker:
                self.widget_data['widget'][i]['default'] = d
            else:
                self.widget_data['widget'][i] = d

    # 操作位置定位(返回光标所在条目索引, 无则 None)
    def _wid_where(self):
        act = getattr(self, '_w_active', None)
        if act is None:
            return None
        return act[0]

    # 写某条目字段 dict
    def _wid_set_fields(self, it, newd):
        if self.widget_maker:
            self.widget_data['widget'][it]['default'] = newd
        else:
            self.widget_data['widget'][it] = newd

    # 新控件默认字段
    def _wid_new_control(self):
        b = {'widget.type': 'scui_widget_type_window', 'widget.clip.w': 'SCUI_HOR_RES',
             'widget.clip.h': 'SCUI_VER_RES', 'widget.myself': 'SCUI_UI_SCENE_'}
        return {'class': 'scui_widget_type_window', 'default': b} if self.widget_maker else b

    # 未保存修改检测(先把输入框回填再与基快照比对)
    def _wid_dirty(self):
        if self.widget_data is None or self._w_base is None:
            return False
        self._wid_sync_back()
        a = json.dumps(self.widget_data, ensure_ascii=False, sort_keys=True)
        b = json.dumps(self._w_base, ensure_ascii=False, sort_keys=True)
        return a != b

    # 切换前确认丢弃修改(是->继续; 否->取消本次切换)
    def _wid_confirm_discard(self):
        if not self._wid_dirty():
            return True
        from tkinter import messagebox
        return messagebox.askyesno('未保存的修改', '当前 json 有未保存的修改，确定丢弃吗？', parent=self.root)

    # 添加控件: 光标所在条目之后插一个新控件(默认字段)
    def _wid_add_control(self):
        from tkinter import messagebox
        if not (self.widget_data is not None and self.widget_path):
            messagebox.showwarning('widget', '请先在中间选择一个 .json 文件')
            return
        self._wid_sync_back()
        pos = self._wid_where()
        if pos is None:
            self._wlog('[widget] 请定位操作位置(点一下标题或某输入框)')
            return
        items = self.widget_data.setdefault('widget', [])
        if pos >= len(items):
            return
        items.insert(pos + 1, self._wid_new_control())
        self._wid_render_json()
        self._wlog('已添加控件(默认字段, 未保存)')

    # 删除控件: 光标所在条目整个删除
    def _wid_del_control(self):
        from tkinter import messagebox
        if not (self.widget_data is not None and self.widget_path):
            messagebox.showwarning('widget', '请先在中间选择一个 .json 文件')
            return
        self._wid_sync_back()
        pos = self._wid_where()
        if pos is None:
            self._wlog('[widget] 请定位操作位置(点一下标题或某输入框)')
            return
        items = self.widget_data.get('widget', [])
        if pos >= len(items):
            return
        del items[pos]
        self._wid_render_json()
        self._wlog('已删除控件(未保存)')

    # 添加字段: 光标所在条目补一个空字段行
    def _wid_add_field(self):
        from tkinter import messagebox
        if not (self.widget_data is not None and self.widget_path):
            messagebox.showwarning('widget', '请先在中间选择一个 .json 文件')
            return
        self._wid_sync_back()
        pos = self._wid_where()
        if pos is None:
            self._wlog('[widget] 请定位操作位置(点一下标题或某输入框)')
            return
        items = self.widget_data.get('widget', [])
        if pos >= len(items):
            return
        f = getattr(self, '_w_active', [None, None])[1]
        fields = self._wid_fields_of(pos)
        newd = {}
        for j, (k, v) in enumerate(fields.items()):
            newd[k] = v
            if f is not None and j == f:
                newd[''] = ''
        if f is None:
            newd[''] = ''
        self._wid_set_fields(pos, newd)
        self._wid_render_json()
        self._wlog('已添加字段(空行, 在输入框填名与值, 未保存)')

    # 删除字段: 光标所在字段行删除
    def _wid_del_field(self):
        from tkinter import messagebox
        if not (self.widget_data is not None and self.widget_path):
            messagebox.showwarning('widget', '请先在中间选择一个 .json 文件')
            return
        self._wid_sync_back()
        act = getattr(self, '_w_active', None)
        if act is None or act[0] is None or act[1] is None:
            self._wlog('[widget] 请定位操作位置(点一下要删除的 key/value 输入框)')
            return
        pos, f = act
        items = self.widget_data.get('widget', [])
        if pos >= len(items):
            return
        fields = self._wid_fields_of(pos)
        ks = list(fields.keys())
        if f >= len(ks):
            return
        kdel = ks[f]
        newd = {k: v for k, v in fields.items() if k != kdel}
        if not newd:                                   # 删空了 -> 删除控件本身, 避免空控件
            del items[pos]
            self._wlog('已删除字段 %s, 控件已空故删除整个控件(未保存)' % kdel)
        else:
            self._wid_set_fields(pos, newd)
            self._wlog('已删除字段 %s(未保存)' % kdel)
        self._wid_render_json()

    # 预览 json: 弹窗显示修改后的完整 json
    def _wid_preview_json(self):
        import tkinter as tk
        from tkinter import scrolledtext
        self._wid_sync_back()
        top = tk.Toplevel(self.root)
        top.title('json 预览(修改后) - %s' % (os.path.basename(self.widget_path) if self.widget_path else '未命名'))
        top.geometry('680x560')
        t = scrolledtext.ScrolledText(top, wrap='none', font=('Consolas', 9), state='normal')
        t.pack(fill='both', expand=True, padx=8, pady=8)
        try:
            s = json.dumps(self.widget_data, ensure_ascii=False, indent=4)
        except Exception as e:
            s = '%r' % e
        t.insert('1.0', s)
        t.configure(state='disabled')

    # maker json 加载(out 下默认配置, 可编辑保存; _silent=自动加载时不弹窗)
    def _wid_load_maker(self, _silent=False):
        from tkinter import messagebox
        path = os.path.join(self.out_abs['widget'], 'scui_ui_maker.json')
        if not os.path.exists(path):
            if not _silent:
                messagebox.showwarning('widget', '未找到 maker 默认配置:\n%s' % path)
            return
        try:
            if not _silent and not self._wid_confirm_discard():   # 手动加载前, 未保存修改需确认
                return
            with open(path, 'r', encoding='utf-8') as fp:
                self.widget_data = json.load(fp)
            self.widget_path = path; self.widget_maker = True
            self._w_base = json.loads(json.dumps(self.widget_data))   # 记录基快照
            self.widget_status.config(text='已加载 maker json(默认配置): %s' % os.path.basename(path))
            self.widget_sub.set('maker json: key/value 直接修改')
            self._wid_switch(True)
            self._wid_render_json()
        except Exception as e:
            self.widget_status.config(text='加载失败: %r' % e)

    # 保存(json 可写; 先回填输入框, 保存才写文件)
    def _wid_save(self):
        if self.widget_data is None or not self.widget_path:
            self._wlog('[save] 无待保存内容')
            return
        if self.widget_path.lower().endswith('.c'):
            self._wlog('[save] .c 只读, 不允许修改写入')
            return
        try:
            self._wid_sync_back()
            with open(self.widget_path, 'w', encoding='utf-8') as fp:
                json.dump(self.widget_data, fp, ensure_ascii=False, indent=4)
            self._w_base = json.loads(json.dumps(self.widget_data))   # 保存后重置基快照
            self.widget_status.config(text='已保存: %s' % os.path.basename(self.widget_path))
            self._wlog('[save] 已写入: %s' % self.widget_path)
        except Exception as e:
            self.widget_status.config(text='保存失败: %r' % e)
            self._wlog('[save] 失败: %r' % e)

#--------------- lang 子界面(左:语言列 | 中:句柄行列表 | 右:行详情) ---------------
    def _build_lang_tab(self):
        import tkinter as tk
        from tkinter import ttk, scrolledtext
        f = ttk.Frame(self.nb, padding=8)
        self.tabs['lang'] = f
        self.lang_data = None     # (languages, rows)
        self.lang_sel = None

        # 顶部
        top = ttk.Frame(f); top.pack(fill='x', pady=(2, 4))
        ttk.Button(top, text='执行 lang 打包', command=lambda: self._run('lang')).pack(side='right')

        # 主体: 三栏(占满, 不再内嵌 LOG)
        hp = ttk.Panedwindow(f, orient='horizontal'); hp.pack(fill='both', expand=True, pady=(6, 0))

        # 左: 语言列表
        lf = ttk.LabelFrame(hp, text=' language ', padding=(4, 4)); hp.add(lf, weight=1)
        self.ltree = ttk.Treeview(lf, show='tree', selectmode='browse')
        lvs = ttk.Scrollbar(lf, orient='vertical', command=self.ltree.yview)
        self.ltree.configure(yscrollcommand=lvs.set)
        self.ltree.pack(side='left', fill='both', expand=True); lvs.pack(side='right', fill='y')

        # 中: 句柄行列表(zh 提示, 带省略号)
        mf = ttk.LabelFrame(hp, text=' 句柄行(zh 预览) ', padding=(4, 4)); hp.add(mf, weight=3)
        self.mtree = ttk.Treeview(mf, columns=('idx',), show='tree headings', selectmode='browse')
        self.mtree.heading('#0', text='句柄行'); self.mtree.heading('idx', text='#')
        self.mtree.column('idx', width=40, anchor='e', stretch=False)
        mvs = ttk.Scrollbar(mf, orient='vertical', command=self.mtree.yview)
        self.mtree.configure(yscrollcommand=mvs.set)
        self.mtree.pack(side='left', fill='both', expand=True); mvs.pack(side='right', fill='y')
        self.mtree.bind('<<TreeviewSelect>>', self._lang_pick)

        # 右: 行详情(各语言内容)
        rf = ttk.LabelFrame(hp, text=' 行详情 ', padding=(4, 4)); hp.add(rf, weight=3)
        self.rtext = scrolledtext.ScrolledText(rf, wrap='word', font=('Consolas', 9), state='disabled')
        self.rtext.pack(fill='both', expand=True)

        self.nb.add(f, text='lang  ')
        self._lang_load()

    def _lang_load(self):
        self.ltree.delete(*self.ltree.get_children())
        self.mtree.delete(*self.mtree.get_children())
        self.rtext.configure(state='normal'); self.rtext.delete('1.0', 'end'); self.rtext.configure(state='disabled')
        self.lang_data = None; self.lang_sel = None

        src = self.in_abs['lang']
        cfg = os.path.join(src, 'scui_lang_parser.json')
        if not os.path.exists(cfg):
            self._append_log('lang', '未找到配置: %s\n' % cfg)
            return
        try:
            with open(cfg, 'r', encoding='utf-8') as fp:
                j = json.load(fp)
            langs = j.get('language', [])
            xlsx_path = os.path.join(src, j.get('xlsx', ''))
            sheet_name = j.get('sheet', '')
            if not os.path.exists(xlsx_path):
                self._append_log('lang', '未找到 xlsx: %s\n' % xlsx_path)
                return
            import openpyxl
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb[wb.sheetnames[0]]
            rows = []
            for row in ws.iter_rows(values_only=True):
                rows.append([str(c) if c is not None else '' for c in row])
            wb.close()
            self.lang_data = (langs, rows)
            # 左: 语言列表
            for i, l in enumerate(langs):
                self.ltree.insert('', 'end', text=l, iid='lang%d' % i)
            # 中: 句柄行(第 0 列为 zh, 带省略号)
            for i, row in enumerate(rows):
                zh = row[0] if row else ''
                show = zh if len(zh) <= 24 else zh[:22] + '…'
                self.mtree.insert('', 'end', iid='r%d' % i, text=show, values=(i,))
            self._append_log('lang', '已加载 %d 行, %d 语言\n' % (len(rows), len(langs)))
        except Exception as e:
            self._append_log('lang', '加载失败: %r\n' % e)

    def _lang_pick(self, _ev=None):
        sel = self.mtree.selection()
        if not sel or self.lang_data is None:
            return
        idx = int(sel[0][1:])
        langs, rows = self.lang_data
        if idx >= len(rows):
            return
        row = rows[idx]
        self.rtext.configure(state='normal')
        self.rtext.delete('1.0', 'end')
        self.rtext.insert('end', '句柄号: SCUI_LANG_0X%04x\n' % idx)
        self.rtext.insert('end', '-' * 40 + '\n')
        for ci, lang in enumerate(langs):
            val = row[ci] if ci < len(row) else ''
            self.rtext.insert('end', '%s: %s\n' % (lang, val))
        self.rtext.configure(state='disabled')

    #--------------- cwf 子界面(左:协议类型+注释 | 中:cwf浏览 | 右:json编辑) ---------------
    def _build_cwf_tab(self):
        import tkinter as tk
        from tkinter import ttk, scrolledtext
        f = ttk.Frame(self.nb, padding=8)
        self.tabs['cwf'] = f
        self.cwf_data = None
        self.cwf_path = None
        self.cwf_edit = None
        self._cwf_base = None
        self._cwf_rows = []
        self._cwf_active = None

        # 顶部工具条
        top = ttk.Frame(f); top.pack(fill='x', pady=(2, 4))
        ttk.Button(top, text='预览 json', command=self._cwf_preview).pack(side='left')
        ttk.Button(top, text='保存 json', command=self._cwf_save).pack(side='left', padx=(6, 0))
        ttk.Button(top, text='执行 cwf 打包', command=lambda: self._run('cwf')).pack(side='right')
        self.cwf_status = ttk.Label(top, text='未选择', foreground='#777')
        self.cwf_status.pack(side='left', padx=(12, 0))

        # 主体: 三栏(占满, 不再内嵌 LOG)
        hp = ttk.Panedwindow(f, orient='horizontal'); hp.pack(fill='both', expand=True, pady=(6, 0))

        # 左: 协议类型 + 注释
        lf = ttk.LabelFrame(hp, text=' cwf 协议类型 ', padding=(4, 4)); hp.add(lf, weight=3)
        self.cbook = ttk.Treeview(lf, columns=('anno',), show='tree headings', selectmode='browse')
        self.cbook.heading('#0', text='type'); self.cbook.heading('anno', text='annotation')
        self.cbook.column('anno', width=120, stretch=False)
        cvs = ttk.Scrollbar(lf, orient='vertical', command=self.cbook.yview)
        self.cbook.configure(yscrollcommand=cvs.set)
        self.cbook.pack(side='left', fill='both', expand=True); cvs.pack(side='right', fill='y')

        # 中: cwf 浏览(满足: 有 image.7z + 同名 .json)
        mf = ttk.LabelFrame(hp, text=' cwf 列表 ', padding=(4, 4)); hp.add(mf, weight=3)
        self.ctree = ttk.Treeview(mf, show='tree', selectmode='browse')
        ctvs = ttk.Scrollbar(mf, orient='vertical', command=self.ctree.yview)
        self.ctree.configure(yscrollcommand=ctvs.set)
        self.ctree.pack(side='left', fill='both', expand=True); ctvs.pack(side='right', fill='y')
        self.ctree.bind('<<TreeviewSelect>>', self._cwf_pick)

        # 右: json 编辑(layout 行列表, 每行 key|value 常驻输入框)
        rf = ttk.LabelFrame(hp, text=' 编辑(json 就地) ', padding=(4, 4)); hp.add(rf, weight=4)
        erow = ttk.Frame(rf); erow.pack(fill='x')
        for t, c in (('添加类型', self._cwf_add_type), ('删除类型', self._cwf_del_type),
                     ('添加字段', self._cwf_add_field), ('删除字段', self._cwf_del_field)):
            ttk.Button(erow, text=t, command=c).pack(side='left', padx=(0, 6))
        self.cwf_sub = tk.StringVar(value='中间选择一个 cwf: 各 layout 行 key|value 直接修改')
        ttk.Label(rf, textvariable=self.cwf_sub, foreground='#777').pack(anchor='w', padx=(2, 2))
        self.ccanv = tk.Canvas(rf, highlightthickness=0)
        csb = ttk.Scrollbar(rf, orient='vertical', command=self.ccanv.yview)
        self.ccanv.configure(yscrollcommand=csb.set)
        self._cbox = ttk.Frame(self.ccanv)
        self._cbox_id = self.ccanv.create_window((0, 0), window=self._cbox, anchor='nw')
        self._cbox.bind('<Configure>', lambda e: self.ccanv.configure(scrollregion=self.ccanv.bbox('all')))
        self.ccanv.bind('<Configure>', lambda e: (self.ccanv.itemconfigure(self._cbox_id, width=e.width),
                                                   self.ccanv.configure(scrollregion=self.ccanv.bbox('all'))))
        self.ccanv.pack(side='left', fill='both', expand=True); csb.pack(side='right', fill='y')
        self.ccanv.bind('<Enter>', lambda e: self.ccanv.bind_all('<MouseWheel>', self._cwf_wheel))
        self.ccanv.bind('<Leave>', lambda e: self.ccanv.unbind_all('<MouseWheel>'))

        self.nb.add(f, text='cwf  ')
        self._cwf_load_proto()
        self._cwf_load_list()

    def _cwf_wheel(self, e):
        self.ccanv.yview_scroll(-1 * (e.delta // 120), 'units')

    # 加载协议 json(类型 + 注释)
    def _cwf_load_proto(self):
        self.cbook.delete(*self.cbook.get_children())
        proto = os.path.join(self.plugs, 'scui_cwf_json_parser.json')
        if not os.path.exists(proto):
            self._append_log('cwf', '未找到协议: %s\n' % proto)
            return
        try:
            with open(proto, 'r', encoding='utf-8') as fp:
                j = json.load(fp)
            for item in j.get('scui_cwf_json_type', []):
                key = item.get('key', '?')
                annos = [v for k, v in item.items() if k == 'annotation']
                anno = ' ; '.join(annos) if annos else ''
                self.cbook.insert('', 'end', text=key, values=(anno,))
        except Exception as e:
            self._append_log('cwf', '协议加载失败: %r\n' % e)

    # 加载 cwf 列表(满足: 有 image.7z + 同名 .json)
    def _cwf_load_list(self):
        self.ctree.delete(*self.ctree.get_children())
        root = self.in_abs['cwf']
        if not os.path.isdir(root):
            self._append_log('cwf', 'cwf 路径不存在: %s\n' % root)
            return
        for name in sorted(os.listdir(root)):
            d = os.path.join(root, name)
            if not os.path.isdir(d):
                continue
            has_7z = os.path.exists(os.path.join(d, 'image.7z'))
            has_json = os.path.exists(os.path.join(d, name + '.json'))
            if has_7z and has_json:
                self.ctree.insert('', 'end', text=name, iid=d)

    # 选中 cwf -> 加载 json
    def _cwf_pick(self, _ev=None):
        sel = self.ctree.selection()
        if not sel:
            return
        path = sel[0]
        if not os.path.isdir(path):
            return
        name = os.path.basename(path)
        json_path = os.path.join(path, name + '.json')
        if not os.path.exists(json_path):
            return
        if not self._cwf_confirm_discard():
            return
        try:
            with open(json_path, 'r', encoding='utf-8') as fp:
                self.cwf_data = json.load(fp)
            self.cwf_path = json_path
            self._cwf_base = json.loads(json.dumps(self.cwf_data))
            self.cwf_status.config(text='已加载: %s' % name)
            self._cwf_render()
        except Exception as e:
            self.cwf_status.config(text='加载失败: %r' % e)

    # 渲染 layout 行(每个条目 = 标题 + key|value 行)
    def _cwf_render(self):
        import tkinter as tk
        from tkinter import ttk
        for w in self._cbox.winfo_children():
            w.destroy()
        self._cwf_rows = []
        if self.cwf_data is None:
            return
        items = self.cwf_data.get('layout', [])
        for i, it in enumerate(items):
            title = '#%d  %s' % (i, it.get('type', '?'))
            tl = ttk.Label(self._cbox, text=title, foreground='#2266cc', font=('Consolas', 9, 'bold'))
            tl.pack(fill='x', anchor='w', pady=(6, 2))
            tl.bind('<Button-1>', lambda e, i=i: self._cwf_active_set(i, None))
            for f, (k, v) in enumerate(it.items()):
                kvar = tk.StringVar(value=str(k))
                vvar = tk.StringVar(value=str(v))
                row = ttk.Frame(self._cbox); row.pack(fill='x', padx=(2, 2))
                ke = ttk.Entry(row, textvariable=kvar, font=('Consolas', 9))
                ve = ttk.Entry(row, textvariable=vvar, font=('Consolas', 9))
                ke.pack(side='left', fill='x', expand=True, padx=(0, 8))
                ve.pack(side='left', fill='x', expand=True)
                ke.bind('<FocusIn>', lambda e, i=i, f=f: self._cwf_active_set(i, f))
                ve.bind('<FocusIn>', lambda e, i=i, f=f: self._cwf_active_set(i, f))
                self._cwf_rows.append([i, f, kvar, vvar])
            ttk.Label(self._cbox, text='').pack()

    def _cwf_active_set(self, i, f):
        self._cwf_active = (i, f)

    def _cwf_fields_of(self, item):
        return self.cwf_data['layout'][item]

    def _cwf_sync_back(self):
        if not self.cwf_data:
            return
        groups = {}
        for i, f, ke, ve in self._cwf_rows:
            groups.setdefault(i, []).append((ke.get(), ve.get()))
        for i, pairs in groups.items():
            if i >= len(self.cwf_data['layout']):
                continue
            d = {}
            for k, v in pairs:
                if k:
                    d[k] = v
            self.cwf_data['layout'][i] = d

    def _cwf_dirty(self):
        if self.cwf_data is None or self._cwf_base is None:
            return False
        self._cwf_sync_back()
        a = json.dumps(self.cwf_data, ensure_ascii=False, sort_keys=True)
        b = json.dumps(self._cwf_base, ensure_ascii=False, sort_keys=True)
        return a != b

    def _cwf_confirm_discard(self):
        if not self._cwf_dirty():
            return True
        from tkinter import messagebox
        return messagebox.askyesno('未保存的修改', '当前 cwf json 有未保存的修改，确定丢弃吗？', parent=self.root)

    def _cwf_new_type(self):
        return {'type': 'scui_cwf_json_type_img_simple', 'x': 0, 'y': 0, 'image_src': []}

    def _cwf_add_type(self):
        from tkinter import messagebox
        if self.cwf_data is None:
            messagebox.showwarning('cwf', '请先在中间选择一个 cwf')
            return
        self._cwf_sync_back()
        items = self.cwf_data.setdefault('layout', [])
        pos = self._cwf_active[0] if self._cwf_active else None
        if pos is None:
            self._append_log('cwf', '请定位操作位置\n')
            return
        items.insert(pos + 1, self._cwf_new_type())
        self._cwf_render()
        self._append_log('cwf', '已添加类型(未保存)\n')

    def _cwf_del_type(self):
        from tkinter import messagebox
        if self.cwf_data is None:
            messagebox.showwarning('cwf', '请先在中间选择一个 cwf')
            return
        self._cwf_sync_back()
        pos = self._cwf_active[0] if self._cwf_active else None
        if pos is None:
            self._append_log('cwf', '请定位操作位置\n')
            return
        items = self.cwf_data.get('layout', [])
        if pos < len(items):
            del items[pos]
            self._cwf_render()
            self._append_log('cwf', '已删除类型(未保存)\n')

    def _cwf_add_field(self):
        from tkinter import messagebox
        if self.cwf_data is None:
            messagebox.showwarning('cwf', '请先在中间选择一个 cwf')
            return
        self._cwf_sync_back()
        act = self._cwf_active
        pos = act[0] if act else None
        if pos is None:
            self._append_log('cwf', '请定位操作位置\n')
            return
        items = self.cwf_data.get('layout', [])
        if pos >= len(items):
            return
        f = act[1] if act else None
        fields = self._cwf_fields_of(pos)
        newd = {}
        for j, (k, v) in enumerate(fields.items()):
            newd[k] = v
            if f is not None and j == f:
                newd[''] = ''
        if f is None:
            newd[''] = ''
        self.cwf_data['layout'][pos] = newd
        self._cwf_render()
        self._append_log('cwf', '已添加字段(未保存)\n')

    def _cwf_del_field(self):
        from tkinter import messagebox
        if self.cwf_data is None:
            messagebox.showwarning('cwf', '请先在中间选择一个 cwf')
            return
        self._cwf_sync_back()
        act = self._cwf_active
        if act is None or act[0] is None or act[1] is None:
            self._append_log('cwf', '请定位操作位置(点 key/value 输入框)\n')
            return
        pos, f = act
        items = self.cwf_data.get('layout', [])
        if pos >= len(items):
            return
        fields = self._cwf_fields_of(pos)
        ks = list(fields.keys())
        if f >= len(ks):
            return
        kdel = ks[f]
        newd = {k: v for k, v in fields.items() if k != kdel}
        if not newd:
            del items[pos]
            self._append_log('cwf', '已删除字段 %s, 类型已空故删除(未保存)\n' % kdel)
        else:
            self.cwf_data['layout'][pos] = newd
            self._append_log('cwf', '已删除字段 %s(未保存)\n' % kdel)
        self._cwf_render()

    def _cwf_preview(self):
        import tkinter as tk
        from tkinter import scrolledtext
        self._cwf_sync_back()
        top = tk.Toplevel(self.root)
        top.title('cwf json 预览 - %s' % (os.path.basename(self.cwf_path) if self.cwf_path else '未命名'))
        top.geometry('680x560')
        t = scrolledtext.ScrolledText(top, wrap='none', font=('Consolas', 9), state='normal')
        t.pack(fill='both', expand=True, padx=8, pady=8)
        try:
            s = json.dumps(self.cwf_data, ensure_ascii=False, indent=4)
        except Exception as e:
            s = '%r' % e
        t.insert('1.0', s)
        t.configure(state='disabled')

    def _cwf_save(self):
        if self.cwf_data is None or not self.cwf_path:
            self._append_log('cwf', '[save] 无待保存内容\n')
            return
        try:
            self._cwf_sync_back()
            with open(self.cwf_path, 'w', encoding='utf-8') as fp:
                json.dump(self.cwf_data, fp, ensure_ascii=False, indent=4)
            self._cwf_base = json.loads(json.dumps(self.cwf_data))
            self.cwf_status.config(text='已保存: %s' % os.path.basename(self.cwf_path))
            self._append_log('cwf', '[save] 已写入: %s\n' % self.cwf_path)
        except Exception as e:
            self.cwf_status.config(text='保存失败: %r' % e)
            self._append_log('cwf', '[save] 失败: %r\n' % e)

    def _build_font_tab(self):
        import tkinter as tk
        from tkinter import ttk
        f = ttk.Frame(self.nb, padding=8)
        self.tabs['font'] = f
        self.font_status = tk.StringVar(value='未加载')

        # 顶部工具条: json 三键 + 执行打包(同一列)
        bar = ttk.Frame(f); bar.pack(fill='x', pady=(2, 4))
        for t, c in (('加载 json', lambda: self._font_load(False)),
                     ('预览 json', self._font_preview),
                     ('保存 json', self._font_save)):
            ttk.Button(bar, text=t, command=c).pack(side='left')
        ttk.Button(bar, text='执行 font 打包', command=lambda: self._run('font')).pack(side='right')
        ttk.Label(bar, textvariable=self.font_status, foreground='#777').pack(side='left', padx=(12, 0))
        ttk.Label(f, text='font.bin 来源于 lv_font_conv 工具，见 font_src/lv_font_conv txt & py',
                  foreground='#888').pack(anchor='w', padx=(2, 0))

        # 主体: 水平可调(树 | 编辑, 占满)
        hp = ttk.Panedwindow(f, orient='horizontal'); hp.pack(fill='both', expand=True, pady=(6, 0))
        lf = ttk.LabelFrame(hp, text=' 字库(json) ', padding=(4, 4)); hp.add(lf, weight=1)
        self.ftree = ttk.Treeview(lf, columns=('size',), show='tree headings', selectmode='browse')
        self.ftree.heading('#0', text='字库条目'); self.ftree.heading('size', text='字号')
        self.ftree.column('size', width=60, anchor='e', stretch=False)
        fvs = ttk.Scrollbar(lf, orient='vertical', command=self.ftree.yview)
        self.ftree.configure(yscrollcommand=fvs.set)
        self.ftree.pack(side='left', fill='both', expand=True); fvs.pack(side='right', fill='y')
        self.ftree.bind('<<TreeviewSelect>>', self._on_font_sel)

        # 右: 六个操作按钮 与 编辑表单 水平并排
        rf = ttk.Frame(hp); hp.add(rf, weight=2)
        self.font_json = None
        self.font_sel = None       # (lang_idx, item_idx)
        ops = ttk.LabelFrame(rf, text='语言 / 字库'); ops.pack(side='left', fill='y', padx=(0, 8))
        for t, c in (('添加语言', self._font_addlang), ('删除语言', self._font_dellang),
                     ('修改语言', self._font_modlang), ('添加字库', self._font_add),
                     ('删除字库', self._font_del), ('修改字库', self._font_modlib)):
            ttk.Button(ops, text=t, command=c).pack(fill='x', pady=2)
        ef = ttk.LabelFrame(rf, text=' 编辑(json 特性) '); ef.pack(side='left', fill='both', expand=True)
        self._build_font_editor(ef)

        self.nb.add(f, text='font  ')
        self._font_load(False)

    def _build_font_editor(self, rf):
        import tkinter as tk
        from tkinter import ttk
        self._fe = {}
        grid = ttk.Frame(rf); grid.pack(fill='x')
        ttk.Label(grid, text='lang').grid(row=0, column=0, sticky='w', padx=(0, 8), pady=3)
        self._fe['lang'] = tk.Entry(grid, width=16, state='readonly')   # 用 tk.Entry, readonly 才生效
        self._fe['lang'].grid(row=0, column=1, sticky='ew', pady=3)
        ttk.Label(grid, text='name').grid(row=1, column=0, sticky='w', padx=(0, 8), pady=3)
        self._fe['name'] = ttk.Entry(grid, width=30)
        self._fe['name'].grid(row=1, column=1, sticky='ew', pady=3)
        ttk.Label(grid, text='size').grid(row=2, column=0, sticky='w', padx=(0, 8), pady=3)
        self._fe['size'] = ttk.Entry(grid, width=10)
        self._fe['size'].grid(row=2, column=1, sticky='w', pady=3)
        ttk.Label(grid, text='base_line_ext').grid(row=3, column=0, sticky='w', padx=(0, 8), pady=3)
        self._fe['base'] = ttk.Entry(grid, width=10)
        self._fe['base'].grid(row=3, column=1, sticky='w', pady=3)
        ttk.Label(grid, text='line_height_ext').grid(row=4, column=0, sticky='w', padx=(0, 8), pady=3)
        self._fe['line'] = ttk.Entry(grid, width=10)
        self._fe['line'].grid(row=4, column=1, sticky='w', pady=3)
        ttk.Label(rf, text='提示: 左树选中条目 → 在右侧表单就地编辑 → 点「修改字库」立即生效；点「保存 json」写文件',
                  foreground='#888', wraplength=340).pack(anchor='w', pady=(6, 0))
        grid.columnconfigure(1, weight=1)

    def _font_json_path(self):
        return os.path.join(self.out_abs['font'], 'scui_font_package.json')

    def _font_rebuild(self):
        """从内存副本重建左树(保留已展开的语言分组, 保证 iid 与数组下标同步)."""
        openli = set()
        if hasattr(self, '_font_lang_iid'):
            for top, li in list(self._font_lang_iid.items()):
                try:
                    if self.ftree.item(top, 'open'):
                        openli.add(li)
                except Exception:
                    pass
        self._font_lang_iid = {}
        self.ftree.delete(*self.ftree.get_children())
        for li, item in enumerate(self.font_json.get('info', [])):
            lang = item.get('lang', '?')
            top = self.ftree.insert('', 'end', text=lang, values=('',),
                                    open=(li in openli))
            self._font_lang_iid[top] = li
            sizes = item.get('size', [])
            for ei, name in enumerate(item.get('name', [])):
                sz = sizes[ei] if ei < len(sizes) else 0
                self.ftree.insert(top, 'end', text=name, values=(sz,),
                                  iid='f%d_%d' % (li, ei))

    def _font_normalize(self):
        """归一化: 保证每个条目的 name/size/ext 长度对齐, 缺失 ext 补默认 0."""
        info = self.font_json.setdefault('info', [])
        for item in info:
            n = len(item.setdefault('name', []))
            item.setdefault('size', [0] * n)
            if len(item['size']) < n:
                item['size'] += [0] * (n - len(item['size']))
            ext = item.setdefault('ext', [None] * n)
            for i in range(n):
                if not isinstance(ext[i], dict):
                    ext[i] = {'base_line_ext': 0, 'line_height_ext': 0}

    def _font_load(self, _notify):
        import tkinter as tk
        p = self._font_json_path()
        try:
            with open(p, 'r', encoding='utf-8') as fp:
                self.font_json = json.load(fp)
            self._font_normalize()
            self._font_rebuild()
            self.font_status.set('已加载: %s' % os.path.basename(p))
        except Exception as e:
            self.font_status.set('加载失败: %r' % e)

    def _on_font_sel(self, _ev=None):
        # 支持直接传 iid(添加条目后填充); 否则从当前 selection 取
        if isinstance(_ev, str):
            iid = _ev
        else:
            sel = self.ftree.selection()
            if not sel:
                return
            iid = sel[0]
        # 顶层 lang 分组: 记录 li, 供添加字号使用
        if iid in getattr(self, '_font_lang_iid', {}):
            self.font_sel = (self._font_lang_iid[iid], -1)
            for k in self._fe:
                self._fe[k].configure(state='normal'); self._fe[k].delete(0, 'end')
                self._fe[k].configure(state='readonly' if k == 'lang' else 'normal')
            self.font_status.set('已选择分组 %s(可点「添加字号」)' % self.font_json['info'][self._font_lang_iid[iid]]['lang'])
            return
        if not iid.startswith('f'):
            return
        li, ei = map(int, iid[1:].split('_'))
        self.font_sel = (li, ei)
        item = self.font_json['info'][li]
        name = item['name'][ei]
        size = item['size'][ei] if ei < len(item['size']) else 0
        self._fe['lang'].configure(state='normal'); self._fe['lang'].delete(0, 'end'); self._fe['lang'].insert(0, item['lang']); self._fe['lang'].configure(state='readonly')
        self._fe['name'].delete(0, 'end'); self._fe['name'].insert(0, name)
        self._fe['size'].delete(0, 'end'); self._fe['size'].insert(0, str(size))
        # ext(可选)
        base = line = 0
        ext = item.get('ext')
        if ext and ei < len(ext) and isinstance(ext[ei], dict):
            base = ext[ei].get('base_line_ext', 0)
            line = ext[ei].get('line_height_ext', 0)
        self._fe['base'].delete(0, 'end'); self._fe['base'].insert(0, str(base))
        self._fe['line'].delete(0, 'end'); self._fe['line'].insert(0, str(line))

    def _font_prune(self):
        """合规性: 移除没有字库条目(空 name)的语言分组, 并从 lang 白名单剔除未被引用的语言."""
        from tkinter import messagebox
        removed = []
        info = self.font_json.get('info', [])
        new_info = []
        for item in info:
            names = [n for n in item.get('name', []) if str(n).strip()]
            if names:
                item['name'] = names
                if len(item['size']) != len(names):
                    item['size'] = [0] * len(names)
                new_info.append(item)
            else:
                removed.append(item.get('lang', '?'))
        self.font_json['info'] = new_info
        used = set(i.get('lang') for i in new_info)
        lang = self.font_json.setdefault('lang', [])
        self.font_json['lang'] = [l for l in lang if l in used]
        if removed:
            messagebox.showinfo('保存 json', '已移除无效语言分组(未含字库条目): %s' % ', '.join(removed), parent=self.root)
        return removed

    def _font_save(self):
        if self.font_json is None:
            return
        if self.font_sel:
            li, ei = self.font_sel
            item = self.font_json['info'][li]
            item['name'][ei] = self._fe['name'].get().strip()
            try:
                item['size'][ei] = int(self._fe['size'].get().strip())
            except Exception:
                pass
            # ext(可选, 空=0)
            def _to_int(s):
                try:
                    return int(str(s).strip())
                except Exception:
                    return 0
            b = _to_int(self._fe['base'].get()); l = _to_int(self._fe['line'].get())
            item.setdefault('ext', [None] * len(item['name']))
            while ei >= len(item['ext']):
                item['ext'].append(None)
            item['ext'][ei] = {'base_line_ext': b, 'line_height_ext': l}
        # 合规性: 去掉只加语言未加字库的无效分组
        self._font_prune()
        try:
            with open(self._font_json_path(), 'w', encoding='utf-8') as fp:
                json.dump(self.font_json, fp, ensure_ascii=False, indent=4)
            # 若当前选中了条目, 就地同步树行文本(不重建, 保持展开)
            if self.font_sel:
                li, ei = self.font_sel
                if li < len(self.font_json['info']):
                    it = self.font_json['info'][li]
                    if ei < len(it['name']):
                        self.ftree.item('f%d_%d' % (li, ei), text=it['name'][ei],
                                        values=(it['size'][ei] if ei < len(it['size']) else 0,))
            self.font_status.set('已保存')
        except Exception as e:
            self.font_status.set('保存失败: %r' % e)

    def _font_add(self):
        from tkinter import messagebox
        if self.font_json is None:
            messagebox.showwarning('font', '尚未加载 json，请先点「加载 json」')
            return
        sel = self.ftree.selection()
        if not sel:
            messagebox.showwarning('font', '请先点选左侧一个语言分组或某个条目，再添加')
            return
        iid = sel[0]
        if iid.startswith('f'):
            li = int(iid[1:].split('_')[0])   # 点的是某 bin -> 所属分组
        else:
            li = self._font_lang_iid.get(iid)
            if li is None:
                messagebox.showwarning('font', '请先点选左侧一个语言分组')
                return
        item = self.font_json['info'][li]
        lang = item['lang']
        item.setdefault('name', [])
        item.setdefault('size', [])
        base = 'font_%s_new_%d.bin' % (lang, len(item['name']))
        item['name'].append(base)
        item['size'].append(0)
        # 就地加数组 + 重建树(iid 与数组同步, 且保留展开), 再选中新条目
        self._font_rebuild()
        new_iid = 'f%d_%d' % (li, len(item['name']) - 1)
        top_iid = next((k for k, v in self._font_lang_iid.items() if v == li), None)
        if top_iid is not None:
            self.ftree.item(top_iid, open=True)
        self.ftree.selection_set(new_iid)
        self._on_font_sel(new_iid)
        self.font_status.set('已添加(未保存，点「保存 json」同步)')

    def _font_addlang(self):
        from tkinter import simpledialog, messagebox
        if self.font_json is None:
            messagebox.showwarning('font', '尚未加载 json，请先点「加载 json」')
            return
        lang = simpledialog.askstring('添加语言', '新语言名(如 en2):', parent=self.root)
        if not lang:
            return
        lang = lang.strip()
        if not lang:
            return
        self.font_json.setdefault('lang', [])
        if lang not in self.font_json['lang']:
            self.font_json['lang'].append(lang)
        self.font_json.setdefault('info', [])
        # 同名已存在则直接选中
        for li, item in enumerate(self.font_json['info']):
            if item.get('lang') == lang:
                self._font_rebuild()
                top = next((k for k, v in self._font_lang_iid.items() if v == li), None)
                if top is not None:
                    self.ftree.item(top, open=True)
                    self.ftree.selection_set(top)
                self.font_status.set('语言 %s 已存在，已选中' % lang)
                return
        self.font_json['info'].append({'lang': lang, 'name': [], 'size': []})
        self._font_rebuild()
        li = len(self.font_json['info']) - 1
        top = next((k for k, v in self._font_lang_iid.items() if v == li), None)
        if top is not None:
            self.ftree.item(top, open=True)
            self.ftree.selection_set(top)
        self.font_status.set('已添加语言 %s(可再点「添加字号」补条目)' % lang)

    def _font_modlang(self):
        from tkinter import simpledialog, messagebox
        if self.font_json is None:
            messagebox.showwarning('font', '尚未加载 json，请先点「加载 json」')
            return
        sel = self.ftree.selection()
        if not sel:
            messagebox.showwarning('font', '请先选中一个语言分组')
            return
        iid = sel[0]
        if iid.startswith('f'):
            li = int(iid[1:].split('_')[0])
        else:
            li = self._font_lang_iid.get(iid)
            if li is None:
                messagebox.showwarning('font', '请先选中一个语言分组')
                return
        item = self.font_json['info'][li]
        old = item.get('lang', '')
        new = simpledialog.askstring('修改语言', '新语言名:', initialvalue=old, parent=self.root)
        if not new:
            return
        new = new.strip()
        if not new:
            return
        item['lang'] = new
        L = self.font_json.setdefault('lang', [])
        if old in L and not any(i.get('lang') == old for i in self.font_json['info']):
            L.remove(old)
        if new not in L:
            L.append(new)
        self._font_rebuild()
        top = next((k for k, v in self._font_lang_iid.items() if v == li), None)
        if top is not None:
            self.ftree.item(top, open=True)
            self.ftree.selection_set(top)
        self.font_status.set('语言已改为 %s(未保存)' % new)

    def _font_dellang(self):
        from tkinter import messagebox
        if self.font_json is None:
            messagebox.showwarning('font', '尚未加载 json，请先点「加载 json」')
            return
        sel = self.ftree.selection()
        if not sel:
            messagebox.showwarning('font', '请先选中一个语言分组')
            return
        iid = sel[0]
        if iid.startswith('f'):
            li = int(iid[1:].split('_')[0])
        else:
            li = self._font_lang_iid.get(iid)
            if li is None:
                messagebox.showwarning('font', '请先选中一个语言分组')
                return
        item = self.font_json['info'][li]
        lang = item.get('lang', '?')
        if not messagebox.askyesno('删除语言', '确认删除整个语言分组 [%s] 及其全部字库条目?' % lang, parent=self.root):
            return
        del self.font_json['info'][li]
        L = self.font_json.setdefault('lang', [])
        if lang in L and not any(i.get('lang') == lang for i in self.font_json['info']):
            L.remove(lang)
        self.ftree.selection_remove(self.ftree.selection())
        self.font_sel = None
        self._font_rebuild()
        self.font_status.set('已删除语言 %s(未保存)' % lang)

    def _font_modlib(self):
        from tkinter import messagebox
        if self.font_json is None:
            messagebox.showwarning('font', '尚未加载 json，请先点「加载 json」')
            return
        sel = self.ftree.selection()
        if not sel or not sel[0].startswith('f'):
            messagebox.showwarning('font', '请先在左侧选中一个具体字库条目(.bin 那一行)')
            return
        iid = sel[0]
        li, ei = map(int, iid[1:].split('_'))
        item = self.font_json['info'][li]
        self.font_sel = (li, ei)
        name = self._fe['name'].get().strip()
        if name:
            item['name'][ei] = name
        try:
            item['size'][ei] = int(self._fe['size'].get().strip())
        except Exception:
            pass
        def _to_int(s):
            try:
                return int(str(s).strip())
            except Exception:
                return 0
        b = _to_int(self._fe['base'].get()); l = _to_int(self._fe['line'].get())
        item.setdefault('ext', [None] * len(item['name']))
        while ei >= len(item['ext']):
            item['ext'].append(None)
        item['ext'][ei] = {'base_line_ext': b, 'line_height_ext': l}
        # 就地同步树行文本
        self.ftree.item('f%d_%d' % (li, ei), text=item['name'][ei],
                        values=(item['size'][ei],))
        self.font_status.set('已修改字库(未写文件，点「保存 json」同步)')

    def _font_del(self):
        from tkinter import messagebox
        if self.font_json is None:
            messagebox.showwarning('font', '尚未加载 json，请先点「加载 json」')
            return
        sel = self.ftree.selection()
        if not sel:
            messagebox.showwarning('font', '请先在左侧选中一个具体的字库条目(.bin 那一行)')
            return
        iid = sel[0]
        if not iid.startswith('f'):
            messagebox.showwarning('font', '请选中具体的字库条目(展开语言分组后点 .bin 那一行)')
            return
        li, ei = map(int, iid[1:].split('_'))
        item = self.font_json['info'][li]
        if ei < len(item['name']):
            del item['name'][ei]
        if ei < len(item['size']):
            del item['size'][ei]
        if ei < len(item.get('ext', [])):
            del item['ext'][ei]
        # 就地删数组 + 重建树(iid 与数组同步, 且保留展开)
        self.ftree.selection_remove(iid)
        self.font_sel = None
        self._font_rebuild()
        self.font_status.set('已删除(未保存，点「保存 json」同步)')

    def _font_preview(self):
        import tkinter as tk
        from tkinter import scrolledtext
        if self.font_json is None:
            return
        top = tk.Toplevel(self.root)
        top.title('json 预览(副本，未保存的修改在此可见)')
        top.geometry('720x560')
        t = scrolledtext.ScrolledText(top, wrap='none', font=('Consolas', 9), state='normal')
        t.pack(fill='both', expand=True, padx=8, pady=8)
        try:
            s = json.dumps(self.font_json, ensure_ascii=False, indent=4)
        except Exception as e:
            s = '%r' % e
        t.insert('1.0', s)
        t.configure(state='disabled')

    #--------------- image 子界面 ---------------
    def _build_image_tab(self):
        import tkinter as tk
        from tkinter import ttk
        f = ttk.Frame(self.nb, padding=8)
        self.tabs['image'] = f

        # row0: 顶部固定(项目名 + 执行)
        top = ttk.Frame(f); top.pack(fill='x', pady=(2, 4))
        pr = ttk.Frame(top); pr.pack(fill='x')
        ttk.Label(pr, text='image 项目名称:').pack(side='left')
        self.proj_var = tk.StringVar(value='prj')
        ttk.Entry(pr, textvariable=self.proj_var, width=16).pack(side='left', padx=(6, 10))
        ttk.Button(pr, text='执行 image 打包', command=lambda: self._run('image')).pack(side='right')

        # 主体: 水平 Panedwindow(树 | 详情[图|文 水平]) -> 可左右拖拽
        mid = ttk.Panedwindow(f, orient='horizontal')
        mid.pack(fill='both', expand=True, pady=(4, 0))

        lf = ttk.LabelFrame(mid, text=' 图片资源 ', padding=(4, 4))
        mid.add(lf, weight=3)
        self.tree = ttk.Treeview(lf, columns=('size',), show='tree headings',
                                 selectmode='browse')
        self.tree.heading('#0', text='路径/文件')
        self.tree.heading('size', text='大小')
        self.tree.column('size', width=80, anchor='e', stretch=False)
        vs = ttk.Scrollbar(lf, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=vs.set)
        self.tree.pack(side='left', fill='both', expand=True)
        vs.pack(side='right', fill='y')

        root_src = self.in_abs['image'] if os.path.isdir(self.in_abs['image']) else self.ui
        self._tree_root_path = root_src
        try:
            root_show = os.path.relpath(root_src, self.in_abs['image'] or root_src)
        except Exception:
            root_show = root_src
        self._tree_root = self.tree.insert('', 'end', iid='root', text=root_show, open=True)
        self._vid_map = {}
        self._vid_count = 0
        self._load_dir(self._tree_root, root_src)
        self.tree.bind('<<TreeviewOpen>>', self._on_tree_open)
        self.tree.bind('<<TreeviewSelect>>', self._on_tree_sel)

        rf = ttk.LabelFrame(mid, text=' 详情(scui_image) ', padding=(6, 4))
        mid.add(rf, weight=2)
        dbody = ttk.Frame(rf); dbody.pack(fill='both', expand=True)
        self.preview = ttk.Label(dbody, text='选择左侧图片查看预览', anchor='center',
                                 width=26, relief='groove')
        self.preview.pack(side='left', fill='y', padx=(0, 6))
        self.field_text = tk.Text(dbody, font=('Consolas', 9), state='disabled', wrap='none')
        self.field_text.pack(side='left', fill='both', expand=True)

        self.nb.add(f, text='image  ')

    def _load_dir(self, parent, path):
        """懒加载: 只填一级目录 + 图片文件."""
        import tkinter as tk
        from PIL import Image, ImageTk
        try:
            entries = sorted(os.listdir(path),
                             key=lambda x: (not os.path.isdir(os.path.join(path, x)), x.lower()))
        except OSError:
            return
        for e in entries:
            full = os.path.join(path, e)
            if os.path.isdir(full):
                node = self.tree.insert(parent, 'end', text=e, values=(u'[目录]',), open=False)
                self.tree.insert(node, 'end', text='…', values=('',))
            elif e.lower().endswith(_IMG_EXT):
                try:
                    im = Image.open(full); im.thumbnail((18, 18))
                    ph = ImageTk.PhotoImage(im); self._imgs.append(ph)
                except Exception:
                    ph = None
                size = self._fsize(full)
                self.tree.insert(parent, 'end', text=e, image=ph, values=(size,),
                                 tags=('img',),
                                 iid='v%d' % self._vid_count)
                self._vid_map[self._vid_count] = full
                self._vid_count += 1

    def _fsize(self, p):
        try:
            n = os.path.getsize(p)
            if n >= 1024 * 1024:
                return '%.2fM' % (n / 1024 / 1024)
            return '%dK' % (n / 1024)
        except OSError:
            return ''

    def _on_tree_open(self, ev):
        node = ev.widget.focus()
        kids = self.tree.get_children(node)
        if len(kids) == 1:
            txt = self.tree.item(kids[0], 'text')
            if txt == '…':
                self.tree.delete(kids[0])
                path = self._node_path(node)
                self._load_dir(node, path)

    def _node_path(self, node):
        parts = []
        cur = node
        while cur:
            parts.insert(0, self.tree.item(cur, 'text'))
            cur = self.tree.parent(cur)
        return os.path.join(self._tree_root_path, *parts[1:])

    def _on_tree_sel(self, ev):
        sel = self.tree.selection()
        if not sel:
            return
        iid = sel[0]
        if iid.startswith('v'):
            path = self._vid_map[int(iid[1:])]
            self._show_img(path)

    def _show_img(self, path):
        import tkinter as tk
        from PIL import Image, ImageTk
        preview_err = None
        # 预览图无条件显示
        try:
            im = Image.open(path)
            im.load()
            im.thumbnail((180, 180))
            ph = ImageTk.PhotoImage(im)
            self._imgs.append(ph)
            self.preview.configure(image=ph, text='')
        except Exception as e:
            preview_err = '%r' % e
            self.preview.configure(image='', text=os.path.basename(path))

        # 构造 file_tag(与打包脚本同规则: 去点/去斜杠/去空格 -> 下划线)
        try:
            rel = os.path.relpath(path, self.in_abs['image'])
        except Exception:
            rel = os.path.basename(path)
        file_tag = rel.replace('.', '').replace('\\', '_').replace('/', '_').replace(' ', '_')

        lines = []
        if preview_err:
            lines.append('【预览失败】 %s' % preview_err)
            lines.append('')
        # 匹配 .h 句柄信息
        h_info = self._image_from_h(file_tag)
        if h_info:
            lines.append('【句柄信息（scui_image_parser.h）】')
            lines.append('    %-18s = %s;' % ('handle_name', h_info['name']))
            lines.append('    %-18s = %s;' % ('handle_value', h_info['value']))
            lines.append('    %-18s = %s;' % ('pixel.width', h_info['width']))
            lines.append('    %-18s = %s;' % ('pixel.height', h_info['height']))
            lines.append('    %-18s = %s;' % ('pixel.size_bin', h_info['size_bin']))
            lines.append('    %-18s = %s;' % ('pixel.size_mem', h_info['size_mem']))
            lines.append('    %-18s = %s;' % ('com_pct', h_info['com_pct']))
            lines.append('')
        # 匹配 .c 结构体信息
        c_info = self._image_from_c(file_tag)
        if c_info:
            lines.append('【资源信息（scui_image_parser.c）】')
            for k, v in c_info.items():
                lines.append('    %-18s = %s;' % (k, v))
            lines.append('')
        # 都没匹配到
        if not h_info and not c_info:
            lines.append('【未匹配到打包记录 — 请先执行 image 打包】')
            lines.append('    (file_tag: %s)' % file_tag)

        self.field_text.configure(state='normal')
        self.field_text.delete('1.0', 'end')
        self.field_text.insert('1.0', '\n'.join(lines))
        self.field_text.configure(state='disabled')

    def _image_from_h(self, file_tag):
        """从 scui_image_parser.h 匹配句柄信息(file_tag 为去点后的相对路径)."""
        out_h = os.path.join(self.out_abs.get('image', ''), 'scui_image_parser.h')
        if not os.path.exists(out_h):
            return None
        try:
            with open(out_h, 'r', encoding='utf-8', errors='replace') as hf:
                content = hf.read()
            # 枚举项以 file_tag 结尾: scui_image_<prefix><file_tag>, // 0xNNNN
            re_enum = r'(scui_image_[A-Za-z0-9_]*%s)\s*,\s*//\s*(0x[0-9a-fA-F]+)' % re.escape(file_tag)
            m = re.search(re_enum, content)
            if not m:
                return None
            name = m.group(1)
            value = m.group(2)
            # 压缩注释行也以同后缀结尾: //< w,h,size_bin,size_mem,com_pct> scui_image_<prefix><file_tag>
            re_comment = r'//<\s*([^>]+)>\s*scui_image_[A-Za-z0-9_]*%s\b' % re.escape(file_tag)
            m2 = re.search(re_comment, content)
            if not m2:
                return None
            # 注释里为十六进制径 w,h,size_bin,size_mem(0x...)
            nums = re.findall(r'0x([0-9a-fA-F]+)', m2.group(1))
            ws = [int(x, 16) for x in nums]
            if len(ws) < 4:
                return None
            w, h, bin_, mem = ws[0], ws[1], ws[2], ws[3]
            return {
                'name': name,
                'value': value,
                'width': w,
                'height': h,
                'size_bin': bin_,
                'size_mem': mem,
                'com_pct': '%.4f' % (float(bin_) / mem) if mem else 1.0,
            }
        except OSError:
            return None

    def _image_from_c(self, file_tag):
        """从 scui_image_parser.c 匹配 scui_image_t 结构体."""
        out_c = os.path.join(self.out_abs.get('image', ''), 'scui_image_parser.c')
        if not os.path.exists(out_c):
            return None
        try:
            with open(out_c, 'r', encoding='utf-8', errors='replace') as cf:
                content = cf.read()
            # 结构体名以 file_tag 结尾(无 scui_image_ 前缀): static const scui_image_t <prefix><file_tag>
            re_struct = r'static\s+const\s+scui_image_t\s+([A-Za-z0-9_]*%s)\s*=\s*\{([^}]+)\}' % re.escape(file_tag)
            m = re.search(re_struct, content, re.DOTALL)
            if not m:
                return None
            body = m.group(2)   # group(1)=结构体名, group(2)={...}体
            info = {}
            # 解析各字段(脚本生成以 ','/';' 结尾)
            for line in body.split('\n'):
                line = line.strip()
                if '=' in line and line[-1] in (';', ','):
                    k, v = line.split('=', 1)
                    k = k.strip().lstrip('.')
                    v = v.strip().rstrip(';').rstrip(',').strip()
                    info[k] = v
            return info
        except OSError:
            return None

    #--------------- 执行打包(子线程, UI 不卡) ---------------
    def _run(self, name):
        if getattr(self, '_busy', False):
            return
        self._busy = True
        self._status('执行中…(%s)' % name)
        src = self.in_abs[name]
        dst = self.out_abs[name]
        proj = getattr(self, 'proj_var', None)
        proj = proj.get() if proj else 'prj'
        # 打包前清空全局日志缓冲, 并弹出共享日志窗口
        self._log_text = ''
        if self._log_win is not None:
            t = getattr(self._log_win, '_text', None)
            if t is not None:
                try:
                    t.configure(state='normal')
                    t.delete('1.0', 'end')
                    t.configure(state='disabled')
                except Exception:
                    pass
        self._log_show()
        self._append_log(name, '==== start: %s ====\n' % name)

        self._logq = queue.Queue()
        wire = _TabLog(self._logq)

        def worker():
            so, se = sys.stdout, sys.stderr
            si = sys.stdin
            sys.stdout = wire
            sys.stderr = wire
            sys.stdin = _NoStdin()          # 无控制台 exe 下避免 input() 阻塞
            try:
                _do_task(name, self.ui, src, dst, proj)
            finally:
                sys.stdout, sys.stderr, sys.stdin = so, se, si
                self._logq.put('__DONE__')

        threading.Thread(target=worker, daemon=True).start()

    def _append_log(self, name, txt):
        # 所有子界面日志共用一份全局内存缓冲 + 全局日志窗口
        del name
        self._log_text += txt
        if self._log_win is not None:
            t = getattr(self._log_win, '_text', None)
            if t is not None:
                try:
                    t.configure(state='normal')
                    t.insert('end', txt)
                    t.see('end')
                    t.configure(state='disabled')
                except Exception:
                    pass

    def _log_show(self):
        # 打开/聚焦全局日志窗口(与菜单「日志」共用一个窗口)
        import tkinter as tk
        if self._log_win is not None:
            try:
                self._log_win.deiconify()
                self._log_win.lift()
                return
            except Exception:
                self._log_win = None
        self._log_win = tk.Toplevel(self.root)
        self._log_win.title('运行日志(全局共享)')
        self._log_win.geometry('820x560')
        from tkinter import scrolledtext
        t = scrolledtext.ScrolledText(self._log_win, wrap='word', font=('Consolas', 9), state='disabled')
        t.pack(fill='both', expand=True, padx=8, pady=8)
        self._log_win._text = t
        static = self._log_text if self._log_text.strip() else '(暂无日志, 请先在任意子界面执行打包)'
        t.configure(state='normal')
        t.insert('1.0', static)
        t.see('end')
        t.configure(state='disabled')
        self._log_win.protocol('WM_DELETE_WINDOW', self._log_hide)

    def _log_hide(self):
        if self._log_win is not None:
            self._log_win.withdraw()

    def _poll(self):
        if self._logq is not None:
            # 每帧限流, 避免海量日志一次性插入拖慢主线程
            n = 0
            while n < 500:
                try:
                    txt = self._logq.get_nowait()
                except queue.Empty:
                    break
                n += 1
                if txt == '__DONE__':
                    self._busy = False
                    self._status('完成')
                    continue
                self._append_log(self._current_name, txt)
        self.root.after(120, self._poll)

    def _status(self, s):
        self.root.title('scui资源工具 - 正在开发中... - %s' % s)

    def _on_tab(self, ev):
        idx = self.nb.index(self.nb.select())
        names = ['widget', 'image', 'font', 'lang', 'cwf']
        self._current_name = names[idx] if idx < len(names) else 'widget'

    def _about(self):
        from tkinter import messagebox
        messagebox.showinfo('scui资源工具',
            '作者：Agent\n'
            '版本：Ver 0.0.1\n'
            '状态：持续开发中...\n'
            '类型：widget/image/font/lang/cwf\n'
            '来源：scui/tools & scui/plugs')

    # 工具「日志」菜单 -> 打开/聚焦全局共享日志窗口
    def _open_logs(self):
        self._log_show()

class _TabLog(object):
    """把 print 写入的文本重定向进指定队列(保留换行), 由 GUI 日志泵写日志区."""
    def __init__(self, q):
        self.q = q
    def write(self, s):
        if s:
            self.q.put(s)
    def flush(self):
        pass

class _NoStdin(object):
    """无控制台环境下替换 stdin, 使 input() 立即抛 EOFError 而非阻塞."""
    def read(self, *a, **k):
        raise EOFError
    def readline(self, *a, **k):
        raise EOFError
    def readlines(self, *a, **k):
        raise EOFError
    def __iter__(self):
        return self
    def __next__(self):
        raise StopIteration

#============================================================
# 入口
#============================================================
def main():
    # 无控制台(exe)下 stdout/stderr 为 None, 替换为 devnull 避免意外崩溃
    if sys.stdout is None:
        sys.stdout = open(os.devnull, 'w')
    if sys.stderr is None:
        sys.stderr = open(os.devnull, 'w')

    # 轻量手动解析: 首个位置参数作为类型(须为已知类型); 未知参数一律忽略, 回退 GUI
    argv = sys.argv[1:]
    typ = ui = src = dst = proj = ''
    i = 0
    while i < len(argv):
        a = argv[i]
        if a == '--ui' and i + 1 < len(argv):
            ui = argv[i + 1]; i += 2
        elif a == '--src' and i + 1 < len(argv):
            src = argv[i + 1]; i += 2
        elif a == '--dst' and i + 1 < len(argv):
            dst = argv[i + 1]; i += 2
        elif a == '--proj' and i + 1 < len(argv):
            proj = argv[i + 1]; i += 2
        elif a.startswith('--'):
            i += 1                            # 未知选项忽略
        else:
            if typ == '' and a in _TASK:
                typ = a                       # 仅首个位置参数可为已知类型
            i += 1                            # 其它位置参数(如误拖入的路径)忽略

    ui_root = os.path.normpath(ui or os.getcwd())
    if typ:
        sdef, ddef = _DEFAULT[typ]
        src = src or os.path.join(ui_root, sdef)
        dst = dst or os.path.join(ui_root, ddef)
        return _do_task(typ, ui_root, src, dst, proj)
    return _launch_gui(ui_root)

if __name__ == '__main__':
    sys.exit(main())