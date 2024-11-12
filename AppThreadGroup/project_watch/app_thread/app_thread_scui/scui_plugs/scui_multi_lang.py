import os
import sys
import json
import openpyxl
# -- coding: utf-8 --**


# for row in xlsx_sheet.rows: # 获取每一行的数据
#     for data in row:        # 获取每一行中单元格的数据
#         print(data.value)  # 打印单元格的值
# print(xlsx_sheet.cell(1, 1).value)


# 编写集成化源文件
def encode_scui_multi_lang_c(file, xlsx_sheet, sheet_row, sheet_col):
    # 写点简要的说明
    file.write('/*本地静态的字符串表\n')
    file.write(' *通过scui_multi_lang.py生成\n */\n\n')
    file.write('#include "scui.h"\n\n')
    # 提取所有外源依赖
    file.write('const char * scui_multi_lang_table[%d * %d] = {\n' % (sheet_row, sheet_col))
    for item in xlsx_sheet.iter_cols():
        for data in item:
            c_array = str([hex(c) for c in str(data.value).encode('utf-8')])
            c_array = c_array.replace('\'', '')
            c_array = c_array.replace('[', '{')
            c_array = c_array.replace(']', ', 0}')
            c_array = c_array.replace('{', '(const char []){')
            file.write('\t//%s\n' % str(data.value).replace('\n', '\\n'))
            file.write('\t{0}\n'.format("%s," % c_array))    # 多国语这里不能很好的对齐,因为不同语言的空格宽度不一致
    file.write('};\n')


# 编写集成化头文件
def encode_scui_multi_lang_h(file, xlsx_sheet, sheet_row, sheet_col, args_list):
    file.write('#ifndef SCUI_MULTI_LANG_H\n')
    file.write('#define SCUI_MULTI_LANG_H\n\n')
    # 写点简要的说明
    file.write('/*本地静态的字符串表\n')
    file.write(' *通过scui_multi_lang.py生成\n */\n\n')
    # 编写头部索引
    file.write('#define SCUI_MULTI_LANG_NUM_TYPE    %s\n' % str(sheet_col))
    file.write('#define SCUI_MULTI_LANG_NUM_STR     %s\n\n' % str(sheet_row))
    # 编写头部索引
    file.write('typedef enum {\n')
    for idx, item in enumerate(args_list[0]):
        file.write('\t%s = %s * %s,\n' % (item, 'SCUI_MULTI_LANG_NUM_STR', str(idx)))
    file.write('} scui_multi_lang_type_t;\n\n')
    # 编写头部索引
    file.write('typedef enum {\n')
    file.write('\t%s = %s,\n' % (args_list[1], args_list[2]))
    for i, item in enumerate(xlsx_sheet.rows):
        c_str = str(item[0].value).replace('\n', '\\n')
        file.write('\tSCUI_MULTI_LANG_0X%04x,\t/* %s */\n' % (i, c_str))
    file.write('} scui_multi_lang_str_t;\n\n')
    file.write('extern const char * scui_multi_lang_table[%d * %d];\n\n' % (sheet_row, sheet_col))
    # 编写固化访问函数
    file.write('/*@brief 转换字符串句柄到多国语字符串句柄\n')
    file.write(' *       字符串句柄默认是第一个语言的字符串句柄\n')
    file.write(' *       其他语言都需要添加偏移做转换\n')
    file.write(' *@param type   语言编号(scui_multi_lang_type_t)\n')
    file.write(' *@param handle 字符串编号(scui_multi_lang_str_t)\n */\n')
    file.write('static inline scui_handle_t scui_multi_lang_switch(scui_handle_t type, scui_handle_t handle)\n')
    file.write('{\n\treturn type + handle;\n}\n')
    file.write('#endif\n')


# 打印重定向到文件
class ScuiRedirectPrint(object):
    def __init__(self, stream=sys.stdout, path='.', file='log.txt'):
        if not os.path.exists(path):
            os.makedirs(path)
        self.log = open(os.path.join(path, file), mode='w+', encoding='utf-8')
        self.terminal = stream

    def write(self, message):
        self.log.write(message)
        self.terminal.write(message)

    def flush(self):
        pass


# 启用集成化资源总表生成
def encode_scui_multi_lang():
    # 参数列表:文件根目录,
    if len(sys.argv) != 3:
        print('argv list not match')
        return
    src_path = sys.argv[1]
    dst_path = sys.argv[2]
    # 获得文件处理根路径
    if not os.path.exists(src_path):
        print('src path is not exist')
        return
    if not os.path.exists(dst_path):
        print('dst path is not exist')
        return
    # print重定向
    sys.stdout = ScuiRedirectPrint(sys.stdout, file=os.path.join(dst_path, 'scui_multi_lang.out'))  # redirect print
    sys.stderr = ScuiRedirectPrint(sys.stderr, file=os.path.join(dst_path, 'scui_multi_lang.err'))  # redirect print
    print('src path:', src_path)
    print('dst path:', dst_path)
    # json转Python字符串并转标准字典
    json_file = open(os.path.join(src_path, 'scui_multi_lang.json'), 'r', encoding='utf-8')
    json_dict = json.loads(json_file.read())
    json_file.close()
    # 仅支持批量化字符串表管理
    if json_dict['type'] != 'scui_multi_lang':
        return
    # 从标准字典获取工作簿名字和单元
    xlsx_file = openpyxl.load_workbook(os.path.join(src_path, json_dict['xlsx']), read_only=False)
    xlsx_sheet = xlsx_file[json_dict['sheet']]
    sheet_row = xlsx_sheet.max_row
    sheet_col = xlsx_sheet.max_column
    # 获得多语言表,和句柄偏移
    args_list = [json_dict['language'], json_dict['offset_name'], json_dict['offset_value']]
    if sheet_col != len(args_list[0]):
        print('args list not match')
        return
    # 开启三个文件
    scui_multi_lang_h = open(os.path.join(dst_path, 'scui_multi_lang.h'), mode='w', encoding='utf-8')
    scui_multi_lang_c = open(os.path.join(dst_path, 'scui_multi_lang.c'), mode='w', encoding='utf-8')
    # 解析
    encode_scui_multi_lang_h(scui_multi_lang_h, xlsx_sheet, sheet_row, sheet_col, args_list)
    encode_scui_multi_lang_c(scui_multi_lang_c, xlsx_sheet, sheet_row, sheet_col)
    # 关闭三个文件
    scui_multi_lang_h.close()
    scui_multi_lang_c.close()
    # 关闭工作簿
    xlsx_file.close()


if __name__ == '__main__':
    try:
        encode_scui_multi_lang()
    except Exception as e:
        print(e)
    print('scui multi lang finish')
    input('请按任意键退出...')
