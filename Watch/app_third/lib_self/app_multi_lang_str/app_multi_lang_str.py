import os
import os.path
import json
import openpyxl

# -- coding: utf-8 --**


# for row in xlsx_sheet.rows: # 获取每一行的数据
#     for data in row:        # 获取每一行中单元格的数据
#         print(data.value)  # 打印单元格的值
# print(xlsx_sheet.cell(1, 1).value)


# 编写集成化源文件
def encode_app_multi_lang_str_c(file, xlsx_sheet, sheet_row, sheet_col):
    # 写点简要的说明
    file.write('/*一个代替注册的,脚本自动生成的,本地静态的字符串表\n')
    file.write(' *通过app_multi_lang_str.py生成\n')
    file.write(' *参考app_multi_lang_str.json中的模式生成源\n */\n\n')
    file.write('#include "app_ext_lib.h"\n')
    file.write('#include "app_multi_lang_str.h"\n\n')
    # 提取所有外源依赖
    file.write('static const char * app_multi_lang_str_table[%d][%d] = {\n' % (sheet_row, sheet_col))
    for item in xlsx_sheet.rows:
        file.write('\t{')
        for data in item:
            file.write('{0}'.format("\"%s\"," % data.value))    # 多国语这里不能很好的对齐,因为不同语言的空格宽度不一致
        file.write('},\n')
    file.write('};\n\n')
    file.write('static uint32_t app_multi_lang_str_index = 0;\n\n')
    # 编写固化访问函数
    file.write('/*@brief设置搜索语言\n')
    file.write(' *@param[in] index语言编号(0~n-1)\n */\n')
    file.write('void app_multi_lang_str_default(uint32_t index)\n')
    file.write('{\n\tapp_multi_lang_str_index = index;\n}\n\n')
    file.write('/*@brief获得多国语字符串\n')
    file.write(' *@param[in] index字符串编号(0~n-1)\n */\n')
    file.write('const char * app_multi_lang_str_get(uint32_t index)\n')
    file.write('{\n\treturn app_multi_lang_str_table[index][app_multi_lang_str_index];\n}\n')


# 编写集成化头文件
def encode_app_multi_lang_str_h(file, xlsx_sheet, sheet_row, sheet_col):
    file.write('#ifndef APP_MULTI_LANG_STR_H\n')
    file.write('#define APP_MULTI_LANG_STR_H\n\n')
    # 写点简要的说明
    file.write('/*一个代替注册的,脚本自动生成的,本地静态的字符串表\n')
    file.write(' *通过app_multi_lang_str.py生成\n')
    file.write(' *参考app_multi_lang_str.json中的模式生成源\n */\n\n')
    # 编写头部索引
    file.write('typedef enum {\n')
    for i, item in enumerate(xlsx_sheet.rows):
        file.write('\tAPP_MULTI_LANG_STR_0X%04x,\t/* %s */\n' % (i, item[0].value))
    file.write('\tAPP_MULTI_LANG_STR_NUM,\n')
    file.write('} app_multi_lang_str_t;\n\n')
    # 编写固化访问函数接口
    file.write('/*@brief设置搜索语言\n')
    file.write(' *@param[in] index语言编号(0~n-1)\n */\n')
    file.write('void app_multi_lang_str_default(uint32_t index);\n\n')
    file.write('/*@brief获得多国语字符串\n')
    file.write(' *@param[in] index字符串编号(0~n-1)\n */\n')
    file.write('const char * app_multi_lang_str_get(uint32_t index);\n\n')
    file.write('#endif\n')


# 启用集成化资源总表生成
def encode_app_multi_lang_str_table():
    # json转Python字符串并转标准字典
    json_file = open('app_multi_lang_str.json', 'r')
    json_dict = json.loads(json_file.read())
    json_file.close()
    # 仅支持批量化字符串表管理
    if json_dict['type'] != 'app_multi_lang_str':
        return
    # 从标准字典获取工作簿名字和单元
    xlsx_file = openpyxl.load_workbook(json_dict['xlsx'], read_only=True)
    xlsx_sheet = xlsx_file[json_dict['sheet']]
    sheet_row = xlsx_sheet.max_row
    sheet_col = xlsx_sheet.max_column
    # 开启三个文件
    file_encode = 'utf-16'
    app_multi_lang_str_h = open('app_multi_lang_str.h', mode='w', encoding=file_encode)
    app_multi_lang_str_c = open('app_multi_lang_str.c', mode='w', encoding=file_encode)
    # 解析
    encode_app_multi_lang_str_h(app_multi_lang_str_h, xlsx_sheet, sheet_row, sheet_col)
    encode_app_multi_lang_str_c(app_multi_lang_str_c, xlsx_sheet, sheet_row, sheet_col)
    # 关闭三个文件
    app_multi_lang_str_h.close()
    app_multi_lang_str_c.close()
    # 关闭工作簿
    xlsx_file.close()


if __name__ == '__main__':
    encode_app_multi_lang_str_table()
    input('\nscript generation completed\n')
